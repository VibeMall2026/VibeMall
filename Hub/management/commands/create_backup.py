"""
Django Management Command: Create Backup
Export all critical VibeMall data to Excel files and sync to Terabox
Usage:
    python manage.py create_backup --type manual
    python manage.py create_backup --type scheduled --frequency daily
"""

import os
import json
import logging
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from django.core.management.base import BaseCommand, CommandError
from django.contrib.auth.models import User
from django.utils import timezone
from django.conf import settings

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from Hub.models import (
    Order, OrderItem, Product, Payment, ReturnRequest,
    BackupConfiguration, BackupLog, TeraboxSettings
)
from Hub.backup_utils import send_backup_notification_email, upload_to_terabox

logger = logging.getLogger(__name__)


class Command(BaseCommand):
    help = 'Create backup of all VibeMall data and sync to Terabox'
    
    def add_arguments(self, parser):
        parser.add_argument(
            '--type',
            type=str,
            default='manual',
            choices=['manual', 'scheduled', 'on-demand'],
            help='Type of backup: manual, scheduled, or on-demand'
        )
        
        parser.add_argument(
            '--frequency',
            type=str,
            default='custom',
            choices=['daily', 'weekly', 'biweekly', 'monthly', 'custom'],
            help='Backup frequency for scheduled backups'
        )
        
        parser.add_argument(
            '--output-dir',
            type=str,
            default=None,
            help='Custom output directory for backup files'
        )
        
        parser.add_argument(
            '--no-cloud',
            action='store_true',
            help='Skip Terabox upload'
        )
        
        parser.add_argument(
            '--no-email',
            action='store_true',
            help='Skip email notification'
        )
    
    def handle(self, *args, **options):
        """Main handler for backup creation."""
        try:
            backup_type = options['type']
            frequency = options['frequency']
            output_dir = options['output_dir']
            skip_cloud = options['no_cloud']
            skip_email = options['no_email']
            
            # Determine output directory
            if not output_dir:
                output_dir = os.path.join(settings.BASE_DIR, 'backups', 'excel')
            
            # Create directory if it doesn't exist
            os.makedirs(output_dir, exist_ok=True)
            
            # Create BackupLog entry
            backup_log = BackupLog.objects.create(
                backup_type=backup_type.upper().replace('-', '_'),
                backup_frequency=frequency.upper(),
                status='IN_PROGRESS'
            )
            
            self.stdout.write(self.style.SUCCESS(f'Starting backup #{backup_log.id}...'))
            
            # Create backup files
            backup_files = self.create_backup_files(output_dir, backup_log)
            
            if backup_files:
                backup_log.end_time = timezone.now()
                backup_log.status = 'SUCCESS'
                backup_log.save()
                
                # Upload to Terabox if enabled
                if not skip_cloud:
                    self.terabox_sync(backup_log, backup_files, output_dir)
                
                # Send notification email
                if not skip_email:
                    self.send_notification(backup_log, backup_files)
                
                self.stdout.write(self.style.SUCCESS(
                    f'\n✅ Backup #{backup_log.id} completed successfully!\n'
                    f'Location: {output_dir}'
                ))
            else:
                backup_log.status = 'FAILED'
                backup_log.error_message = 'No backup files generated'
                backup_log.save()
                self.stdout.write(self.style.ERROR('Backup failed - no files generated'))
        
        except Exception as e:
            logger.error(f'Backup creation failed: {str(e)}', exc_info=True)
            if 'backup_log' in locals():
                backup_log.status = 'FAILED'
                backup_log.error_message = str(e)
                backup_log.error_trace = logger.formatException(e)
                backup_log.end_time = timezone.now()
                backup_log.save()
            self.stdout.write(self.style.ERROR(f'Error: {str(e)}'))
            raise CommandError(f'Backup failed: {str(e)}')
    
    def create_backup_files(self, output_dir, backup_log):
        """Create all backup files."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_files = {}
        
        try:
            # Users backup
            users_file = self.export_users(output_dir, timestamp, backup_log)
            if users_file:
                backup_files['users'] = users_file
            
            # Orders backup
            orders_file = self.export_orders(output_dir, timestamp, backup_log)
            if orders_file:
                backup_files['orders'] = orders_file
            
            # Payments backup
            payments_file = self.export_payments(output_dir, timestamp, backup_log)
            if payments_file:
                backup_files['payments'] = payments_file
            
            # Products backup
            products_file = self.export_products(output_dir, timestamp, backup_log)
            if products_file:
                backup_files['products'] = products_file
            
            # Returns backup
            returns_file = self.export_returns(output_dir, timestamp, backup_log)
            if returns_file:
                backup_files['returns'] = returns_file
            
            # Analytics backup
            analytics_file = self.export_analytics(output_dir, timestamp, backup_log)
            if analytics_file:
                backup_files['analytics'] = analytics_file
            
            return backup_files
        
        except Exception as e:
            logger.error(f'Error creating backup files: {str(e)}', exc_info=True)
            raise
    
    def export_users(self, output_dir, timestamp, backup_log):
        """Export all user data to Excel file."""
        try:
            users = User.objects.all().values(
                'id', 'username', 'email', 'first_name', 'last_name',
                'phone', 'date_joined', 'last_login', 'is_active'
            )
            
            df = pd.DataFrame(list(users))
            
            if df.empty:
                self.stdout.write(self.style.WARNING('No users to backup'))
                return None
            
            filename = f'users_backup_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            # Create Excel with styling
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Users', index=False)
                
                # Add summary sheet
                summary_df = pd.DataFrame({
                    'Metric': ['Total Users', 'Active Users', 'Inactive Users', 'Backup Date'],
                    'Value': [
                        len(df),
                        len(df[df['is_active'] == True]),
                        len(df[df['is_active'] == False]),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                })
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                # Format sheets
                self.format_worksheet(writer, 'Users')
                self.format_worksheet(writer, 'Summary')
            
            backup_log.users_count = len(df)
            backup_log.local_file_path = filepath
            
            self.stdout.write(self.style.SUCCESS(f'✓ Exported {len(df)} users'))
            return filepath
        
        except Exception as e:
            logger.error(f'Error exporting users: {str(e)}', exc_info=True)
            return None
    
    def export_orders(self, output_dir, timestamp, backup_log):
        """Export all orders with detailed information."""
        try:
            orders = Order.objects.select_related('user').all()
            
            if not orders.exists():
                self.stdout.write(self.style.WARNING('No orders to backup'))
                return None
            
            # Main orders sheet
            orders_data = []
            all_items_data = []
            
            for order in orders:
                orders_data.append({
                    'Order ID': order.id,
                    'User': order.user.username,
                    'Email': order.user.email,
                    'Total Amount': float(order.total_amount),
                    'Discount': float(order.discount) if order.discount else 0,
                    'Tax': float(order.tax) if order.tax else 0,
                    'Shipping': float(order.shipping_cost) if order.shipping_cost else 0,
                    'Payment Status': order.payment_status,
                    'Order Status': order.order_status,
                    'Created At': order.created_at,
                    'Delivery Date': order.delivery_date,
                    'Notes': order.notes or '',
                })
                
                # Order items
                for item in order.items.all():
                    all_items_data.append({
                        'Order ID': order.id,
                        'Product': item.product.name,
                        'Category': item.product.category.name if item.product.category else '',
                        'Quantity': item.quantity,
                        'Price': float(item.price),
                        'Discount': float(item.discount) if item.discount else 0,
                        'Total': float(item.total),
                    })
            
            orders_df = pd.DataFrame(orders_data)
            items_df = pd.DataFrame(all_items_data)
            
            filename = f'orders_backup_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                orders_df.to_excel(writer, sheet_name='Orders', index=False)
                items_df.to_excel(writer, sheet_name='Order Items', index=False)
                
                # Summary
                summary_df = pd.DataFrame({
                    'Metric': ['Total Orders', 'Total Revenue', 'Avg Order Value', 'Backup Date'],
                    'Value': [
                        len(orders_df),
                        float(orders_df['Total Amount'].sum()),
                        float(orders_df['Total Amount'].mean()),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                })
                summary_df.to_excel(writer, sheet_name='Summary', index=False)
                
                self.format_worksheet(writer, 'Orders')
                self.format_worksheet(writer, 'Order Items')
                self.format_worksheet(writer, 'Summary')
            
            backup_log.orders_count = len(orders_df)
            
            self.stdout.write(self.style.SUCCESS(f'✓ Exported {len(orders_df)} orders'))
            return filepath
        
        except Exception as e:
            logger.error(f'Error exporting orders: {str(e)}', exc_info=True)
            return None
    
    def export_payments(self, output_dir, timestamp, backup_log):
        """Export payment data and analytics."""
        try:
            from Hub.models import Payment
            
            payments = Payment.objects.select_related('order').all()
            
            if not payments.exists():
                self.stdout.write(self.style.WARNING('No payments to backup'))
                return None
            
            payments_data = []
            for payment in payments:
                payments_data.append({
                    'Payment ID': payment.razorpay_payment_id,
                    'Order ID': payment.order_id,
                    'Amount': float(payment.amount),
                    'Currency': payment.currency,
                    'Status': payment.payment_status,
                    'Method': payment.payment_method or 'N/A',
                    'Created At': payment.created_at,
                    'Updated At': payment.updated_at,
                })
            
            payments_df = pd.DataFrame(payments_data)
            
            filename = f'payments_backup_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                payments_df.to_excel(writer, sheet_name='Payments', index=False)
                
                # Summary by status
                status_summary = payments_df.groupby('Status').agg({
                    'Amount': ['count', 'sum', 'mean']
                }).round(2)
                
                status_summary_df = pd.DataFrame({
                    'Status': status_summary.index,
                    'Count': status_summary[('Amount', 'count')].values,
                    'Total': status_summary[('Amount', 'sum')].values,
                    'Average': status_summary[('Amount', 'mean')].values,
                })
                
                status_summary_df.to_excel(writer, sheet_name='Status Summary', index=False)
                
                # Overall summary
                overall_df = pd.DataFrame({
                    'Metric': ['Total Payments', 'Total Revenue', 'Avg Payment', 'Backup Date'],
                    'Value': [
                        len(payments_df),
                        float(payments_df['Amount'].sum()),
                        float(payments_df['Amount'].mean()),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                })
                overall_df.to_excel(writer, sheet_name='Summary', index=False)
                
                self.format_worksheet(writer, 'Payments')
                self.format_worksheet(writer, 'Status Summary')
                self.format_worksheet(writer, 'Summary')
            
            backup_log.payments_count = len(payments_df)
            
            self.stdout.write(self.style.SUCCESS(f'✓ Exported {len(payments_df)} payments'))
            return filepath
        
        except Exception as e:
            logger.error(f'Error exporting payments: {str(e)}', exc_info=True)
            return None
    
    def export_products(self, output_dir, timestamp, backup_log):
        """Export product inventory data."""
        try:
            products = Product.objects.select_related('category').all()
            
            if not products.exists():
                self.stdout.write(self.style.WARNING('No products to backup'))
                return None
            
            products_data = []
            for product in products:
                products_data.append({
                    'Product ID': product.id,
                    'Name': product.name,
                    'Category': product.category.name if product.category else '',
                    'SKU': product.sku or '',
                    'Price': float(product.price),
                    'Cost Price': float(product.cost_price) if product.cost_price else 0,
                    'Stock': product.stock,
                    'Sold': product.sold or 0,
                    'Rating': float(product.rating) if product.rating else 0,
                    'Is Active': product.is_active,
                    'Visibility': product.visibility_status,
                    'Created At': product.created_at,
                })
            
            products_df = pd.DataFrame(products_data)
            
            filename = f'products_backup_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                products_df.to_excel(writer, sheet_name='Products', index=False)
                
                # Low stock alert
                low_stock_df = products_df[products_df['Stock'] < 10].copy()
                if not low_stock_df.empty:
                    low_stock_df.to_excel(writer, sheet_name='Low Stock Alert', index=False)
                
                # Category summary
                category_summary = products_df.groupby('Category').agg({
                    'Product ID': 'count',
                    'Stock': 'sum',
                    'Sold': 'sum',
                    'Price': 'mean'
                }).round(2)
                
                category_summary_df = pd.DataFrame({
                    'Category': category_summary.index,
                    'Count': category_summary['Product ID'].values,
                    'Total Stock': category_summary['Stock'].values,
                    'Total Sold': category_summary['Sold'].values,
                    'Avg Price': category_summary['Price'].values,
                })
                category_summary_df.to_excel(writer, sheet_name='Category Summary', index=False)
                
                # Overall summary
                overall_df = pd.DataFrame({
                    'Metric': ['Total Products', 'Total Stock', 'Total Sold', 'Avg Price', 'Backup Date'],
                    'Value': [
                        len(products_df),
                        int(products_df['Stock'].sum()),
                        int(products_df['Sold'].sum()),
                        float(products_df['Price'].mean()),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                })
                overall_df.to_excel(writer, sheet_name='Summary', index=False)
                
                self.format_worksheet(writer, 'Products')
                if not low_stock_df.empty:
                    self.format_worksheet(writer, 'Low Stock Alert')
                self.format_worksheet(writer, 'Category Summary')
                self.format_worksheet(writer, 'Summary')
            
            backup_log.products_count = len(products_df)
            
            self.stdout.write(self.style.SUCCESS(f'✓ Exported {len(products_df)} products'))
            return filepath
        
        except Exception as e:
            logger.error(f'Error exporting products: {str(e)}', exc_info=True)
            return None
    
    def export_returns(self, output_dir, timestamp, backup_log):
        """Export return requests data."""
        try:
            from Hub.models import ReturnRequest
            
            returns = ReturnRequest.objects.select_related('order', 'user').all()
            
            if not returns.exists():
                self.stdout.write(self.style.WARNING('No return requests to backup'))
                return None
            
            returns_data = []
            for ret in returns:
                returns_data.append({
                    'Return ID': ret.id,
                    'Order ID': ret.order_id,
                    'User': ret.user.username,
                    'Email': ret.user.email,
                    'Reason': ret.reason,
                    'Status': ret.status,
                    'Refund Amount': float(ret.refund_amount) if ret.refund_amount else 0,
                    'Created At': ret.created_at,
                    'Updated At': ret.updated_at,
                })
            
            returns_df = pd.DataFrame(returns_data)
            
            filename = f'returns_backup_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                returns_df.to_excel(writer, sheet_name='Returns', index=False)
                
                # Status summary
                status_summary = returns_df.groupby('Status').agg({
                    'Return ID': 'count',
                    'Refund Amount': 'sum'
                }).round(2)
                
                status_summary_df = pd.DataFrame({
                    'Status': status_summary.index,
                    'Count': status_summary['Return ID'].values,
                    'Total Refunds': status_summary['Refund Amount'].values,
                })
                status_summary_df.to_excel(writer, sheet_name='Status Summary', index=False)
                
                # Overall summary
                overall_df = pd.DataFrame({
                    'Metric': ['Total Returns', 'Total Refunds', 'Avg Refund', 'Backup Date'],
                    'Value': [
                        len(returns_df),
                        float(returns_df['Refund Amount'].sum()),
                        float(returns_df['Refund Amount'].mean()),
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    ]
                })
                overall_df.to_excel(writer, sheet_name='Summary', index=False)
                
                self.format_worksheet(writer, 'Returns')
                self.format_worksheet(writer, 'Status Summary')
                self.format_worksheet(writer, 'Summary')
            
            backup_log.returns_count = len(returns_df)
            
            self.stdout.write(self.style.SUCCESS(f'✓ Exported {len(returns_df)} returns'))
            return filepath
        
        except Exception as e:
            logger.error(f'Error exporting returns: {str(e)}', exc_info=True)
            return None
    
    def export_analytics(self, output_dir, timestamp, backup_log):
        """Export financial analytics and business metrics."""
        try:
            orders = Order.objects.select_related('user').all()
            products = Product.objects.all()
            
            if not orders.exists():
                self.stdout.write(self.style.WARNING('No data for analytics backup'))
                return None
            
            # Calculate metrics
            total_revenue = sum(float(o.total_amount) for o in orders)
            total_orders = len(orders)
            
            # Order status breakdown
            status_breakdown = {}
            for order in orders:
                status = order.order_status
                if status not in status_breakdown:
                    status_breakdown[status] = 0
                status_breakdown[status] += 1
            
            # Payment status breakdown
            payment_breakdown = {}
            for order in orders:
                status = order.payment_status
                if status not in payment_breakdown:
                    payment_breakdown[status] = 0
                payment_breakdown[status] += 1
            
            # Top products
            items = OrderItem.objects.select_related('product').all()
            product_sales = {}
            for item in items:
                if item.product.id not in product_sales:
                    product_sales[item.product.id] = {'name': item.product.name, 'qty': 0, 'revenue': 0}
                product_sales[item.product.id]['qty'] += item.quantity
                product_sales[item.product.id]['revenue'] += float(item.total)
            
            top_products = sorted(product_sales.items(), key=lambda x: x[1]['revenue'], reverse=True)[:10]
            
            filename = f'analytics_backup_{timestamp}.xlsx'
            filepath = os.path.join(output_dir, filename)
            
            with Workbook() as wb:
                # Business Metrics
                ws = wb.active
                ws.title = 'Metrics'
                ws['A1'] = 'VibeMall Business Analytics'
                ws['A1'].font = Font(bold=True, size=14)
                
                metrics = [
                    ['Metric', 'Value'],
                    ['Total Revenue', f'₹{total_revenue:,.2f}'],
                    ['Total Orders', total_orders],
                    ['Avg Order Value', f'₹{total_revenue/total_orders:,.2f}' if total_orders else '₹0'],
                    ['Total Products', products.count()],
                    ['Backup Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ]
                
                for row_idx, row_data in enumerate(metrics, 1):
                    for col_idx, val in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 30
                
                # Order Status Breakdown
                ws = wb.create_sheet('Order Status')
                ws['A1'] = 'Order Status Breakdown'
                ws['A1'].font = Font(bold=True)
                for idx, (status, count) in enumerate(status_breakdown.items(), 2):
                    ws[f'A{idx}'] = status
                    ws[f'B{idx}'] = count
                
                # Payment Status Breakdown
                ws = wb.create_sheet('Payment Status')
                ws['A1'] = 'Payment Status Breakdown'
                ws['A1'].font = Font(bold=True)
                for idx, (status, count) in enumerate(payment_breakdown.items(), 2):
                    ws[f'A{idx}'] = status
                    ws[f'B{idx}'] = count
                
                # Top Products
                ws = wb.create_sheet('Top Products')
                ws['A1'] = 'Product Name'
                ws['B1'] = 'Quantity Sold'
                ws['C1'] = 'Revenue'
                for idx, (prod_id, data) in enumerate(top_products, 2):
                    ws[f'A{idx}'] = data['name']
                    ws[f'B{idx}'] = data['qty']
                    ws[f'C{idx}'] = data['revenue']
                
                ws.column_dimensions['A'].width = 30
                ws.column_dimensions['B'].width = 15
                ws.column_dimensions['C'].width = 15
                
                wb.save(filepath)
            
            self.stdout.write(self.style.SUCCESS('✓ Exported analytics'))
            return filepath
        
        except Exception as e:
            logger.error(f'Error exporting analytics: {str(e)}', exc_info=True)
            return None
    
    def format_worksheet(self, writer, sheet_name):
        """Format worksheet with colors and borders."""
        worksheet = writer.sheets[sheet_name]
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
        
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.border = border
    
    def terabox_sync(self, backup_log, backup_files, output_dir):
        """Upload backup files to Terabox."""
        try:
            terabox_settings = TeraboxSettings.objects.first()
            if not terabox_settings or not terabox_settings.is_connected:
                self.stdout.write(self.style.WARNING('⚠ Terabox not connected, skipping cloud sync'))
                return
            
            self.stdout.write('Syncing to Terabox...')
            
            for file_type, filepath in backup_files.items():
                filename = os.path.basename(filepath)
                
                # Upload file
                success = upload_to_terabox(terabox_settings, filepath, filename)
                
                if success:
                    self.stdout.write(self.style.SUCCESS(f'✓ Uploaded {file_type} to Terabox'))
                    backup_log.terabox_synced = True
                else:
                    self.stdout.write(self.style.WARNING(f'⚠ Failed to upload {file_type}'))
            
            backup_log.save()
        
        except Exception as e:
            logger.error(f'Terabox sync failed: {str(e)}', exc_info=True)
            self.stdout.write(self.style.WARNING(f'⚠ Cloud sync failed: {str(e)}'))
    
    def send_notification(self, backup_log, backup_files):
        """Send backup notification email."""
        try:
            config = BackupConfiguration.objects.first()
            if not config or not config.notification_emails:
                return
            
            emails = config.get_notification_emails()
            if not emails:
                return
            
            send_backup_notification_email(backup_log, backup_files, emails)
            backup_log.email_sent = True
            backup_log.save()
            
            self.stdout.write(self.style.SUCCESS(f'✓ Notification sent to {len(emails)} recipients'))
        
        except Exception as e:
            logger.error(f'Email notification failed: {str(e)}', exc_info=True)
