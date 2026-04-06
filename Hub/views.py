from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages 
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import JsonResponse, HttpResponse, HttpRequest
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.cache import never_cache
from django.db.models import Count, Q, Avg, Sum, F, DecimalField, ExpressionWrapper, Case, When, Value, IntegerField, QuerySet, Min, Max, Prefetch
from django.db.models.functions import Coalesce, Lower, Trim
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger, Page
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone
from django.utils.dateparse import parse_datetime, parse_date
from django.utils.html import strip_tags
from django.conf import settings
from django.urls import reverse
from django.core.mail import send_mail, EmailMultiAlternatives
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.template.loader import render_to_string
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator
from django.conf import settings
from django.core.cache import cache
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.files import File
from decimal import Decimal, InvalidOperation
from datetime import timedelta, datetime, time
from urllib.parse import urlencode, urlparse
from typing import Dict, List, Optional, Tuple, Any, Union, Callable
import ipaddress
import re
import json
import base64
import os
import uuid
import requests
import shutil
from io import BytesIO
from PIL import Image as PILImage, UnidentifiedImageError
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError
import logging

logger = logging.getLogger(__name__)

RETURN_WINDOW_DAYS = 7
CANCEL_WINDOW_HOURS = 2
MAX_RETURN_ATTEMPTS = 1
NON_RETURNABLE_CATEGORIES = set()

RETURN_STATUS_FLOW = {
    'REQUESTED': ['APPROVED', 'REJECTED', 'CANCELLED'],
    'APPROVED': ['PICKUP_SCHEDULED'],
    'PICKUP_SCHEDULED': ['RECEIVED', 'UNABLE_TO_REACH'],
    'UNABLE_TO_REACH': ['RESCHEDULED'],
    'RESCHEDULED': ['PICKUP_SCHEDULED'],
    'RECEIVED': ['QC_PENDING'],
    'QC_PENDING': ['REFUND_PENDING', 'WRONG_RETURN', 'REPLACED'],
    'REFUND_PENDING': ['REFUNDED', 'WRONG_RETURN', 'REPLACED'],
    'WRONG_RETURN': ['REFUNDED'],
    'QC_PASSED': ['REFUND_PENDING', 'WRONG_RETURN', 'REFUNDED', 'REPLACED'],
    'QC_FAILED': ['REFUND_PENDING', 'WRONG_RETURN', 'REFUNDED', 'REPLACED'],
}

RTO_STATUS_FLOW = {
    'DELIVERY_FAILED': ['RTO_INITIATED'],
    'RTO_INITIATED': ['RTO_IN_TRANSIT', 'RTO_RECEIVED'],
    'RTO_IN_TRANSIT': ['RTO_RECEIVED'],
    'RTO_RECEIVED': ['RTO_CLOSED'],
}


def _ops_risk_level(return_count: int, rto_count: int, delivered_count: int) -> Dict[str, Any]:
    total_events = return_count + rto_count
    completed_orders = max(delivered_count, 1)
    issue_rate = total_events / completed_orders

    if rto_count >= 3 or total_events >= 6 or issue_rate >= 0.45:
        return {'level': 'high', 'label': 'High Risk', 'color': 'danger'}
    if rto_count >= 1 or total_events >= 3 or issue_rate >= 0.20:
        return {'level': 'medium', 'label': 'Medium Risk', 'color': 'warning'}
    return {'level': 'low', 'label': 'Low Risk', 'color': 'success'}


def _build_customer_ops_profiles(user_ids: List[int]) -> Dict[int, Dict[str, Any]]:
    clean_ids = [user_id for user_id in set(user_ids) if user_id]
    if not clean_ids:
        return {}

    return_counts = {
        row['user_id']: row['total']
        for row in ReturnRequest.objects.filter(user_id__in=clean_ids)
        .values('user_id')
        .annotate(total=Count('id'))
    }
    rto_counts = {
        row['order__user_id']: row['total']
        for row in RTOCase.objects.filter(order__user_id__in=clean_ids)
        .values('order__user_id')
        .annotate(total=Count('id'))
    }
    delivered_counts = {
        row['user_id']: row['total']
        for row in Order.objects.filter(user_id__in=clean_ids, order_status='DELIVERED')
        .values('user_id')
        .annotate(total=Count('id'))
    }

    profiles: Dict[int, Dict[str, Any]] = {}
    for user_id in clean_ids:
        return_count = int(return_counts.get(user_id, 0) or 0)
        rto_count = int(rto_counts.get(user_id, 0) or 0)
        delivered_count = int(delivered_counts.get(user_id, 0) or 0)
        risk_meta = _ops_risk_level(return_count, rto_count, delivered_count)
        profiles[user_id] = {
            'return_count': return_count,
            'rto_count': rto_count,
            'delivered_count': delivered_count,
            'issue_rate': round(((return_count + rto_count) / max(delivered_count, 1)) * 100, 1),
            **risk_meta,
        }
    return profiles


def _return_ops_recommendation(status: str) -> str:
    mapping = {
        'REQUESTED': 'Review the request, verify reason, and move quickly to approve or reject.',
        'APPROVED': 'Schedule reverse pickup and lock courier ownership.',
        'PICKUP_SCHEDULED': 'Track pickup SLA and monitor first-attempt completion.',
        'UNABLE_TO_REACH': 'Reconnect with the customer and reschedule pickup.',
        'RESCHEDULED': 'Watch the new pickup slot and avoid repeated misses.',
        'RECEIVED': 'Move the item into QC immediately.',
        'QC_PENDING': 'Complete QC and decide refund, replacement, or wrong-return handling.',
        'QC_PASSED': 'Route to refund/replacement and confirm inventory disposition.',
        'QC_FAILED': 'Hold item, document findings, and review wrong-return or partial-refund options.',
        'REFUND_PENDING': 'Process refund and close the case.',
        'WRONG_RETURN': 'Document mismatch evidence and confirm adjusted refund handling.',
    }
    return mapping.get(status, 'Monitor the case and move it to closure.')


def _rto_ops_recommendation(status: str) -> str:
    mapping = {
        'DELIVERY_FAILED': 'Confirm the failed-delivery reason and start the RTO workflow.',
        'RTO_INITIATED': 'Track courier handoff and keep customer risk tagged.',
        'RTO_IN_TRANSIT': 'Monitor return scans and expected warehouse arrival.',
        'RTO_RECEIVED': 'Run intake QC and decide restock, hold, reship, or disposal.',
        'RTO_CLOSED': 'Ensure final disposition and notes are complete.',
    }
    return mapping.get(status, 'Review the case and assign the next operational action.')


def _log_rto_history(rto_case: 'RTOCase', old_status: str, new_status: str, user: User, notes: str = '') -> None:
    RTOHistory.objects.create(
        rto_case=rto_case,
        old_status=old_status,
        new_status=new_status,
        changed_by=user,
        notes=notes,
    )

from .models import CategoryIcon, SubCategory, Slider, Feature, Banner, Product, DealCountdown, UserProfile, Address, Cart, Wishlist, ProductImage, ProductReview, ReviewImage, ReviewVote, ProductQuestion, Order, OrderItem, OrderStatusHistory, OrderCancellationRequest, ReturnRequest, ReturnItem, ReturnHistory, ReturnAttachment, ReturnLabel, RTOCase, RTOHistory, AdminEmailSettings, ProductStockNotification, BrandPartner, SiteSettings, LoyaltyPoints, PointsTransaction, MainPageProduct, MainPageSubCategoryBanner, MainPageBanner, ChatThread, ChatMessage, ChatAttachment, NewsletterSubscription
from .models_content_management import FAQCategory, FAQ
from .email_utils import send_order_confirmation_email, send_order_status_update_email, send_admin_order_notification, build_invoice_context
from .view_helpers import (
    _split_full_name,
    _get_checkout_items,
    _get_checkout_total_quantity,
    _get_resell_link_matching_quantity,
    _get_checkout_min_unit_price,
    _verify_upi_with_razorpay,
    _validate_indian_pincode,
    _lookup_ifsc_details,
)

# ===== ADMIN PANEL VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_test(request):
    """Admin Test Page"""
    return render(request, 'admin_panel/test.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
@never_cache
def admin_dashboard(request: HttpRequest) -> HttpResponse:
    """Admin Dashboard with comprehensive e-commerce statistics"""
    from datetime import timedelta
    from django.db.models.functions import TruncDate, TruncMonth, ExtractHour
    import calendar
    
    # Get current date for filtering
    today = timezone.now().date()
    last_7_days = today - timedelta(days=7)
    last_30_days = today - timedelta(days=30)
    current_month_start = today.replace(day=1)
    
    # Basic Statistics
    total_products = Product.objects.count()
    active_products = Product.objects.filter(is_active=True).count()
    total_users = User.objects.count()
    total_orders = Order.objects.count()

    # Paid orders (for consistent revenue calculations)
    paid_orders_qs = Order.objects.filter(payment_status='PAID')
    profit_orders_qs = _paid_orders_qs()
    
    # Revenue Statistics
    total_revenue = paid_orders_qs.aggregate(total=Sum('total_amount'))['total'] or 0
    
    monthly_revenue = paid_orders_qs.filter(
        created_at__gte=current_month_start
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    weekly_revenue = paid_orders_qs.filter(
        created_at__gte=last_7_days
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Order Statistics
    pending_orders = Order.objects.filter(order_status='PENDING').count()
    processing_orders = Order.objects.filter(order_status='PROCESSING').count()
    shipped_orders = Order.objects.filter(order_status='SHIPPED').count()
    delivered_orders = Order.objects.filter(order_status='DELIVERED').count()
    
    # Recent Orders
    recent_orders = Order.objects.all().select_related('user').order_by('-created_at')[:10]

    # Recent Order Items for dashboard table (show first 7 recent orders with their items)
    recent_order_items = Order.objects.all().select_related('user').prefetch_related('items').order_by('-created_at')[:7]
    
    # Top Selling Products (by total quantity sold)
    top_products = (
        Product.objects
        .annotate(sales_count=Sum('orderitem__quantity'))
        .filter(sales_count__gt=0)
        .order_by('-sales_count')[:5]
    )
    
    # Low Stock Products
    low_stock_products = Product.objects.filter(stock__lte=10, is_active=True).order_by('stock')[:5]
    
    # Recent Reviews
    recent_reviews = ProductReview.objects.select_related('user', 'product').order_by('-created_at')[:5]
    
    # Daily Sales Chart Data (Last 7 Days)
    daily_sales = paid_orders_qs.filter(
        created_at__gte=last_7_days
    ).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(
        total=Sum('total_amount')
    ).order_by('date')

    # Visitors (new users) in last 7 days
    new_users_last_7 = User.objects.filter(date_joined__date__gte=last_7_days).count()
    prev_7_start = last_7_days - timedelta(days=7)
    new_users_prev_7 = User.objects.filter(
        date_joined__date__gte=prev_7_start,
        date_joined__date__lt=last_7_days
    ).count()
    visitors_weekly_percent = round((new_users_last_7 / new_users_prev_7) * 100, 1) if new_users_prev_7 else (100.0 if new_users_last_7 else 0.0)

    visitors_series_qs = User.objects.filter(date_joined__date__gte=last_7_days).annotate(
        date=TruncDate('date_joined')
    ).values('date').annotate(count=Count('id')).order_by('date')
    visitors_series_map = {row['date']: row['count'] for row in visitors_series_qs}
    visitors_week_labels = [today - timedelta(days=i) for i in range(6, -1, -1)]
    visitors_weekly_series = [visitors_series_map.get(day, 0) for day in visitors_week_labels]

    # Visitors (today, yesterday, last 30 days)
    yesterday = today - timedelta(days=1)

    visitors_today_qs = User.objects.filter(date_joined__date=today).annotate(
        hour=ExtractHour('date_joined')
    ).values('hour').annotate(count=Count('id')).order_by('hour')
    visitors_today_map = {row['hour']: row['count'] for row in visitors_today_qs}
    visitors_today_series = [visitors_today_map.get(h, 0) for h in range(24)]
    visitors_today_count = sum(visitors_today_series)

    visitors_yesterday_qs = User.objects.filter(date_joined__date=yesterday).annotate(
        hour=ExtractHour('date_joined')
    ).values('hour').annotate(count=Count('id')).order_by('hour')
    visitors_yesterday_map = {row['hour']: row['count'] for row in visitors_yesterday_qs}
    visitors_yesterday_series = [visitors_yesterday_map.get(h, 0) for h in range(24)]
    visitors_yesterday_count = sum(visitors_yesterday_series)

    last_30_start = today - timedelta(days=29)
    visitors_last_month_qs = User.objects.filter(date_joined__date__gte=last_30_start).annotate(
        date=TruncDate('date_joined')
    ).values('date').annotate(count=Count('id')).order_by('date')
    visitors_last_month_map = {row['date']: row['count'] for row in visitors_last_month_qs}
    visitors_last_month_labels = [last_30_start + timedelta(days=i) for i in range(30)]
    visitors_last_month_series = [visitors_last_month_map.get(day, 0) for day in visitors_last_month_labels]
    visitors_last_month_count = sum(visitors_last_month_series)

    # Activity (orders) last 7 days vs previous 7 days
    activity_last_7 = Order.objects.filter(created_at__date__gte=last_7_days).count()
    activity_prev_7 = Order.objects.filter(
        created_at__date__gte=prev_7_start,
        created_at__date__lt=last_7_days
    ).count()
    activity_weekly_percent = round((activity_last_7 / activity_prev_7) * 100, 1) if activity_prev_7 else (100.0 if activity_last_7 else 0.0)

    activity_series_qs = Order.objects.filter(created_at__date__gte=last_7_days).annotate(
        date=TruncDate('created_at')
    ).values('date').annotate(count=Count('id')).order_by('date')
    activity_series_map = {row['date']: row['count'] for row in activity_series_qs}
    activity_weekly_series = [activity_series_map.get(day, 0) for day in visitors_week_labels]

    # Sales (ranges)
    last_30_revenue = paid_orders_qs.filter(created_at__date__gte=last_30_days).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_today = paid_orders_qs.filter(created_at__date=today).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_yesterday = paid_orders_qs.filter(created_at__date=today - timedelta(days=1)).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_last_month = last_30_revenue
    last_6_start = (current_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=30 * 5)
    sales_last_6_months = paid_orders_qs.filter(created_at__date__gte=last_6_start).aggregate(total=Sum('total_amount'))['total'] or 0
    prev_30_start = last_30_days - timedelta(days=30)
    prev_30_revenue = paid_orders_qs.filter(
        created_at__date__gte=prev_30_start,
        created_at__date__lt=last_30_days
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    sales_change_percent = round(((last_30_revenue - prev_30_revenue) / prev_30_revenue) * 100, 1) if prev_30_revenue else (100.0 if last_30_revenue else 0.0)

    # Profit & Expenses (margin-based)
    profit_items_last_30 = _order_item_profit(
        OrderItem.objects.filter(order__in=profit_orders_qs, order__created_at__date__gte=last_30_days)
    )
    return_fee_last_30 = _return_fee_profit(start_date=last_30_days)
    profit_last_30 = profit_items_last_30 + return_fee_last_30
    expenses_last_30 = max(last_30_revenue - profit_items_last_30, Decimal('0.00'))
    expenses_percent = round((expenses_last_30 / last_30_revenue) * 100, 1) if last_30_revenue else 0
    expenses_remaining_percent = max(0, 100 - expenses_percent)

    # Transactions (paid orders) & change
    transactions_last_30 = paid_orders_qs.filter(created_at__date__gte=last_30_days).count()
    transactions_prev_30 = paid_orders_qs.filter(
        created_at__date__gte=prev_30_start,
        created_at__date__lt=last_30_days
    ).count()
    transactions_change_percent = round(((transactions_last_30 - transactions_prev_30) / transactions_prev_30) * 100, 1) if transactions_prev_30 else (100.0 if transactions_last_30 else 0.0)

    # Yearly income/expense overview
    current_year_start = today.replace(month=1, day=1)
    income_by_month_qs = paid_orders_qs.filter(created_at__date__gte=current_year_start).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(total=Sum('total_amount')).order_by('month')
    income_month_map = {row['month'].month: row['total'] for row in income_by_month_qs if row['month']}
    month_labels = [calendar.month_abbr[i] for i in range(1, 13)]
    income_yearly_series = [float(income_month_map.get(i, 0) or 0) for i in range(1, 13)]

    profit_expr = ExpressionWrapper(
        Coalesce('margin_amount', 'product__margin', Value(0)) * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    profit_by_month_qs = (
        OrderItem.objects.filter(order__in=profit_orders_qs, order__created_at__date__gte=current_year_start)
        .annotate(month=TruncMonth('order__created_at'))
        .values('month')
        .annotate(total=Sum(profit_expr))
        .order_by('month')
    )
    profit_month_map = {row['month'].month: row['total'] for row in profit_by_month_qs if row['month']}

    return_fee_month_qs = (
        ReturnRequest.objects.filter(status='REFUNDED', resolved_at__date__gte=current_year_start)
        .annotate(month=TruncMonth('resolved_at'))
        .values('month')
        .annotate(total=Sum('refund_fee'))
        .order_by('month')
    )
    return_fee_month_map = {row['month'].month: row['total'] for row in return_fee_month_qs if row['month']}

    profit_yearly_series = [
        float((profit_month_map.get(i, 0) or 0) + (return_fee_month_map.get(i, 0) or 0))
        for i in range(1, 13)
    ]
    expense_yearly_series = [
        float(max((income_month_map.get(i, 0) or 0) - (profit_month_map.get(i, 0) or 0), 0))
        for i in range(1, 13)
    ]

    # Monthly report summary (current month vs previous)
    previous_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
    previous_month_end = current_month_start - timedelta(days=1)
    prev_month_revenue = paid_orders_qs.filter(
        created_at__date__gte=previous_month_start,
        created_at__date__lte=previous_month_end
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    report_income = monthly_revenue
    profit_items_month = _order_item_profit(
        OrderItem.objects.filter(order__in=profit_orders_qs, order__created_at__date__gte=current_month_start)
    )
    return_fee_month = _return_fee_profit(start_date=current_month_start)
    report_profit = profit_items_month + return_fee_month
    report_expense = max(monthly_revenue - profit_items_month, Decimal('0.00'))
    report_income_change = round(((monthly_revenue - prev_month_revenue) / prev_month_revenue) * 100, 1) if prev_month_revenue else (100.0 if monthly_revenue else 0.0)

    # Performance (last 6 months)
    last_6_months = [((current_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=30 * i)) for i in range(5, -1, -1)]
    perf_month_labels = [calendar.month_abbr[m.month] for m in last_6_months]
    perf_income_series = []
    perf_profit_series = []
    for m in last_6_months:
        next_m = (m + timedelta(days=32)).replace(day=1)
        total = paid_orders_qs.filter(created_at__date__gte=m, created_at__date__lt=next_m).aggregate(total=Sum('total_amount'))['total'] or 0
        perf_income_series.append(float(total))
        profit_items = _order_item_profit(
            OrderItem.objects.filter(order__in=profit_orders_qs, order__created_at__date__gte=m, order__created_at__date__lt=next_m)
        )
        return_fee = _return_fee_profit(start_date=m, end_date=next_m - timedelta(days=1))
        perf_profit_series.append(float(profit_items + return_fee))
    perf_sales_series = [round(val / 1000, 2) for val in perf_income_series]

    # Conversion funnel
    cart_count = Cart.objects.count()
    checkout_count = Order.objects.count()
    purchased_count = paid_orders_qs.count()
    impressions_count = max(cart_count * 8, total_users * 5, 1)
    conversion_rate = round((purchased_count / impressions_count) * 100, 2) if impressions_count else 0

    # Order statistics (weekly percentage)
    weekly_orders = Order.objects.filter(created_at__date__gte=last_7_days).count()
    order_statistics_percent = round((weekly_orders / total_orders) * 100, 1) if total_orders else 0

    # Finance tab series (last 7 months)
    finance_months = [((current_month_start - timedelta(days=1)).replace(day=1) - timedelta(days=30 * i)) for i in range(6, -1, -1)]
    finance_labels = [calendar.month_abbr[m.month] for m in finance_months]
    finance_income_series = []
    finance_expense_series = []
    finance_profit_series = []
    for m in finance_months:
        next_m = (m + timedelta(days=32)).replace(day=1)
        total = paid_orders_qs.filter(created_at__date__gte=m, created_at__date__lt=next_m).aggregate(total=Sum('total_amount'))['total'] or 0
        profit_items = _order_item_profit(
            OrderItem.objects.filter(order__in=profit_orders_qs, order__created_at__date__gte=m, order__created_at__date__lt=next_m)
        )
        return_fee = _return_fee_profit(start_date=m, end_date=next_m - timedelta(days=1))
        finance_profit = profit_items + return_fee
        finance_expense = max(total - profit_items, Decimal('0.00'))
        finance_income_series.append(float(total))
        finance_expense_series.append(float(finance_expense))
        finance_profit_series.append(float(finance_profit))

    weekly_expense = expenses_last_30 / 4 if expenses_last_30 else 0
    weekly_expense_percent = round((weekly_expense / expenses_last_30) * 100, 1) if expenses_last_30 else 0

    # Transactions list (recent orders)
    recent_transactions = []
    for order in recent_orders[:6]:
        recent_transactions.append({
            'method': order.payment_method,
            'label': order.user.get_full_name() or order.user.username,
            'amount': order.total_amount,
            'status': order.payment_status
        })

    # Top selling product (last 30 days)
    top_selling_item = (
        OrderItem.objects
        .filter(order__payment_status='PAID', order__created_at__date__gte=last_30_days)
        .values('product_id', 'product_name')
        .annotate(total_qty=Sum('quantity'), total_value=Sum('subtotal'))
        .order_by('-total_qty', '-total_value')
        .first()
    )
    top_selling_name = top_selling_item['product_name'] if top_selling_item else 'N/A'
    top_selling_sales = top_selling_item['total_value'] if top_selling_item else 0
    top_selling_qty = top_selling_item['total_qty'] if top_selling_item else 0
    top_selling_target_percent = min(round((top_selling_sales / (monthly_revenue or Decimal('1'))) * 100, 1), 100) if monthly_revenue else 0

    # Revenue growth vs previous month
    revenue_growth_percent = round(((monthly_revenue - prev_month_revenue) / prev_month_revenue) * 100, 1) if prev_month_revenue else (100.0 if monthly_revenue else 0.0)

    # Sales target
    sales_target = monthly_revenue * Decimal('1.2') if monthly_revenue else Decimal('10000')
    sales_target_percent = min(round((monthly_revenue / sales_target) * 100, 1), 100) if sales_target else 0
    
    # Recent Customers
    recent_customers = User.objects.filter(is_staff=False).order_by('-date_joined')[:5]
    
    # ===== RETURN METRICS =====
    total_returns = ReturnRequest.objects.count()
    pending_returns = ReturnRequest.objects.exclude(status__in=['REFUNDED', 'REPLACED', 'REJECTED']).count()
    refunded_returns = ReturnRequest.objects.filter(status='REFUNDED')
    total_refund_amount = refunded_returns.aggregate(
        total=Sum(Coalesce('refund_amount_net', 'refund_amount'))
    )['total'] or Decimal('0.00')
    
    returns_last_7 = ReturnRequest.objects.filter(requested_at__date__gte=last_7_days).count()
    total_delivered = Order.objects.filter(order_status='DELIVERED').count()
    return_rate = round((total_returns / total_delivered * 100), 2) if total_delivered > 0 else 0
    
    context = {
        # Basic Stats
        'total_products': total_products,
        'active_products': active_products,
        'total_users': total_users,
        'total_orders': total_orders,
        
        # Revenue Stats
        'total_revenue': total_revenue,
        'monthly_revenue': monthly_revenue,
        'weekly_revenue': weekly_revenue,
        
        # Order Stats
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        
        # Lists
        'recent_orders': recent_orders,
        'recent_order_items': recent_order_items,
        'top_products': top_products,
        'low_stock_products': low_stock_products,
        'recent_reviews': recent_reviews,
        'recent_customers': recent_customers,
        
        # Chart Data
        'daily_sales': list(daily_sales),

        # Dashboard UI data
        'top_selling_name': top_selling_name,
        'top_selling_sales': top_selling_sales,
        'top_selling_qty': top_selling_qty,
        'top_selling_target_percent': top_selling_target_percent,

        'visitors_weekly_percent': visitors_weekly_percent,
        'visitors_weekly_series': visitors_weekly_series,
        'visitors_today_series': visitors_today_series,
        'visitors_yesterday_series': visitors_yesterday_series,
        'visitors_last_month_series': visitors_last_month_series,
        'visitors_today_count': visitors_today_count,
        'visitors_yesterday_count': visitors_yesterday_count,
        'visitors_last_month_count': visitors_last_month_count,
        'activity_weekly_percent': activity_weekly_percent,
        'activity_weekly_series': activity_weekly_series,

        'last_30_revenue': last_30_revenue,
        'sales_today': sales_today,
        'sales_yesterday': sales_yesterday,
        'sales_last_month': sales_last_month,
        'sales_last_6_months': sales_last_6_months,
        'sales_change_percent': sales_change_percent,
        'profit_last_30': profit_last_30,
        'expenses_last_30': expenses_last_30,
        'expenses_percent': expenses_percent,
        'expenses_remaining_percent': expenses_remaining_percent,
        'transactions_last_30': transactions_last_30,
        'transactions_change_percent': transactions_change_percent,

        'month_labels': month_labels,
        'income_yearly_series': income_yearly_series,
        'expense_yearly_series': expense_yearly_series,
        'profit_yearly_series': profit_yearly_series,

        'report_income': report_income,
        'report_expense': report_expense,
        'report_profit': report_profit,
        'report_income_change': report_income_change,

        'perf_month_labels': perf_month_labels,
        'perf_income_series': perf_income_series,
        'perf_profit_series': perf_profit_series,
        'perf_sales_series': perf_sales_series,

        'impressions_count': impressions_count,
        'cart_count': cart_count,
        'checkout_count': checkout_count,
        'purchased_count': purchased_count,
        'conversion_rate': conversion_rate,

        'order_statistics_percent': order_statistics_percent,
        'weekly_orders': weekly_orders,
        'total_sales_count': purchased_count,

        'finance_labels': finance_labels,
        'finance_income_series': finance_income_series,
        'finance_expense_series': finance_expense_series,
        'finance_profit_series': finance_profit_series,
        'weekly_expense': weekly_expense,
        'weekly_expense_percent': weekly_expense_percent,

        'recent_transactions': recent_transactions,

        'revenue_growth_percent': revenue_growth_percent,
        'sales_target': sales_target,
        'sales_target_percent': sales_target_percent,
        
        # Return Metrics
        'total_returns': total_returns,
        'pending_returns': pending_returns,
        'total_refund_amount': total_refund_amount,
        'returns_last_7': returns_last_7,
        'return_rate': return_rate,
    }
    return render(request, 'admin_panel/dashboard.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_new_dashboard(request):
    """New admin analytics dashboard (read-only)"""
    from datetime import timedelta
    from django.db.models.functions import TruncDate
    import json

    today = timezone.localdate()
    range_key = request.GET.get('range', 'this_week')
    range_key = range_key if range_key in {'today', 'yesterday', 'this_week', 'last_week', 'this_month', 'this_year'} else 'this_week'

    if range_key == 'today':
        start_date = today
        end_date = today
        range_label = 'Today'
    elif range_key == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
        range_label = 'Yesterday'
    elif range_key == 'last_week':
        start_date = today - timedelta(days=13)
        end_date = today - timedelta(days=7)
        range_label = 'Last Week'
    elif range_key == 'this_month':
        start_date = today.replace(day=1)
        end_date = today
        range_label = 'This Month'
    elif range_key == 'this_year':
        start_date = today.replace(month=1, day=1)
        end_date = today
        range_label = 'This Year'
    else:
        start_date = today - timedelta(days=6)
        end_date = today
        range_label = 'This Week'

    paid_orders = Order.objects.filter(payment_status='PAID')
    profit_orders_qs = _paid_orders_qs()
    profit_orders_qs = _paid_orders_qs()
    profit_orders_qs = _paid_orders_qs()

    revenue_qs = paid_orders.filter(created_at__date__gte=start_date, created_at__date__lte=end_date)
    revenue_by_day = revenue_qs.annotate(day=TruncDate('created_at')).values('day').annotate(total=Sum('total_amount'))
    orders_by_day = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).annotate(day=TruncDate('created_at')).values('day').annotate(total=Count('id'))
    transactions_by_day = revenue_qs.annotate(day=TruncDate('created_at')).values('day').annotate(total=Count('id'))
    visitors_by_day = User.objects.filter(date_joined__date__gte=start_date, date_joined__date__lte=end_date).annotate(day=TruncDate('date_joined')).values('day').annotate(total=Count('id'))
    customers_by_day = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).annotate(day=TruncDate('created_at')).values('day').annotate(total=Count('user', distinct=True))

    revenue_map = {item['day']: item['total'] or Decimal('0') for item in revenue_by_day}
    orders_map = {item['day']: item['total'] or 0 for item in orders_by_day}
    transactions_map = {item['day']: item['total'] or 0 for item in transactions_by_day}
    visitors_map = {item['day']: item['total'] or 0 for item in visitors_by_day}
    customers_map = {item['day']: item['total'] or 0 for item in customers_by_day}

    labels = []
    revenue_series = []
    orders_series = []
    profit_series = []
    expense_series = []
    visitors_series = []
    customers_series = []
    transactions_series = []

    profit_expr = ExpressionWrapper(
        Coalesce('margin_amount', 'product__margin', Value(0)) * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    profit_by_day_qs = (
        OrderItem.objects.filter(order__in=profit_orders_qs, order__created_at__date__gte=start_date, order__created_at__date__lte=end_date)
        .annotate(day=TruncDate('order__created_at'))
        .values('day')
        .annotate(total=Sum(profit_expr))
    )
    profit_day_map = {item['day']: item['total'] or Decimal('0.00') for item in profit_by_day_qs}

    return_fee_by_day_qs = (
        ReturnRequest.objects.filter(status='REFUNDED', resolved_at__date__gte=start_date, resolved_at__date__lte=end_date)
        .annotate(day=TruncDate('resolved_at'))
        .values('day')
        .annotate(total=Sum('refund_fee'))
    )
    return_fee_day_map = {item['day']: item['total'] or Decimal('0.00') for item in return_fee_by_day_qs}

    day_count = (end_date - start_date).days + 1
    for i in range(day_count):
        day = start_date + timedelta(days=i)
        labels.append(day.strftime('%b %d'))
        day_revenue = revenue_map.get(day, Decimal('0'))
        revenue_series.append(float(day_revenue))
        orders_series.append(int(orders_map.get(day, 0)))
        day_profit_items = profit_day_map.get(day, Decimal('0.00'))
        day_return_fee = return_fee_day_map.get(day, Decimal('0.00'))
        profit_series.append(float(day_profit_items + day_return_fee))
        expense_series.append(float(max(day_revenue - day_profit_items, Decimal('0.00'))))
        visitors_series.append(int(visitors_map.get(day, 0)))
        customers_series.append(int(customers_map.get(day, 0)))
        transactions_series.append(int(transactions_map.get(day, 0)))

    total_revenue = sum(revenue_series)
    total_profit = sum(profit_series)
    total_expenses = sum(expense_series)
    total_orders = sum(orders_series)
    total_visitors = sum(visitors_series)
    total_customers = Order.objects.filter(created_at__date__gte=start_date, created_at__date__lte=end_date).values('user').distinct().count()
    total_transactions = sum(transactions_series)

    cards = [
        {'key': 'orders', 'title': 'Orders', 'value': total_orders, 'is_money': False, 'series': json.dumps(orders_series), 'color': '#2563eb', 'chart_id': 'ordersChart'},
        {'key': 'expenses', 'title': 'Expenses', 'value': total_expenses, 'is_money': True, 'series': json.dumps(expense_series), 'color': '#dc2626', 'chart_id': 'expensesChart'},
        {'key': 'profit', 'title': 'Profit', 'value': total_profit, 'is_money': True, 'series': json.dumps(profit_series), 'color': '#16a34a', 'chart_id': 'profitChart'},
        {'key': 'sales', 'title': 'Sales', 'value': total_revenue, 'is_money': True, 'series': json.dumps(revenue_series), 'color': '#0ea5e9', 'chart_id': 'salesChart'},
        {'key': 'visitors', 'title': 'Visitors', 'value': total_visitors, 'is_money': False, 'series': json.dumps(visitors_series), 'color': '#8b5cf6', 'chart_id': 'visitorsChart'},
        {'key': 'customers', 'title': 'Customers', 'value': total_customers, 'is_money': False, 'series': json.dumps(customers_series), 'color': '#f59e0b', 'chart_id': 'customersChart'},
        {'key': 'transactions', 'title': 'Transactions', 'value': total_transactions, 'is_money': False, 'series': json.dumps(transactions_series), 'color': '#14b8a6', 'chart_id': 'transactionsChart'},
    ]

    excluded_categories = {'TOP_DEALS', 'TOP_SELLING', 'TOP_FEATURED', 'RECOMMENDED'}
    category_rows = (
        OrderItem.objects.filter(
            order__created_at__date__gte=start_date,
            order__created_at__date__lte=end_date
        )
        .values('product__category')
        .annotate(total=Sum('quantity'))
    )
    category_totals_map = {row['product__category']: int(row['total'] or 0) for row in category_rows}

    category_colors = [
        '#60a5fa',
        '#f59e0b',
        '#fb7185',
        '#22c55e',
        '#a78bfa',
        '#38bdf8',
        '#f97316',
        '#84cc16',
        '#c084fc',
        '#f472b6',
        '#2dd4bf',
        '#94a3b8',
    ]

    category_labels = []
    category_counts = []
    category_list = []
    icon_categories = CategoryIcon.objects.filter(is_active=True).order_by('order', 'id')

    for idx, icon in enumerate(icon_categories):
        category_key = icon.category_key
        if category_key in excluded_categories:
            continue
        total = category_totals_map.get(category_key, 0)
        color = category_colors[idx % len(category_colors)]
        category_labels.append(icon.name)
        category_counts.append(total)
        category_list.append({'label': icon.name, 'total': total, 'color': color})

    icon_keys = {icon.category_key for icon in icon_categories}
    other_total = 0
    for category_key, total in category_totals_map.items():
        if category_key in icon_keys or category_key in excluded_categories:
            continue
        other_total += total
    if other_total:
        color = category_colors[len(category_labels) % len(category_colors)]
        category_labels.append('Other')
        category_counts.append(other_total)
        category_list.append({'label': 'Other', 'total': other_total, 'color': color})

    context = {
        'range_key': range_key,
        'range_label': range_label,
        'chart_labels': json.dumps(labels),
        'cards': cards,
        'category_labels': json.dumps(category_labels),
        'category_counts': json.dumps(category_counts),
        'category_list': category_list,
        'category_colors': json.dumps(category_colors[:len(category_labels)]),
    }

    return render(request, 'admin_panel/new_dashboard.html', context)


def _ensure_session_key(request):
    if not request.session.session_key:
        request.session.save()


def _get_admin_chat_email():
    admin_settings = AdminEmailSettings.objects.filter(is_active=True).first()
    return admin_settings.admin_email if admin_settings else 'info.vibemall@gmail.com'


def _get_from_email():
    return (
        getattr(settings, 'DEFAULT_FROM_EMAIL', '').strip() or
        getattr(settings, 'EMAIL_HOST_USER', '').strip() or
        'info.vibemall@gmail.com'
    )


def _get_thread_for_request(request, thread_id=None):
    thread = None
    if thread_id:
        try:
            thread = ChatThread.objects.get(id=thread_id)
        except ChatThread.DoesNotExist:
            thread = None

        if thread:
            if request.user.is_authenticated:
                if thread.user_id != request.user.id:
                    thread = None
            else:
                _ensure_session_key(request)
                if thread.session_key != request.session.session_key:
                    thread = None

    if not thread:
        if request.user.is_authenticated:
            thread = ChatThread.objects.filter(user=request.user, status='OPEN').order_by('-last_message_at', '-created_at').first()
        else:
            _ensure_session_key(request)
            thread = ChatThread.objects.filter(session_key=request.session.session_key, status='OPEN').order_by('-last_message_at', '-created_at').first()

    if not thread:
        thread = ChatThread.objects.create(
            user=request.user if request.user.is_authenticated else None,
            session_key=request.session.session_key if not request.user.is_authenticated else ''
        )

    return thread


def _send_admin_chat_email(thread, message_text):
    admin_email = _get_admin_chat_email()
    if not admin_email:
        return

    subject = f"New chat message - {thread.display_name()}"
    html_content = render_to_string('emails/chat_admin_notify.html', {
        'thread': thread,
        'message': message_text,
    })
    text_content = f"""New customer message

From: {thread.display_name()}
Email: {thread.user.email if thread.user else thread.guest_email}

Message:
{message_text}
"""

    email = EmailMultiAlternatives(subject, text_content, _get_from_email(), [admin_email])
    email.attach_alternative(html_content, "text/html")
    email.send(fail_silently=True)


def _send_user_chat_email(thread, message_text):
    recipient = None
    if thread.user and thread.user.email:
        recipient = thread.user.email
    elif thread.guest_email:
        recipient = thread.guest_email

    if not recipient:
        return

    subject = "Support reply from VibeMall"
    html_content = render_to_string('emails/chat_user_reply.html', {
        'thread': thread,
        'message': message_text,
    })
    text_content = f"""Support reply

Hello {thread.display_name()},

{message_text}
"""

    email = EmailMultiAlternatives(subject, text_content, _get_from_email(), [recipient])
    email.attach_alternative(html_content, "text/html")
    email.send(fail_silently=True)


def _validate_chat_attachments(files):
    allowed_ext = {'.jpg', '.jpeg', '.png', '.webp', '.mp4', '.pdf', '.doc', '.docx'}
    allowed_types = {
        'image/jpeg', 'image/png', 'image/webp',
        'video/mp4',
        'application/pdf',
        'application/msword',
        'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
    }
    max_size = 25 * 1024 * 1024

    for f in files:
        if f.size > max_size:
            return f"File {f.name} exceeds 25 MB."
        name_lower = (f.name or '').lower()
        if not any(name_lower.endswith(ext) for ext in allowed_ext):
            return f"File type not allowed: {f.name}."
        if f.content_type and f.content_type not in allowed_types:
            return f"File type not allowed: {f.name}."
    return None


def chat_thread(request):
    thread_id = request.GET.get('thread_id')
    thread = _get_thread_for_request(request, thread_id)

    messages_qs = thread.messages.prefetch_related('attachments').order_by('created_at')[:50]
    messages_data = [
        {
            'sender': msg.sender_type,
            'message': msg.message,
            'attachments': [
                {
                    'name': att.original_name,
                    'url': att.file.url,
                    'content_type': att.content_type,
                    'is_image': (att.content_type or '').startswith('image/')
                }
                for att in msg.attachments.all()
            ],
            'created_at': msg.created_at.strftime('%d %b %H:%M')
        }
        for msg in messages_qs
    ]

    requires_profile = False
    if not request.user.is_authenticated:
        requires_profile = not (thread.guest_name and thread.guest_email)

    return JsonResponse({
        'thread_id': thread.id,
        'messages': messages_data,
        'requires_profile': requires_profile,
        'guest_name': thread.guest_name,
        'guest_email': thread.guest_email,
    })


@require_POST
def chat_message(request):
    message_text = (request.POST.get('message') or '').strip()
    files = request.FILES.getlist('attachments')
    if not message_text and not files:
        return JsonResponse({'error': 'Message or attachment is required.'}, status=400)

    error = _validate_chat_attachments(files)
    if error:
        return JsonResponse({'error': error}, status=400)

    thread_id = request.POST.get('thread_id')
    thread = _get_thread_for_request(request, thread_id)

    if not request.user.is_authenticated:
        guest_name = (request.POST.get('guest_name') or '').strip()
        guest_email = (request.POST.get('guest_email') or '').strip()

        if not (thread.guest_name and thread.guest_email):
            if not guest_name or not guest_email:
                return JsonResponse({'error': 'Name and email are required.'}, status=400)

        if guest_name and not thread.guest_name:
            thread.guest_name = guest_name
        if guest_email and not thread.guest_email:
            thread.guest_email = guest_email

    message = ChatMessage.objects.create(
        thread=thread,
        sender_type='USER',
        message=message_text
    )
    for f in files:
        ChatAttachment.objects.create(
            message=message,
            file=f,
            original_name=f.name,
            content_type=f.content_type or '',
            size_bytes=f.size
        )
    thread.last_message_at = timezone.now()
    thread.save(update_fields=['last_message_at', 'guest_name', 'guest_email'])

    notify_text = message_text or "Sent attachment(s)."
    _send_admin_chat_email(thread, notify_text)

    return JsonResponse({
        'status': 'ok',
        'message': {
            'sender': message.sender_type,
            'message': message.message,
            'attachments': [
                {
                    'name': att.original_name,
                    'url': att.file.url,
                    'content_type': att.content_type,
                    'is_image': (att.content_type or '').startswith('image/')
                }
                for att in message.attachments.all()
            ],
            'created_at': message.created_at.strftime('%d %b %H:%M')
        }
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_chat_list(request):
    threads = ChatThread.objects.select_related('user').order_by('-last_message_at', '-created_at')
    return render(request, 'admin_panel/chat_list.html', {'threads': threads})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_chat_detail(request, thread_id):
    thread = get_object_or_404(ChatThread, id=thread_id)

    if request.method == 'POST':
        message_text = (request.POST.get('message') or '').strip()
        files = request.FILES.getlist('attachments')
        if not message_text and not files:
            messages.error(request, 'Message or attachment is required.')
        else:
            error = _validate_chat_attachments(files)
            if error:
                messages.error(request, error)
            else:
                message = ChatMessage.objects.create(
                    thread=thread,
                    sender_type='ADMIN',
                    message=message_text
                )
                for f in files:
                    ChatAttachment.objects.create(
                        message=message,
                        file=f,
                        original_name=f.name,
                        content_type=f.content_type or '',
                        size_bytes=f.size
                    )
                thread.last_message_at = timezone.now()
                thread.save(update_fields=['last_message_at'])
                notify_text = message_text or "Sent attachment(s)."
                _send_user_chat_email(thread, notify_text)
                messages.success(request, 'Reply sent.')
                return redirect('admin_chat_detail', thread_id=thread.id)

    messages_qs = thread.messages.prefetch_related('attachments').order_by('created_at')
    context = {
        'thread': thread,
        'messages_list': messages_qs,
    }
    return render(request, 'admin_panel/chat_detail.html', context)

def calc_discount_percent(current_price, original_price):
    try:
        original_value = Decimal(str(original_price))
    except Exception:
        return 0

    if original_value <= 0:
        return 0

    try:
        current_value = Decimal(str(current_price))
    except Exception:
        current_value = Decimal('0')

    if current_value >= original_value:
        return 0

    return int(((original_value - current_value) / original_value) * Decimal('100'))


def normalize_decimal_input(raw_value):
    if raw_value is None:
        return ''

    cleaned = re.sub(r'[^0-9,.-]', '', str(raw_value))

    if ',' in cleaned and '.' in cleaned:
        cleaned = cleaned.replace(',', '')
    else:
        cleaned = cleaned.replace(',', '.')

    return cleaned


def crop_image_height(file_obj, target_height=738):
    if not file_obj:
        return file_obj

    try:
        _register_heif_opener()
        file_obj.seek(0)
        image = PILImage.open(file_obj)
        image.load()
        width, height = image.size

        if height > target_height:
            image = image.crop((0, 0, width, target_height))

        image = image.convert('RGBA')
        output = BytesIO()
        image.save(output, format='PNG')
        output.seek(0)

        base_name, _ = os.path.splitext(file_obj.name)
        png_name = f"{base_name}.png"

        return InMemoryUploadedFile(
            output,
            'ImageField',
            png_name,
            'image/png',
            output.getbuffer().nbytes,
            None,
        )
    except Exception:
        file_obj.seek(0)
        return file_obj
def _register_heif_opener():
    try:
        import pillow_heif
        pillow_heif.register_heif_opener()
    except Exception:
        pass


def _safe_edit_photo_path_from_url(url):
    media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
    media_url = getattr(settings, 'MEDIA_URL', '/media/')
    if not url.startswith(media_url):
        return None, None
    rel_path = url[len(media_url):].lstrip('/')
    if not rel_path.startswith('edit_photo/'):
        return None, None
    abs_path = os.path.join(media_root, rel_path)
    if not os.path.isfile(abs_path):
        return None, None
    return abs_path, rel_path


def _crop_png_by_ratio(abs_path, ratio, index=None):
    _register_heif_opener()
    image = PILImage.open(abs_path)
    image.load()
    img_w, img_h = image.size
    rx, ry, rw, rh = ratio
    x = max(0, min(int(rx * img_w), img_w - 1))
    y = max(0, min(int(ry * img_h), img_h - 1))
    width = max(1, min(int(rw * img_w), img_w - x))
    height = max(1, min(int(rh * img_h), img_h - y))
    image = image.crop((x, y, x + width, y + height))
    image = image.convert('RGBA')

    if index is not None:
        filename = f"{index}.png"
    else:
        filename = f"edit_photo_crop_{uuid.uuid4().hex}.png"
    rel_path = os.path.join('edit_photo', filename)
    media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
    abs_out = os.path.join(media_root, rel_path)
    os.makedirs(os.path.dirname(abs_out), exist_ok=True)
    image.save(abs_out, format='PNG')
    public_url = f"{settings.MEDIA_URL}{rel_path.replace(os.sep, '/')}"
    return public_url, filename


def _resize_uploaded_image(file_obj, target_width=None, target_height=None, keep_aspect=True):
    _register_heif_opener()
    file_obj.seek(0)
    image = PILImage.open(file_obj)
    image.load()

    original_width, original_height = image.size

    if not target_width and not target_height:
        raise ValueError('Please provide width or height.')

    if keep_aspect:
        if target_width and target_height:
            scale = min(target_width / original_width, target_height / original_height)
            scale = max(scale, 1 / max(original_width, original_height))
            new_width = max(1, int(round(original_width * scale)))
            new_height = max(1, int(round(original_height * scale)))
        elif target_width:
            new_width = max(1, int(target_width))
            new_height = max(1, int(round((original_height * new_width) / original_width)))
        else:
            new_height = max(1, int(target_height))
            new_width = max(1, int(round((original_width * new_height) / original_height)))
    else:
        new_width = max(1, int(target_width or original_width))
        new_height = max(1, int(target_height or original_height))

    try:
        resample_filter = PILImage.Resampling.LANCZOS
    except AttributeError:
        resample_filter = PILImage.LANCZOS

    image = image.convert('RGBA').resize((new_width, new_height), resample_filter)

    filename = f"edit_photo_resize_{uuid.uuid4().hex}.png"
    rel_path = os.path.join('edit_photo', filename)
    media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
    abs_path = os.path.join(media_root, rel_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    image.save(abs_path, format='PNG')
    public_url = f"{settings.MEDIA_URL}{rel_path.replace(os.sep, '/')}"

    return {
        'url': public_url,
        'filename': filename,
        'original_width': original_width,
        'original_height': original_height,
        'new_width': new_width,
        'new_height': new_height,
    }


def _cleanup_orphan_orders() -> int:
    """Delete invalid orders that no longer have any order items."""
    orphan_order_ids = list(
        Order.objects.annotate(item_count=Count('items'))
        .filter(item_count=0)
        .values_list('id', flat=True)
    )
    if not orphan_order_ids:
        return 0

    Order.objects.filter(id__in=orphan_order_ids).delete()
    logger.warning("Cleaned up %s orphan orders with no items.", len(orphan_order_ids))
    return len(orphan_order_ids)


def convert_url_to_png(image_url, crop_box=None, crop_ratio=None):
    response = requests.get(image_url, timeout=15)
    response.raise_for_status()
    content_type = (response.headers.get('Content-Type') or '').lower()
    
    # Allow AVIF content type
    if content_type and 'image/' not in content_type and 'avif' not in content_type:
        raise ValueError('URL did not return an image.')

    # Register AVIF plugin
    try:
        import pillow_avif
    except ImportError:
        pass  # Plugin not available, continue anyway
    
    _register_heif_opener()

    image = None
    first_error = None
    
    # Try PIL first (with AVIF plugin if available)
    try:
        image = PILImage.open(BytesIO(response.content))
        image.load()
    except Exception as e:
        first_error = str(e)
        # If PIL failed, raise error with details
        raise ValueError(f'Unsupported image format. Error: {first_error}. Try JPG or PNG.')

    if crop_box is None and crop_ratio:
        try:
            rx, ry, rw, rh = crop_ratio
            img_w, img_h = image.size
            crop_box = (
                int(rx * img_w),
                int(ry * img_h),
                int(rw * img_w),
                int(rh * img_h),
            )
        except Exception:
            crop_box = None

    if crop_box:
        x, y, width, height = crop_box
        img_w, img_h = image.size
        x = max(0, min(int(x), img_w - 1))
        y = max(0, min(int(y), img_h - 1))
        width = max(1, min(int(width), img_w - x))
        height = max(1, min(int(height), img_h - y))
        image = image.crop((x, y, x + width, y + height))

    # Convert to RGB for better compatibility
    if image.mode in ('RGBA', 'LA', 'P'):
        # Create white background for transparency
        background = PILImage.new('RGB', image.size, (255, 255, 255))
        if image.mode == 'P':
            image = image.convert('RGBA')
        background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
        image = background
    elif image.mode != 'RGB':
        image = image.convert('RGB')

    filename = f"edit_photo_{uuid.uuid4().hex}.png"
    rel_path = os.path.join('edit_photo', filename)
    media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
    abs_path = os.path.join(media_root, rel_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)

    image.save(abs_path, format='PNG')
    public_url = f"{settings.MEDIA_URL}{rel_path.replace(os.sep, '/')}"

    return public_url, filename


def convert_uploaded_file_to_png(uploaded_file, upload_to='converted_images'):
    """
    Convert an uploaded file (including AVIF) to PNG format.
    Returns a relative path that can be saved to an ImageField.
    """
    _register_heif_opener()
    
    try:
        # Try to open the uploaded file with PIL
        image = PILImage.open(uploaded_file)
        image.load()
    except UnidentifiedImageError:
        # If PIL can't identify it, try pillow_heif for AVIF/HEIF
        try:
            import pillow_heif
            uploaded_file.seek(0)
            heif = pillow_heif.read_heif(uploaded_file.read())
            image = heif.to_pillow()
        except Exception as exc:
            raise ValueError('Unsupported image format. Try JPG or PNG.') from exc
    
    # Convert to RGB (PNG doesn't need RGBA unless transparency is needed)
    if image.mode in ('RGBA', 'LA', 'P'):
        image = image.convert('RGBA')
    else:
        image = image.convert('RGB')
    
    # Generate unique filename
    filename = f"{uuid.uuid4().hex}.png"
    rel_path = os.path.join(upload_to, filename)
    media_root = getattr(settings, 'MEDIA_ROOT', None) or os.path.join(settings.BASE_DIR, 'media')
    abs_path = os.path.join(media_root, rel_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    
    # Save as PNG
    image.save(abs_path, format='PNG')
    
    # Return the relative path for Django's ImageField
    return rel_path.replace(os.sep, '/')


@staff_member_required(login_url='login')
def admin_add_product(request):
    """Admin Add Product Page"""
    if request.method == 'POST':
        try:
            # Get basic form data
            name = request.POST.get('name')
            price = normalize_decimal_input(request.POST.get('price'))
            margin_raw = normalize_decimal_input(request.POST.get('margin', '0'))
            old_price = normalize_decimal_input(request.POST.get('old_price'))
            stock = request.POST.get('stock')
            sold = request.POST.get('sold')
            is_active = request.POST.get('is_active') == 'on'
            is_top_deal = request.POST.get('is_top_deal') == 'on'

            try:
                base_price = Decimal(price or '0')
            except Exception:
                base_price = Decimal('0')

            try:
                margin_value = Decimal(margin_raw or '0')
            except Exception:
                margin_value = Decimal('0')

            final_price = base_price + margin_value
            old_price_value = old_price if old_price else None
            discount_percent = calc_discount_percent(final_price, old_price_value)
            
            # Convert rating and review_count to proper types
            try:
                rating = float(request.POST.get('rating', 0)) or 0
            except (TypeError, ValueError):
                rating = 0
            
            try:
                review_count = int(float(request.POST.get('review_count', 0))) or 0
            except (TypeError, ValueError):
                review_count = 0
            
            # Get new fields
            category = request.POST.get('category')
            sub_category = request.POST.get('sub_category', '').strip()
            sku = request.POST.get('sku', '')
            brand = request.POST.get('brand', '')
            product_link = request.POST.get('product_link', '').strip()
            description = request.POST.get('description', '')
            weight = request.POST.get('weight', '')
            color = request.POST.get('color', '')
            # Handle multiple size selections from checkboxes
            size_list = request.POST.getlist('size')
            size = ', '.join(size_list) if size_list else ''
            
            # Return & Payment Policy
            is_returnable = request.POST.get('is_returnable') == 'on'
            return_days = int(request.POST.get('return_days', 7))
            return_policy = request.POST.get('return_policy', '')
            
            # Payment Methods (multiple selection)
            payment_methods_list = request.POST.getlist('payment_methods')
            payment_methods = ','.join(payment_methods_list) if payment_methods_list else 'COD,ONLINE,UPI,CARD'
            
            # Get images
            image = crop_image_height(request.FILES.get('image'))
            descriptionImage = request.FILES.get('descriptionImage')
            
            # Process gallery images with error handling
            gallery_images = []
            gallery_files = request.FILES.getlist('gallery_images')
            for idx, gallery_file in enumerate(gallery_files, 1):
                try:
                    processed_image = crop_image_height(gallery_file)
                    if processed_image:
                        gallery_images.append(processed_image)
                    else:
                        messages.warning(request, f'Failed to process gallery image {idx}. Using original file.')
                        gallery_images.append(gallery_file)
                except Exception as e:
                    messages.warning(request, f'Error processing gallery image {idx}: {str(e)}. Skipping this image.')
                    continue

            # Process color variant images with optional color labels
            variant_images = []
            variant_colors = request.POST.getlist('variant_colors')
            variant_files = request.FILES.getlist('variant_images')
            for idx, variant_file in enumerate(variant_files, 1):
                try:
                    processed_image = crop_image_height(variant_file)
                    color_value = (variant_colors[idx - 1].strip() if idx - 1 < len(variant_colors) else '').strip()
                    if processed_image:
                        variant_images.append({'image': processed_image, 'color': color_value})
                    else:
                        messages.warning(request, f'Failed to process color variant image {idx}. Using original file.')
                        variant_images.append({'image': variant_file, 'color': color_value})
                except Exception as e:
                    messages.warning(request, f'Error processing color variant image {idx}: {str(e)}. Skipping this image.')
                    continue

            # Optional direct reel upload
            reel_video_file = request.FILES.get('reel_video_file')
            reel_thumbnail = request.FILES.get('reel_thumbnail')
            reel_title = (request.POST.get('reel_title') or '').strip()
            reel_description = (request.POST.get('reel_description') or '').strip()
            reel_is_published = request.POST.get('reel_is_published') == 'on'
            reel_duration_raw = (request.POST.get('reel_duration') or '0').strip()
            reel_duration = 0
            reel_order = int(request.POST.get('reel_order', 0))
            if reel_video_file:
                reel_content_type = (getattr(reel_video_file, 'content_type', '') or '').lower()
                if reel_content_type and not reel_content_type.startswith('video/'):
                    messages.error(request, 'Invalid reel file type. Please upload a video file.')
                    return redirect('admin_add_product')
                try:
                    reel_duration = max(int(reel_duration_raw), 0)
                except (TypeError, ValueError):
                    reel_duration = 0

            use_saved_images = request.POST.get('use_saved_images') == 'on'
            saved_info = request.session.get('edit_photo_saved') if use_saved_images else None
            if saved_info and not image:
                base_path = saved_info.get('base_path')
                target_dir = saved_info.get('target_dir')
                main_category = saved_info.get('main_category')

                if not category and main_category:
                    category_icon = CategoryIcon.objects.filter(name__iexact=main_category).first()
                    if category_icon:
                        category = category_icon.category_key

                if not sub_category:
                    sub_category = (saved_info.get('sub_category') or '').strip()

                def file_from_path(path):
                    return File(open(path, 'rb'), name=os.path.basename(path))

                main_path = os.path.join(target_dir, '1.png') if target_dir else None
                gallery_path = os.path.join(target_dir, '2.png') if target_dir else None
                desc_path = os.path.join(target_dir, '3.png') if target_dir else None

                try:
                    if main_path and os.path.isfile(main_path):
                        image = file_from_path(main_path)
                except Exception:
                    pass

                try:
                    if gallery_path and os.path.isfile(gallery_path):
                        gallery_images.append(file_from_path(gallery_path))
                except Exception:
                    pass

                try:
                    if desc_path and os.path.isfile(desc_path):
                        descriptionImage = file_from_path(desc_path)
                except Exception:
                    pass
            
            def normalize_sku(raw_sku):
                cleaned = (raw_sku or '').strip()
                if not cleaned:
                    return None
                unique_sku = cleaned
                counter = 1
                while Product.objects.filter(sku=unique_sku).exists():
                    unique_sku = f"{cleaned}-{counter}"
                    counter += 1
                return unique_sku

            sku = normalize_sku(sku)

            # Create product
            product = Product.objects.create(
                name=name,
                price=final_price,
                margin=margin_value,
                old_price=old_price_value,
                discount_percent=discount_percent,
                stock=stock,
                sold=sold or 0,
                rating=rating,
                review_count=review_count,
                is_active=True,
                is_top_deal=False,
                image=image,
                descriptionImage=descriptionImage,
                category=category if category else None,
                sub_category=sub_category,
                sku=sku,
                brand=brand,
                product_link=product_link,
                description=description,
                weight=weight,
                color=color,
                size=size,
                is_returnable=is_returnable,
                return_days=return_days,
                return_policy=return_policy,
                payment_methods=payment_methods
            )
            
            # Add gallery images if provided
            current_order = 0
            for idx, gallery_image in enumerate(gallery_images, start=1):
                current_order = idx
                ProductImage.objects.create(
                    product=product,
                    image=gallery_image,
                    order=idx,
                    is_active=True
                )

            # Add color variant images if provided
            for idx, variant in enumerate(variant_images, start=current_order + 1):
                ProductImage.objects.create(
                    product=product,
                    image=variant['image'],
                    color=variant['color'],
                    order=idx,
                    is_active=True
                )

            # Auto-generate approved reviews if rating/review_count provided
            if review_count > 0 and rating > 0:
                generate_auto_reviews(product, review_count, rating, request.user)

            if reel_video_file:
                from Hub.models import Reel

                Reel.objects.create(
                    title=reel_title or f'{product.name} Reel',
                    description=reel_description,
                    product=product,
                    video_file=reel_video_file,
                    thumbnail=reel_thumbnail,
                    duration=reel_duration,
                    order=reel_order,
                    is_published=reel_is_published,
                    is_processing=False,
                    created_by=request.user,
                )

            if use_saved_images:
                request.session.pop('edit_photo_saved', None)
                request.session.modified = True
            
            if reel_video_file:
                messages.success(
                    request,
                    f'Product "{product.name}" added successfully with {len(gallery_images)} gallery images and 1 reel.'
                )
            else:
                messages.success(
                    request,
                    f'Product "{product.name}" added successfully with {len(gallery_images)} gallery images!'
                )
            return redirect('admin_add_product')
            
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    # Get categories for dropdown
    categories = CategoryIcon.objects.filter(is_active=True).order_by('order', 'id')
    saved_info = request.session.get('edit_photo_saved')
    saved_images = None
    if saved_info:
        target_dir = saved_info.get('target_dir')
        saved_images = []
        for index, label in ((1, 'Main Product Image'), (2, 'Gallery Image'), (3, 'Description Image')):
            filename = f"{index}.png"
            path = os.path.join(target_dir, filename) if target_dir else ''
            saved_images.append({
                'label': label,
                'filename': filename,
                'path': path,
                'exists': bool(path and os.path.isfile(path)),
            })

    sub_categories = list(
        SubCategory.objects.filter(is_active=True)
        .order_by('order', 'name')
        .values('category_key', 'name')
    )

    return render(request, 'admin_panel/add_product.html', {
        'categories': categories,
        'saved_info': saved_info,
        'saved_images': saved_images,
        'sub_categories': sub_categories,
        'sub_categories_json': json.dumps(sub_categories),
    })


def generate_auto_reviews(product, review_total, target_avg, reviewer_user=None):
    """Generate approved reviews and rating distribution for a product."""
    try:
        review_total = int(review_total)
    except (TypeError, ValueError):
        return

    try:
        target_avg = float(target_avg)
    except (TypeError, ValueError):
        return

    if review_total <= 0 or target_avg <= 0:
        return

    target_avg = max(1.0, min(5.0, target_avg))

    # Gaussian-like weights around average
    weights = {}
    for r in range(1, 6):
        dist = (r - target_avg)
        weights[r] = pow(2.718281828, -(dist ** 2) / 2)

    weight_sum = sum(weights.values()) or 1
    raw = {r: (weights[r] / weight_sum) * review_total for r in range(1, 6)}

    counts = {r: int(raw[r]) for r in range(1, 6)}
    remainder = review_total - sum(counts.values())
    if remainder > 0:
        remainders = sorted(
            [(r, raw[r] - counts[r]) for r in range(1, 6)],
            key=lambda x: x[1],
            reverse=True
        )
        for r, _ in remainders[:remainder]:
            counts[r] += 1

    target_sum = int(round(target_avg * review_total))
    current_sum = sum(r * c for r, c in counts.items())

    while current_sum < target_sum:
        moved = False
        for r in range(4, 0, -1):
            if counts[r] > 0:
                counts[r] -= 1
                counts[r + 1] += 1
                current_sum += 1
                moved = True
                if current_sum >= target_sum:
                    break
        if not moved:
            break

    while current_sum > target_sum:
        moved = False
        for r in range(2, 6):
            if counts[r] > 0:
                counts[r] -= 1
                counts[r - 1] += 1
                current_sum -= 1
                moved = True
                if current_sum <= target_sum:
                    break
        if not moved:
            break

    if reviewer_user is None:
        reviewer_user = (
            User.objects.filter(is_superuser=True).first()
            or User.objects.filter(is_staff=True).first()
            or User.objects.first()
        )

    if reviewer_user is None:
        return

    reviewer_name = "Admin"
    reviewer_email = "admin@vibemall.local"
    if reviewer_user:
        reviewer_name = reviewer_user.get_full_name().strip() or reviewer_user.username
        reviewer_email = reviewer_user.email or reviewer_email

    for r in range(5, 0, -1):
        for _ in range(counts.get(r, 0)):
            ProductReview.objects.create(
                product=product,
                user=reviewer_user,
                rating=r,
                name=reviewer_name,
                email=reviewer_email,
                comment="",
                is_approved=True,
                is_auto_generated=True
            )

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_product_list(request):
    """Admin Product List Page"""
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'add_sub_category':
            sub_name = (request.POST.get('sub_category_name') or '').strip()
            sub_category_key = (request.POST.get('sub_category_key') or '').strip()
            if sub_name and sub_category_key:
                SubCategory.objects.get_or_create(
                    category_key=sub_category_key,
                    name=sub_name,
                    defaults={'is_active': True}
                )
                messages.success(request, f'Sub-category "{sub_name}" added.')
            else:
                messages.error(request, 'Please select a category and enter sub-category name.')
            return redirect('admin_product_list')


    products = (
        Product.objects
        .annotate(
            reel_count=Count('watch_shop_reels', distinct=True),
            total_reel_views=Coalesce(
                Sum('watch_shop_reels__view_count'),
                Value(0),
                output_field=IntegerField(),
            ),
            total_reel_likes=Coalesce(
                Sum('watch_shop_reels__like_count'),
                Value(0),
                output_field=IntegerField(),
            ),
        )
        .order_by('-id')
    )

    # Filters
    status_filter = request.GET.get('status', 'all')
    category_filter = request.GET.get('category', 'all')
    stock_filter = request.GET.get('stock', 'all')
    search_query = request.GET.get('q', '').strip()

    if status_filter == 'active':
        products = products.filter(is_active=True)
    elif status_filter == 'inactive':
        products = products.filter(is_active=False)

    if stock_filter == 'in':
        products = products.filter(stock__gt=0)
    elif stock_filter == 'out':
        products = products.filter(stock__lte=0)
    elif stock_filter == 'low':
        products = products.filter(stock__gt=0, stock__lte=10)

    if category_filter and category_filter != 'all':
        products = products.filter(category=category_filter)

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(brand__icontains=search_query)
        )

    # Export CSV
    if request.GET.get('export') == '1':
        import csv
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="products.csv"'
        writer = csv.writer(response)
        writer.writerow(['ID', 'Name', 'Category', 'Stock', 'SKU', 'Price', 'Status'])
        for product in products:
            writer.writerow([
                product.id,
                product.name,
                'General',
                product.stock,
                product.sku or '',
                product.price,
                'Active' if product.is_active else 'Inactive'
            ])
        return response
    
    # Pagination
    per_page = request.GET.get('per_page', '10')
    try:
        per_page = int(per_page)
    except ValueError:
        per_page = 10
    per_page = per_page if per_page in [10, 20, 50, 100] else 10

    paginator = Paginator(products, per_page)
    page = request.GET.get('page')
    
    try:
        products = paginator.page(page)
    except PageNotAnInteger:
        products = paginator.page(1)
    except EmptyPage:
        products = paginator.page(paginator.num_pages)

    # Summary cards
    paid_orders = Order.objects.filter(payment_status='PAID')
    in_store_orders = paid_orders.filter(payment_method='COD')
    website_orders = paid_orders.exclude(payment_method='COD')

    in_store_sales = in_store_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    website_sales = website_orders.aggregate(total=Sum('total_amount'))['total'] or 0
    discount_totals = paid_orders.aggregate(subtotal=Sum('subtotal'), total=Sum('total_amount'))
    discount_total = (discount_totals['subtotal'] or 0) - (discount_totals['total'] or 0)
    affiliate_sales = paid_orders.filter(is_resell=True).aggregate(total=Sum('total_amount'))['total'] or 0

    in_store_orders_count = in_store_orders.count()
    website_orders_count = website_orders.count()
    discount_orders_count = paid_orders.filter(subtotal__gt=0).count()
    affiliate_orders_count = paid_orders.filter(is_resell=True).count()

    query_params = request.GET.copy()
    query_params.pop('page', None)
    query_params.pop('export', None)

    category_choices = list(
        CategoryIcon.objects.filter(is_active=True)
        .order_by('order', 'id')
        .values_list('category_key', 'name')
    )

    sub_categories = list(
        SubCategory.objects.filter(is_active=True)
        .order_by('order', 'name')
        .values('category_key', 'name')
    )

    context = {
        'products': products,
        'status_filter': status_filter,
        'category_filter': category_filter,
        'stock_filter': stock_filter,
        'search_query': search_query,
        'per_page': per_page,
        'category_choices': category_choices,
        'sub_categories': sub_categories,

        'in_store_sales': in_store_sales,
        'website_sales': website_sales,
        'discount_total': discount_total,
        'affiliate_sales': affiliate_sales,
        'in_store_orders_count': in_store_orders_count,
        'website_orders_count': website_orders_count,
        'discount_orders_count': discount_orders_count,
        'affiliate_orders_count': affiliate_orders_count,
        'querystring': query_params.urlencode(),
    }
    return render(request, 'admin_panel/product_list.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_toggle_stock(request, product_id):
    product = get_object_or_404(Product, id=product_id)
    previous_stock = product.stock

    if product.stock > 0:
        product.stock = 0
    else:
        product.stock = 1

    product.is_active = True
    product.save(update_fields=['stock', 'is_active'])

    # Notify subscribers when restocked from zero
    if previous_stock <= 0 and product.stock > 0:
        notifications = ProductStockNotification.objects.filter(product=product, is_sent=False)
        for note in notifications:
            try:
                send_mail(
                    subject=f"{product.name} is back in stock",
                    message=f"Good news! {product.name} is available again. Visit the product page to purchase.",
                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                    recipient_list=[note.email],
                    fail_silently=False,
                )
                note.mark_sent()
            except Exception:
                pass

    return JsonResponse({
        'success': True,
        'stock': product.stock,
    })

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_product(request, product_id):
    """Admin Edit Product Page"""
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            # Update basic fields
            product.name = request.POST.get('name')
            price_raw = normalize_decimal_input(request.POST.get('price'))
            margin_raw = normalize_decimal_input(request.POST.get('margin', '0'))

            try:
                base_price = Decimal(price_raw or '0')
            except Exception:
                base_price = Decimal('0')

            try:
                margin_value = Decimal(margin_raw or '0')
            except Exception:
                margin_value = Decimal('0')

            product.price = base_price + margin_value
            product.margin = margin_value
            old_price_value = normalize_decimal_input(request.POST.get('old_price')) or None
            product.old_price = old_price_value
            product.discount_percent = calc_discount_percent(product.price, product.old_price)
            product.stock = request.POST.get('stock')
            product.sold = request.POST.get('sold') or 0
            product.category = request.POST.get('category') or None
            product.sub_category = request.POST.get('sub_category', '').strip()
            product.rating = request.POST.get('rating', 0)
            product.review_count = request.POST.get('review_count', 0)
            product.is_active = request.POST.get('is_active') == 'on'
            product.is_top_deal = request.POST.get('is_top_deal') == 'on'
            
            # Update new fields
            product.sku = request.POST.get('sku', '')
            product.brand = request.POST.get('brand', '')
            product.product_link = request.POST.get('product_link', '').strip()
            product.tags = request.POST.get('tags', '')
            product.description = request.POST.get('description', '')
            product.weight = request.POST.get('weight', '')
            product.dimensions = request.POST.get('dimensions', '')
            product.color = request.POST.get('color', '')
            size_list = request.POST.getlist('size')
            product.size = ', '.join(size_list) if size_list else ''
            product.shipping_info = request.POST.get('shipping_info', '')
            product.care_info = request.POST.get('care_info', '')
            
            # Update image if provided
            if 'image' in request.FILES:
                product.image = crop_image_height(request.FILES['image'])
            
            # Update description image if provided
            if 'descriptionImage' in request.FILES:
                product.descriptionImage = request.FILES['descriptionImage']
            
            product.save()
            
            # Handle gallery images if provided
            if 'gallery_images' in request.FILES:
                gallery_files = request.FILES.getlist('gallery_images')
                processed_gallery_images = []
                
                for idx, gallery_file in enumerate(gallery_files, 1):
                    try:
                        processed_image = crop_image_height(gallery_file)
                        if processed_image:
                            processed_gallery_images.append(processed_image)
                        else:
                            messages.warning(request, f'Failed to process gallery image {idx}. Using original file.')
                            processed_gallery_images.append(gallery_file)
                    except Exception as e:
                        messages.warning(request, f'Error processing gallery image {idx}: {str(e)}. Skipping this image.')
                        continue
                
                current_gallery_count = ProductImage.objects.filter(product=product).count()
                for idx, gallery_image in enumerate(processed_gallery_images, start=current_gallery_count + 1):
                    ProductImage.objects.create(
                        product=product,
                        image=gallery_image,
                        order=idx,
                        is_active=True
                    )

            reel_added = False

            # Optional direct reel upload from edit product page
            new_reel_video_file = request.FILES.get('new_reel_video_file')
            if new_reel_video_file:
                from Hub.models import Reel

                reel_content_type = (getattr(new_reel_video_file, 'content_type', '') or '').lower()
                if reel_content_type and not reel_content_type.startswith('video/'):
                    messages.error(request, 'Invalid reel file type. Please upload a video file.')
                    return redirect('admin_edit_product', product_id=product_id)

                new_reel_title = (request.POST.get('new_reel_title') or '').strip() or f'{product.name} Reel'
                new_reel_description = (request.POST.get('new_reel_description') or '').strip()
                new_reel_thumbnail = request.FILES.get('new_reel_thumbnail')
                new_reel_is_published = request.POST.get('new_reel_is_published') == 'on'
                new_reel_duration_raw = (request.POST.get('new_reel_duration') or '0').strip()
                try:
                    new_reel_duration = max(int(new_reel_duration_raw), 0)
                except (TypeError, ValueError):
                    new_reel_duration = 0
                new_reel_order = int(request.POST.get('new_reel_order', 0))

                Reel.objects.create(
                    title=new_reel_title,
                    description=new_reel_description,
                    product=product,
                    video_file=new_reel_video_file,
                    thumbnail=new_reel_thumbnail,
                    duration=new_reel_duration,
                    order=new_reel_order,
                    is_published=new_reel_is_published,
                    is_processing=False,
                    created_by=request.user,
                )
                reel_added = True
            
            if reel_added:
                messages.success(request, f'Product "{product.name}" updated successfully and reel added.')
            else:
                messages.success(request, f'Product "{product.name}" updated successfully!')
            return redirect('admin_product_list')
        
        except Exception as e:
            messages.error(request, f'Error updating product: {str(e)}')
            return redirect('admin_edit_product', product_id=product_id)
    
    try:
        base_price = Decimal(str(product.price or 0)) - Decimal(str(product.margin or 0))
    except Exception:
        base_price = product.price

    product_reels_qs = product.watch_shop_reels.all().order_by('-created_at')
    reel_totals = product.watch_shop_reels.aggregate(
        total_reels=Count('id'),
        total_views=Coalesce(Sum('view_count'), Value(0), output_field=IntegerField()),
        total_likes=Coalesce(Sum('like_count'), Value(0), output_field=IntegerField()),
    )

    context = {
        'product': product,
        'base_price': base_price,
        'product_reels': product_reels_qs[:10],
        'product_reel_count': reel_totals.get('total_reels', 0) or 0,
        'product_reel_views': reel_totals.get('total_views', 0) or 0,
        'product_reel_likes': reel_totals.get('total_likes', 0) or 0,
        'categories': CategoryIcon.objects.filter(is_active=True).order_by('order', 'id'),
        'gallery_images': product.additional_images.filter(is_active=True).order_by('order'),
        'sub_categories': list(
            SubCategory.objects.filter(is_active=True)
            .order_by('order', 'name')
            .values('category_key', 'name')
        ),
        'sub_categories_json': json.dumps(list(
            SubCategory.objects.filter(is_active=True)
            .order_by('order', 'name')
            .values('category_key', 'name')
        )),
    }
    return render(request, 'admin_panel/edit_product.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_product(request, product_id):
    """Delete Product"""
    product = get_object_or_404(Product, id=product_id)
    product_name = product.name
    product.delete()
    
    messages.success(request, f'Product "{product_name}" deleted successfully!')
    return redirect('admin_product_list')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_gallery_image(request, image_id):
    """Delete Gallery Image"""
    gallery_image = get_object_or_404(ProductImage, id=image_id)
    product_name = gallery_image.product.name
    gallery_image.delete()
    
    messages.success(request, f'Gallery image deleted from product "{product_name}"!')
    return redirect('admin_edit_product', product_id=gallery_image.product.id)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_categories(request):
    """Admin Category Management Page"""
    categories = CategoryIcon.objects.all().order_by('order', 'id')
    
    context = {
        'categories': categories,
    }
    return render(request, 'admin_panel/categories.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_subcategories(request):
    """Admin Sub-Category Management Page"""
    categories = CategoryIcon.objects.all().order_by('order', 'id')
    sub_categories = SubCategory.objects.all().order_by('category_key', 'order', 'name')

    context = {
        'categories': categories,
        'sub_categories': sub_categories,
        'sub_categories_json': json.dumps(
            list(sub_categories.values(
                'category_key',
                'name',
                'icon_class',
                'icon_color',
                'icon_size',
                'background_gradient',
                'icon_image',
                'is_active'
            ))
        ),
    }
    return render(request, 'admin_panel/subcategories.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_save_subcategory_icon(request):
    """Create or update sub-category icon settings."""
    category_key = request.POST.get('category_key', '').strip()
    sub_category_name = request.POST.get('sub_category_name', '').strip()
    custom_name = request.POST.get('sub_category_name_custom', '').strip()
    if custom_name:
        sub_category_name = custom_name

    if not category_key or not sub_category_name:
        messages.error(request, 'Please select a category and sub-category.')
        return redirect('admin_categories')

    icon_class = request.POST.get('icon_class', '').strip()
    icon_color = request.POST.get('icon_color', '#0288d1')
    background_gradient = request.POST.get(
        'background_gradient',
        'linear-gradient(135deg, #e0f7ff 0%, #b3e5fc 100%)'
    )
    icon_size_raw = request.POST.get('icon_size', '').strip()
    is_active = request.POST.get('is_active') == 'on'

    try:
        icon_size = int(icon_size_raw)
    except (TypeError, ValueError):
        icon_size = 48

    icon_size = max(16, min(icon_size, 128))

    sub_category, _ = SubCategory.objects.get_or_create(
        category_key=category_key,
        name=sub_category_name
    )

    sub_category.icon_class = icon_class
    sub_category.icon_color = icon_color
    sub_category.background_gradient = background_gradient
    sub_category.icon_size = icon_size
    sub_category.is_active = is_active

    if request.FILES.get('icon_image'):
        sub_category.icon_image = request.FILES['icon_image']

    sub_category.save()

    messages.success(
        request,
        f'Sub-category "{sub_category.name}" updated successfully!'
    )
    return redirect('admin_categories')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_category(request):
    """Admin Add Category Page"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            category_key = request.POST.get('category_key')
            icon_class = request.POST.get('icon_class', '')
            icon_color = request.POST.get('icon_color', '#0288d1')
            background_gradient = request.POST.get('background_gradient', 'linear-gradient(135deg, #e0f7ff 0%, #b3e5fc 100%)')
            icon_size_raw = request.POST.get('icon_size', '').strip()
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'

            try:
                icon_size = int(icon_size_raw)
            except (TypeError, ValueError):
                icon_size = 48

            icon_size = max(16, min(icon_size, 128))
            
            category = CategoryIcon.objects.create(
                name=name,
                category_key=category_key,
                icon_class=icon_class,
                icon_color=icon_color,
                background_gradient=background_gradient,
                icon_size=icon_size,
                order=order,
                is_active=is_active
            )
            
            # Handle icon image upload
            if request.FILES.get('icon_image'):
                category.icon_image = request.FILES['icon_image']

            # Handle category card image upload (homepage photo card)
            if request.FILES.get('card_image'):
                category.card_image = request.FILES['card_image']

            if request.FILES.get('icon_image') or request.FILES.get('card_image'):
                category.save()
            
            messages.success(request, f'Category "{category.name}" added successfully!')
            return redirect('admin_categories')
        
        except Exception as e:
            messages.error(request, f'Error adding category: {str(e)}')
            return redirect('admin_add_category')
    
    return render(request, 'admin_panel/add_category.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_category(request, category_id):
    """Admin Edit Category Page"""
    category = get_object_or_404(CategoryIcon, id=category_id)
    
    if request.method == 'POST':
        try:
            category.name = request.POST.get('name')
            category.category_key = request.POST.get('category_key')
            category.icon_class = request.POST.get('icon_class', '')
            category.icon_color = request.POST.get('icon_color', '#0288d1')
            category.background_gradient = request.POST.get('background_gradient', 'linear-gradient(135deg, #e0f7ff 0%, #b3e5fc 100%)')
            icon_size_raw = request.POST.get('icon_size', '').strip()
            category.order = request.POST.get('order', 0)
            category.is_active = request.POST.get('is_active') == 'on'

            try:
                icon_size = int(icon_size_raw)
            except (TypeError, ValueError):
                icon_size = 48

            category.icon_size = max(16, min(icon_size, 128))
            
            # Handle icon image upload
            if request.FILES.get('icon_image'):
                category.icon_image = request.FILES['icon_image']

            # Handle category card image upload/removal
            if request.POST.get('remove_card_image') == 'on' and category.card_image:
                category.card_image.delete(save=False)
                category.card_image = None
            if request.FILES.get('card_image'):
                category.card_image = request.FILES['card_image']
            
            category.save()
            
            messages.success(request, f'Category "{category.name}" updated successfully!')
            return redirect('admin_categories')
        
        except Exception as e:
            messages.error(request, f'Error updating category: {str(e)}')
            return redirect('admin_edit_category', category_id=category_id)
    
    context = {
        'category': category,
    }
    return render(request, 'admin_panel/edit_category.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_photo(request):
    results = []
    errors = []
    show_downloads = False
    saved_message = ''
    resize_result = None

    product_image_root = r"D:\VibeMallProduct\ProductImage"

    def build_folder_index(root_path):
        folder_index = {}
        if not os.path.isdir(root_path):
            return folder_index
        for category_name in sorted(os.listdir(root_path)):
            category_path = os.path.join(root_path, category_name)
            if not os.path.isdir(category_path):
                continue
            sub_index = {}
            for sub_name in sorted(os.listdir(category_path)):
                sub_path = os.path.join(category_path, sub_name)
                if not os.path.isdir(sub_path):
                    continue
                numbers = set()
                for entry in os.listdir(sub_path):
                    entry_path = os.path.join(sub_path, entry)
                    name = os.path.splitext(entry)[0]
                    if os.path.isdir(entry_path) and entry.isdigit():
                        numbers.add(int(entry))
                    elif name.isdigit():
                        numbers.add(int(name))
                sub_index[sub_name] = sorted(numbers)
            folder_index[category_name] = sub_index
        return folder_index

    if request.method == 'POST':
        action = request.POST.get('action') or 'convert'

        if action == 'resize_image':
            uploaded_image = request.FILES.get('resize_image')
            width_raw = (request.POST.get('resize_width') or '').strip()
            height_raw = (request.POST.get('resize_height') or '').strip()
            keep_aspect = request.POST.get('keep_aspect') == 'on'

            if not uploaded_image:
                errors.append('Please upload an image to resize.')
            else:
                try:
                    target_width = int(width_raw) if width_raw else None
                except (TypeError, ValueError):
                    target_width = None
                    errors.append('Width must be a valid number.')

                try:
                    target_height = int(height_raw) if height_raw else None
                except (TypeError, ValueError):
                    target_height = None
                    errors.append('Height must be a valid number.')

                if target_width is not None and (target_width < 1 or target_width > 10000):
                    errors.append('Width must be between 1 and 10000.')
                if target_height is not None and (target_height < 1 or target_height > 10000):
                    errors.append('Height must be between 1 and 10000.')

                if not errors:
                    try:
                        resize_result = _resize_uploaded_image(
                            uploaded_image,
                            target_width=target_width,
                            target_height=target_height,
                            keep_aspect=keep_aspect
                        )
                    except Exception as exc:
                        errors.append(f'Resize failed: {exc}')

        elif action == 'save_to_folder':
            converted_urls = request.POST.getlist('converted_urls')
            main_category = (request.POST.get('main_category') or '').strip()
            sub_category = (request.POST.get('sub_category') or '').strip()
            product_folder = (request.POST.get('product_folder') or '').strip()
            create_folder = request.POST.get('create_folder') == 'on'

            if not main_category or not sub_category or not product_folder:
                errors.append('Please select Main category, Sub_Category, and Product_folder.')
            elif not product_folder.isdigit():
                errors.append('Product_folder must be a number like 1, 2, 3.')
            else:
                target_dir = os.path.normpath(
                    os.path.join(product_image_root, main_category, sub_category, product_folder)
                )
                root_norm = os.path.normpath(product_image_root)
                if os.path.commonpath([root_norm, target_dir]) != root_norm:
                    errors.append('Invalid target folder path.')
                else:
                    if create_folder:
                        os.makedirs(target_dir, exist_ok=True)
                    if not os.path.isdir(target_dir):
                        errors.append('Target folder does not exist. Enable create folder.')
                    else:
                        to_delete = []
                        for idx, url in enumerate(converted_urls, start=1):
                            safe_path, _ = _safe_edit_photo_path_from_url(url)
                            if not safe_path:
                                errors.append(f"{url}: Invalid converted image.")
                                continue
                            dest_path = os.path.join(target_dir, f"{idx}.png")
                            try:
                                shutil.copyfile(safe_path, dest_path)
                                to_delete.append(safe_path)
                            except Exception as exc:
                                errors.append(f"{url}: {exc}")
                        if not errors:
                            saved_message = f"Saved {len(converted_urls)} images to {target_dir}"
                            request.session['edit_photo_saved'] = {
                                'main_category': main_category,
                                'sub_category': sub_category,
                                'product_folder': product_folder,
                                'base_path': product_image_root,
                                'target_dir': target_dir,
                            }
                            request.session.modified = True
                            for path in to_delete:
                                try:
                                    if os.path.isfile(path):
                                        os.remove(path)
                                except Exception:
                                    pass
                            show_downloads = True

            for url in converted_urls:
                filename = os.path.basename(url)
                results.append({
                    'source': url,
                    'url': url,
                    'filename': filename,
                    'download_name': filename,
                })
                show_downloads = True

        elif action == 'apply_crop':
            converted_urls = request.POST.getlist('converted_urls')
            ratio_x = request.POST.get('crop_ratio_x')
            ratio_y = request.POST.get('crop_ratio_y')
            ratio_w = request.POST.get('crop_ratio_w')
            ratio_h = request.POST.get('crop_ratio_h')

            try:
                crop_ratio = (float(ratio_x), float(ratio_y), float(ratio_w), float(ratio_h))
            except (TypeError, ValueError):
                crop_ratio = None

            if not crop_ratio:
                errors.append('Please crop the selected image before applying to all.')
            else:
                for idx, url in enumerate(converted_urls, start=1):
                    safe_path, _ = _safe_edit_photo_path_from_url(url)
                    if not safe_path:
                        errors.append(f"{url}: Invalid converted image.")
                        continue
                    try:
                        public_url, filename = _crop_png_by_ratio(safe_path, crop_ratio, index=idx)
                        results.append({
                            'source': url,
                            'url': public_url,
                            'filename': filename,
                            'download_name': filename,
                        })
                    except Exception as exc:
                        errors.append(f"{url}: {exc}")
                if results:
                    show_downloads = True
        else:
            urls = request.POST.getlist('image_urls')
            for raw_url in urls:
                url = (raw_url or '').strip()
                if not url:
                    continue

                try:
                    public_url, filename = convert_url_to_png(url)
                    results.append({
                        'source': url,
                        'url': public_url,
                        'filename': filename,
                    })
                except Exception as exc:
                    errors.append(f"{url}: {exc}")

    folder_index = build_folder_index(product_image_root)
    return render(request, 'admin_panel/edit_photo.html', {
        'results': results,
        'errors': errors,
        'show_downloads': show_downloads,
        'saved_message': saved_message,
        'resize_result': resize_result,
        'folder_index_json': json.dumps(folder_index),
        'category_options': sorted(folder_index.keys()),
        'product_image_root': product_image_root,
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_photo_preview(request):
    raw_url = (request.GET.get('url') or '').strip()
    if not raw_url:
        return HttpResponse('Missing url', status=400)

    parsed = urlparse(raw_url)
    if parsed.scheme not in ('http', 'https'):
        return HttpResponse('Invalid url', status=400)

    host = (parsed.hostname or '').lower()
    if host in ('localhost', '127.0.0.1', '::1') or host.endswith('.local'):
        return HttpResponse('Blocked host', status=400)

    try:
        ip = ipaddress.ip_address(host)
        if ip.is_private or ip.is_loopback or ip.is_link_local:
            return HttpResponse('Blocked host', status=400)
    except ValueError:
        pass

    try:
        response = requests.get(raw_url, timeout=10)
        response.raise_for_status()
    except Exception:
        return HttpResponse('Failed to fetch', status=400)

    content_type = (response.headers.get('Content-Type') or '').lower()
    if content_type and 'image/' not in content_type:
        return HttpResponse('URL did not return an image', status=400)

    content = response.content
    _register_heif_opener()
    try:
        image = PILImage.open(BytesIO(content))
        image.load()
        image = image.convert('RGBA')
        output = BytesIO()
        image.save(output, format='PNG')
        output.seek(0)
        return HttpResponse(output.getvalue(), content_type='image/png')
    except Exception:
        pass
    if len(content) > 8 * 1024 * 1024:
        return HttpResponse('File too large', status=400)

    content_type = response.headers.get('Content-Type') or 'image/jpeg'
    return HttpResponse(content, content_type=content_type)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_category(request, category_id):
    """Delete Category"""
    category = get_object_or_404(CategoryIcon, id=category_id)
    category_name = category.name
    category.delete()
    
    messages.success(request, f'Category "{category_name}" deleted successfully!')
    return redirect('admin_categories')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_brand_partners(request):
    """Admin Brand Partners Management Page"""
    brand_partners = BrandPartner.objects.all().order_by('order', 'name')
    
    context = {
        'brand_partners': brand_partners,
    }
    return render(request, 'admin_panel/brand_partners.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_brand_partner(request):
    """Add New Brand Partner"""
    if request.method == 'POST':
        name = request.POST.get('name')
        logo = request.FILES.get('logo')
        link_url = request.POST.get('link_url', '')
        order = request.POST.get('order', 0)
        is_active = request.POST.get('is_active') == 'on'
        
        if name and logo:
            partner = BrandPartner.objects.create(
                name=name,
                logo=logo,
                link_url=link_url,
                order=order,
                is_active=is_active
            )
            messages.success(request, f'Brand Partner "{name}" added successfully!')
            return redirect('admin_brand_partners')
        else:
            messages.error(request, 'Please provide brand name and logo!')
    
    return render(request, 'admin_panel/add_brand_partner.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_brand_partner(request, partner_id):
    """Edit Brand Partner"""
    partner = get_object_or_404(BrandPartner, id=partner_id)
    
    if request.method == 'POST':
        partner.name = request.POST.get('name')
        link_url = request.POST.get('link_url', '')
        partner.link_url = link_url
        partner.order = request.POST.get('order', 0)
        partner.is_active = request.POST.get('is_active') == 'on'
        
        # Update logo if new one uploaded
        if request.FILES.get('logo'):
            partner.logo = request.FILES.get('logo')
        
        partner.save()
        messages.success(request, f'Brand Partner "{partner.name}" updated successfully!')
        return redirect('admin_brand_partners')
    
    context = {
        'partner': partner,
    }
    return render(request, 'admin_panel/edit_brand_partner.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_site_settings(request):
    """Admin Site Settings - Logo, Name, Contact Info"""
    settings_obj = SiteSettings.get_settings()
    
    if request.method == 'POST':
        settings_obj.site_name = request.POST.get('site_name', 'VibeMall')
        settings_obj.site_name_html = request.POST.get('site_name_html', '')
        settings_obj.tagline = request.POST.get('tagline', '')
        settings_obj.contact_email = request.POST.get('contact_email', 'info.vibemall@gmail.com')
        settings_obj.contact_phone = request.POST.get('contact_phone', '+91 1234567890')
        settings_obj.facebook_url = request.POST.get('facebook_url', '')
        settings_obj.instagram_url = request.POST.get('instagram_url', '')
        settings_obj.twitter_url = request.POST.get('twitter_url', '')
        settings_obj.youtube_url = request.POST.get('youtube_url', '')
        
        # Handle logo uploads
        if request.FILES.get('site_logo'):
            settings_obj.site_logo = request.FILES.get('site_logo')
        
        if request.FILES.get('site_favicon'):
            settings_obj.site_favicon = request.FILES.get('site_favicon')
        
        if request.FILES.get('admin_logo'):
            settings_obj.admin_logo = request.FILES.get('admin_logo')
        
        settings_obj.save()
        messages.success(request, 'Site settings updated successfully!')
        return redirect('admin_site_settings')
    
    context = {
        'settings': settings_obj,
    }
    return render(request, 'admin_panel/site_settings.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_newsletter_subscribers(request):
    """Newsletter subscriber tracking page in custom admin panel."""
    draft_subject = ''
    draft_title = ''
    draft_body = ''
    draft_cta_text = ''
    draft_cta_url = ''
    draft_recipient_scope = 'active'

    if request.method == 'POST' and request.POST.get('action') == 'send_newsletter':
        draft_subject = (request.POST.get('subject') or '').strip()
        draft_title = (request.POST.get('title') or '').strip()
        draft_body = (request.POST.get('body') or '').strip()
        draft_cta_text = (request.POST.get('cta_text') or '').strip()
        draft_cta_url = (request.POST.get('cta_url') or '').strip()
        draft_recipient_scope = (request.POST.get('recipient_scope') or 'active').strip().lower()

        if not draft_subject or not draft_title or not draft_body:
            messages.error(request, 'Subject, title, and newsletter content are required.')
        elif bool(draft_cta_text) != bool(draft_cta_url):
            messages.error(request, 'Please provide both CTA text and CTA URL, or leave both empty.')
        elif draft_cta_url and not (
            draft_cta_url.startswith('http://') or
            draft_cta_url.startswith('https://') or
            draft_cta_url.startswith('/')
        ):
            messages.error(request, 'CTA URL must start with http://, https://, or /.')
        else:
            recipients_qs = NewsletterSubscription.objects.all()
            if draft_recipient_scope == 'active':
                recipients_qs = recipients_qs.filter(is_active=True)
            elif draft_recipient_scope == 'inactive':
                recipients_qs = recipients_qs.filter(is_active=False)

            recipient_emails = list(
                recipients_qs.values_list('email', flat=True).distinct()
            )

            if not recipient_emails:
                messages.warning(request, 'No subscribers found for the selected recipient scope.')
            else:
                site_settings_obj = SiteSettings.get_settings()
                site_url = (getattr(settings, 'SITE_URL', '') or '').rstrip('/')
                if not site_url:
                    site_url = request.build_absolute_uri('/').rstrip('/')

                logo_url = ''
                try:
                    if site_settings_obj and site_settings_obj.site_logo:
                        raw_logo_url = site_settings_obj.site_logo.url
                        if raw_logo_url.startswith('http://') or raw_logo_url.startswith('https://'):
                            logo_url = raw_logo_url
                        else:
                            logo_url = f"{site_url}{raw_logo_url}"
                except Exception:
                    logo_url = ''

                from_email = (
                    getattr(settings, 'DEFAULT_FROM_EMAIL', None) or
                    getattr(settings, 'EMAIL_HOST_USER', None) or
                    'no-reply@vibemall.com'
                )

                normalized_cta_url = draft_cta_url
                if normalized_cta_url.startswith('/'):
                    normalized_cta_url = f"{site_url}{normalized_cta_url}"

                success_count = 0
                failed_count = 0
                email_logs = []

                for recipient_email in recipient_emails:
                    try:
                        html_content = render_to_string('emails/newsletter_campaign.html', {
                            'site_settings': site_settings_obj,
                            'site_url': site_url,
                            'site_logo_url': logo_url,
                            'newsletter_title': draft_title,
                            'newsletter_body': draft_body,
                            'cta_text': draft_cta_text,
                            'cta_url': normalized_cta_url,
                            'recipient_email': recipient_email,
                            'now': timezone.now(),
                        })

                        text_content = f"{draft_title}\n\n{draft_body}\n\n"
                        if draft_cta_text and normalized_cta_url:
                            text_content += f"{draft_cta_text}: {normalized_cta_url}\n\n"
                        text_content += (
                            f"Support: {getattr(site_settings_obj, 'contact_email', 'info.vibemall@gmail.com')}\n"
                            f"You received this email because you subscribed to "
                            f"{getattr(site_settings_obj, 'site_name', 'VibeMall')} newsletter."
                        )
                        text_content = strip_tags(text_content)

                        email_message = EmailMultiAlternatives(
                            subject=draft_subject,
                            body=text_content,
                            from_email=from_email,
                            to=[recipient_email]
                        )
                        email_message.attach_alternative(html_content, "text/html")
                        email_message.send(fail_silently=False)

                        success_count += 1
                        email_logs.append(EmailLog(
                            user=request.user,
                            email_to=recipient_email,
                            email_type='PROMOTIONAL',
                            subject=draft_subject,
                            sent_successfully=True
                        ))
                    except Exception as exc:
                        failed_count += 1
                        logger.exception("Newsletter send failed for %s", recipient_email)
                        email_logs.append(EmailLog(
                            user=request.user,
                            email_to=recipient_email,
                            email_type='PROMOTIONAL',
                            subject=draft_subject,
                            sent_successfully=False,
                            error_message=str(exc)[:2000]
                        ))

                if email_logs:
                    EmailLog.objects.bulk_create(email_logs, batch_size=250)

                if failed_count == 0:
                    messages.success(
                        request,
                        f'Newsletter sent successfully to {success_count} subscriber(s).'
                    )
                elif success_count > 0:
                    messages.warning(
                        request,
                        f'Newsletter sent to {success_count} subscriber(s), failed for {failed_count}. '
                        'Check Email Logs for details.'
                    )
                else:
                    messages.error(
                        request,
                        f'Newsletter could not be sent. Failed attempts: {failed_count}.'
                    )

                return redirect('admin_newsletter_subscribers')

    search_query = (request.GET.get('search') or '').strip()
    status_filter = (request.GET.get('status') or 'all').strip().lower()
    source_filter = (request.GET.get('source') or 'all').strip()

    subscribers_qs = NewsletterSubscription.objects.all().order_by('-subscribed_at')

    if search_query:
        subscribers_qs = subscribers_qs.filter(email__icontains=search_query)

    if status_filter == 'active':
        subscribers_qs = subscribers_qs.filter(is_active=True)
    elif status_filter == 'inactive':
        subscribers_qs = subscribers_qs.filter(is_active=False)

    if source_filter != 'all':
        if source_filter == '__blank__':
            subscribers_qs = subscribers_qs.filter(source_page='')
        else:
            subscribers_qs = subscribers_qs.filter(source_page=source_filter)

    source_options = (
        NewsletterSubscription.objects.exclude(source_page='')
        .values_list('source_page', flat=True)
        .distinct()
        .order_by('source_page')
    )

    total_subscribers = NewsletterSubscription.objects.count()
    active_subscribers = NewsletterSubscription.objects.filter(is_active=True).count()
    inactive_subscribers = NewsletterSubscription.objects.filter(is_active=False).count()
    today_subscribers = NewsletterSubscription.objects.filter(
        subscribed_at__date=timezone.localdate()
    ).count()

    paginator = Paginator(subscribers_qs, 25)
    page_number = request.GET.get('page')
    subscribers = paginator.get_page(page_number)

    context = {
        'subscribers': subscribers,
        'search_query': search_query,
        'status_filter': status_filter,
        'source_filter': source_filter,
        'source_options': source_options,
        'total_subscribers': total_subscribers,
        'active_subscribers': active_subscribers,
        'inactive_subscribers': inactive_subscribers,
        'today_subscribers': today_subscribers,
        'draft_subject': draft_subject,
        'draft_title': draft_title,
        'draft_body': draft_body,
        'draft_cta_text': draft_cta_text,
        'draft_cta_url': draft_cta_url,
        'draft_recipient_scope': draft_recipient_scope,
    }
    return render(request, 'admin_panel/newsletter_subscribers.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_brand_partner(request, partner_id):
    """Delete Brand Partner"""
    partner = get_object_or_404(BrandPartner, id=partner_id)
    partner_name = partner.name
    partner.delete()
    
    messages.success(request, f'Brand Partner "{partner_name}" deleted successfully!')
    return redirect('admin_brand_partners')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_main_page_products(request):
    """Manage homepage sections: main categories, sub-category banners and ready-to-ship styles."""
    from .models import MainPageProduct, MainPageSubCategoryBanner, ReadyShipStyle

    valid_sections = {'categories', 'subcat_banners', 'ready_ship'}
    active_section = (request.GET.get('section') or '').strip()
    if active_section not in valid_sections:
        active_section = ''

    def redirect_with_section(section_name=''):
        target_section = section_name if section_name in valid_sections else active_section
        redirect_url = reverse('admin_main_page_products')
        if target_section:
            redirect_url = f"{redirect_url}?section={target_section}"
        return redirect(redirect_url)

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add':
            product_id = request.POST.get('product_id')
            category = request.POST.get('category')

            if product_id and category:
                try:
                    product = Product.objects.get(id=product_id)
                    existing = MainPageProduct.objects.filter(product=product, category=category).first()
                    if not existing:
                        MainPageProduct.objects.create(product=product, category=category)
                        messages.success(request, f'✓ {product.name} added to {category}')
                    else:
                        messages.warning(request, f'⚠ {product.name} already in {category}')
                except Product.DoesNotExist:
                    messages.error(request, 'Product not found')
            return redirect_with_section('categories')

        elif action == 'remove':
            item_id = request.POST.get('item_id')
            try:
                item = MainPageProduct.objects.get(id=item_id)
                product_name = item.product.name
                item.delete()
                messages.success(request, f'✓ {product_name} removed')
            except MainPageProduct.DoesNotExist:
                messages.error(request, 'Item not found')
            return redirect_with_section('categories')

        elif action == 'update_order':
            items = MainPageProduct.objects.all()
            for item in items:
                new_order = request.POST.get(f'order_{item.id}')
                if not new_order:
                    continue
                try:
                    item.order = int(new_order)
                    item.save()
                except (TypeError, ValueError):
                    continue
            messages.success(request, '✓ Display order updated')
            return redirect_with_section('categories')

        elif action == 'add_subcategory_banner':
            sub_category_id = request.POST.get('sub_category_id')
            title = (request.POST.get('banner_title') or '').strip()
            order_value = request.POST.get('banner_order', '0')
            image = request.FILES.get('banner_image')
            is_active = request.POST.get('banner_is_active') == 'on'

            try:
                order = int(order_value or 0)
            except (TypeError, ValueError):
                order = 0

            if not sub_category_id:
                messages.error(request, 'Please select a sub-category.')
                return redirect_with_section('subcat_banners')

            if not image:
                messages.error(request, 'Please upload an image for the sub-category banner.')
                return redirect_with_section('subcat_banners')

            try:
                sub_category = SubCategory.objects.get(id=sub_category_id, is_active=True)
            except SubCategory.DoesNotExist:
                messages.error(request, 'Selected sub-category not found.')
                return redirect_with_section('subcat_banners')

            # Convert AVIF/HEIF to PNG if needed
            try:
                converted_path = convert_uploaded_file_to_png(image, upload_to='main_page_subcategory_banners')
                # Create a File object from the converted path
                from django.core.files import File
                with open(os.path.join(settings.MEDIA_ROOT, converted_path), 'rb') as f:
                    image_file = File(f, name=os.path.basename(converted_path))
                    banner = MainPageSubCategoryBanner.objects.create(
                        title=title,
                        sub_category=sub_category,
                        order=order,
                        is_active=is_active,
                    )
                    banner.image.save(os.path.basename(converted_path), image_file, save=True)
                messages.success(request, f'✓ Banner added for {sub_category.name} (converted to PNG)')
            except ValueError as e:
                # If conversion fails, try to save the original file
                try:
                    MainPageSubCategoryBanner.objects.create(
                        title=title,
                        sub_category=sub_category,
                        image=image,
                        order=order,
                        is_active=is_active,
                    )
                    messages.success(request, f'✓ Banner added for {sub_category.name}')
                except Exception as save_error:
                    messages.error(request, f'Failed to save banner: {save_error}')
                    return redirect_with_section('subcat_banners')
            
            return redirect_with_section('subcat_banners')

        elif action == 'update_subcategory_banner':
            banner_id = request.POST.get('banner_id')
            if not banner_id:
                messages.error(request, 'Banner item not found.')
                return redirect_with_section('subcat_banners')

            banner_item = get_object_or_404(MainPageSubCategoryBanner, id=banner_id)
            sub_category_id = request.POST.get('sub_category_id')
            title = (request.POST.get('banner_title') or '').strip()
            order_value = request.POST.get('banner_order', '')
            new_image = request.FILES.get('banner_image')
            remove_image = request.POST.get('remove_banner_image') == 'on'
            is_active = request.POST.get('banner_is_active') == 'on'
            old_image_name = banner_item.image.name if banner_item.image else ''

            try:
                order = int(order_value) if order_value != '' else banner_item.order
            except (TypeError, ValueError):
                order = banner_item.order

            if sub_category_id:
                try:
                    banner_item.sub_category = SubCategory.objects.get(id=sub_category_id, is_active=True)
                except SubCategory.DoesNotExist:
                    messages.error(request, 'Selected sub-category not found.')
                    return redirect_with_section('subcat_banners')

            banner_item.title = title
            banner_item.order = order
            banner_item.is_active = is_active

            if remove_image and not new_image:
                messages.error(request, 'Please upload a replacement image when removing current image.')
                return redirect_with_section('subcat_banners')

            if new_image:
                # Convert AVIF/HEIF to PNG if needed
                try:
                    converted_path = convert_uploaded_file_to_png(new_image, upload_to='main_page_subcategory_banners')
                    from django.core.files import File
                    with open(os.path.join(settings.MEDIA_ROOT, converted_path), 'rb') as f:
                        image_file = File(f, name=os.path.basename(converted_path))
                        banner_item.image.save(os.path.basename(converted_path), image_file, save=False)
                except ValueError:
                    # If conversion fails, use original file
                    banner_item.image = new_image

            if not banner_item.image:
                messages.error(request, 'Banner image is required.')
                return redirect_with_section('subcat_banners')

            banner_item.save()
            if remove_image and old_image_name and old_image_name != banner_item.image.name:
                banner_item.image.storage.delete(old_image_name)
            messages.success(request, '✓ Sub-category banner updated')
            return redirect_with_section('subcat_banners')

        elif action == 'delete_subcategory_banner':
            banner_id = request.POST.get('banner_id')
            try:
                banner_item = MainPageSubCategoryBanner.objects.get(id=banner_id)
                banner_name = banner_item.display_title
                banner_item.delete()
                messages.success(request, f'✓ "{banner_name}" banner removed')
            except MainPageSubCategoryBanner.DoesNotExist:
                messages.error(request, 'Banner not found')
            return redirect_with_section('subcat_banners')

        elif action == 'add_ready_ship':
            product_id = request.POST.get('product_id')
            title = (request.POST.get('title') or '').strip()
            order_value = request.POST.get('order', '0')
            is_active = request.POST.get('is_active') == 'on'
            image = request.FILES.get('image')

            try:
                order = int(order_value or 0)
            except (TypeError, ValueError):
                order = 0

            if not product_id:
                messages.error(request, 'Please select a product.')
                return redirect_with_section('ready_ship')

            try:
                product = Product.objects.get(id=product_id, is_active=True)
            except Product.DoesNotExist:
                messages.error(request, 'Selected product not found.')
                return redirect_with_section('ready_ship')

            duplicate = ReadyShipStyle.objects.filter(product=product).exists()
            if duplicate:
                messages.warning(request, f'"{product.name}" is already added in Ready-to-Ship styles.')
                return redirect_with_section('ready_ship')

            ReadyShipStyle.objects.create(
                product=product,
                title=title,
                image=image,
                order=order,
                is_active=is_active,
            )
            messages.success(request, 'Ready-to-Ship style added successfully.')
            return redirect_with_section('ready_ship')

        elif action == 'update_ready_ship':
            item_id = request.POST.get('item_id')
            if not item_id:
                messages.error(request, 'Invalid style item.')
                return redirect_with_section('ready_ship')

            item = get_object_or_404(ReadyShipStyle, id=item_id)
            product_id = request.POST.get('product_id')
            title = (request.POST.get('title') or '').strip()
            order_value = request.POST.get('order', '0')
            is_active = request.POST.get('is_active') == 'on'
            remove_image = request.POST.get('remove_image') == 'on'
            new_image = request.FILES.get('image')

            try:
                order = int(order_value or 0)
            except (TypeError, ValueError):
                order = item.order

            if product_id:
                try:
                    product = Product.objects.get(id=product_id, is_active=True)
                except Product.DoesNotExist:
                    messages.error(request, 'Selected product not found.')
                    return redirect_with_section('ready_ship')

                duplicate = ReadyShipStyle.objects.filter(product=product).exclude(id=item.id).exists()
                if duplicate:
                    messages.warning(request, f'"{product.name}" is already assigned to another Ready-to-Ship item.')
                    return redirect_with_section('ready_ship')
                item.product = product

            item.title = title
            item.order = order
            item.is_active = is_active

            if remove_image and item.image:
                item.image.delete(save=False)
                item.image = None

            if new_image:
                item.image = new_image

            item.save()
            messages.success(request, 'Ready-to-Ship style updated successfully.')
            return redirect_with_section('ready_ship')

        elif action == 'delete_ready_ship':
            item_id = request.POST.get('item_id')
            item = get_object_or_404(ReadyShipStyle, id=item_id)
            item_name = item.display_title
            item.delete()
            messages.success(request, f'"{item_name}" removed successfully.')
            return redirect_with_section('ready_ship')

    # Main page category products
    categories = [
        ('category1', 'Category 1'),
        ('category2', 'Category 2'),
        ('category3', 'Category 3'),
        ('category4', 'Category 4'),
    ]
    
    main_page_items = MainPageProduct.objects.select_related('product').order_by('category', 'order')
    
    # Organize by category
    category_products = {}
    for cat_key, cat_name in categories:
        category_products[cat_key] = {
            'name': cat_name,
            'products': main_page_items.filter(category=cat_key)
        }
    
    # Get available products
    available_products = Product.objects.filter(is_active=True).exclude(
        id__in=MainPageProduct.objects.values_list('product_id', flat=True)
    ).order_by('-sold', 'name')

    sub_categories = SubCategory.objects.filter(is_active=True).order_by('category_key', 'order', 'name')
    main_page_subcategory_banners = (
        MainPageSubCategoryBanner.objects
        .select_related('sub_category')
        .order_by('order', 'id')
    )

    # Ready-to-Ship section data
    ready_ship_items = ReadyShipStyle.objects.select_related('product').order_by('order', 'id')
    used_ready_ship_product_ids = ready_ship_items.values_list('product_id', flat=True)
    ready_ship_available_products = (
        Product.objects
        .filter(is_active=True)
        .exclude(id__in=used_ready_ship_product_ids)
        .order_by('name')
    )
    ready_ship_all_products = Product.objects.filter(is_active=True).order_by('name')
    ready_ship_hero_id = (
        ready_ship_items
        .filter(is_active=True)
        .values_list('id', flat=True)
        .first()
    )

    context = {
        'active_section': active_section,
        'categories': categories,
        'category_products': category_products,
        'available_products': available_products[:50],  # Limit to 50 for dropdown
        'all_products_count': Product.objects.filter(is_active=True).count(),
        'main_category_item_count': main_page_items.count(),
        'sub_categories': sub_categories,
        'main_page_subcategory_banners': main_page_subcategory_banners,
        'subcat_banner_count': main_page_subcategory_banners.count(),
        'ready_ship_items': ready_ship_items,
        'ready_ship_available_products': ready_ship_available_products[:200],
        'ready_ship_all_products': ready_ship_all_products[:500],
        'ready_ship_hero_id': ready_ship_hero_id,
        'ready_ship_count': ready_ship_items.count(),
    }

    return render(request, 'admin_panel/main_page_products.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_ready_ship_styles(request):
    """Backward-compatible route: redirect to main page section selector."""
    return redirect(f"{reverse('admin_main_page_products')}?section=ready_ship")


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reviews(request):
    """Admin Review Management Page"""
    from django.db.models import Avg

    def refresh_product_rating(product):
        """Recalculate product rating and review count from approved reviews"""
        approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
        if approved_reviews.exists():
            product.review_count = approved_reviews.count()
            avg_rating = approved_reviews.aggregate(Avg('rating'))['rating__avg']
            product.rating = round(avg_rating, 1) if avg_rating else 0
        else:
            product.review_count = 0
            product.rating = 0
        product.save(update_fields=['review_count', 'rating'])

    # Handle bulk actions
    if request.method == 'POST':
        action = request.POST.get('action')
        review_ids = request.POST.getlist('review_ids')
        
        if action == 'approve' and review_ids:
            product_ids = list(ProductReview.objects.filter(id__in=review_ids).values_list('product_id', flat=True))
            ProductReview.objects.filter(id__in=review_ids).update(is_approved=True)
            for product in Product.objects.filter(id__in=product_ids):
                refresh_product_rating(product)
            messages.success(request, f'{len(review_ids)} review(s) approved successfully!')
        elif action == 'reject' and review_ids:
            product_ids = list(ProductReview.objects.filter(id__in=review_ids).values_list('product_id', flat=True))
            ProductReview.objects.filter(id__in=review_ids).update(is_approved=False)
            for product in Product.objects.filter(id__in=product_ids):
                refresh_product_rating(product)
            messages.success(request, f'{len(review_ids)} review(s) rejected successfully!')
        elif action == 'delete' and review_ids:
            product_ids = list(ProductReview.objects.filter(id__in=review_ids).values_list('product_id', flat=True))
            ProductReview.objects.filter(id__in=review_ids).delete()
            for product in Product.objects.filter(id__in=product_ids):
                refresh_product_rating(product)
            messages.success(request, f'{len(review_ids)} review(s) deleted successfully!')
        
        return redirect('admin_reviews')
    
    # Filter reviews
    filter_status = request.GET.get('status', 'all')
    reviews = ProductReview.objects.all().select_related('product', 'user').order_by('-created_at')
    
    if filter_status == 'pending':
        reviews = reviews.filter(is_approved=False)
    elif filter_status == 'approved':
        reviews = reviews.filter(is_approved=True)
    
    # Count stats
    total_reviews = ProductReview.objects.count()
    pending_reviews = ProductReview.objects.filter(is_approved=False).count()
    approved_reviews = ProductReview.objects.filter(is_approved=True).count()
    
    # Pagination
    paginator = Paginator(reviews, 20)
    page = request.GET.get('page')
    
    try:
        reviews = paginator.page(page)
    except PageNotAnInteger:
        reviews = paginator.page(1)
    except EmptyPage:
        reviews = paginator.page(paginator.num_pages)
    
    context = {
        'reviews': reviews,
        'filter_status': filter_status,
        'total_reviews': total_reviews,
        'pending_reviews': pending_reviews,
        'approved_reviews': approved_reviews,
    }
    return render(request, 'admin_panel/reviews.html', context)


# ===== ORDER MANAGEMENT (ADMIN PANEL) =====

# Helper Functions for Order Management
def send_order_status_email(order):
    """Send email notification to customer when order status changes"""
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    
    try:
        subject = f'Order {order.order_number} - Status Update'
        message = render_to_string('emails/order_status_update.html', {
            'order': order,
            'customer_name': order.user.get_full_name() or order.user.username
        })
        send_mail(
            subject,
            '',
            settings.DEFAULT_FROM_EMAIL,
            [order.user.email],
            html_message=message,
            fail_silently=True
        )
    except Exception:
        logger.exception("Email sending failed for order status update order_id=%s", order.id)


def send_admin_new_order_notification(order):
    """Send email notification to admin when new order is received"""
    from django.core.mail import send_mail
    from django.template.loader import render_to_string
    
    try:
        # Get admin email from settings (configurable)
        admin_settings = AdminEmailSettings.objects.first()
        admin_email = admin_settings.admin_email if admin_settings and admin_settings.is_active else 'info.vibemall@gmail.com'
        
        subject = f'New Order Received - {order.order_number}'
        message = f"""
        New Order Received!
        
        Order Number: {order.order_number}
        Customer: {order.user.get_full_name() or order.user.username}
        Email: {order.user.email}
        Phone: {order.user.userprofile.mobile_number if hasattr(order.user, 'userprofile') else 'N/A'}
        
        Order Details:
        Total Amount: ₹{order.total_amount}
        Payment Method: {order.get_payment_method_display()}
        Payment Status: {order.get_payment_status_display()}
        Order Status: {order.get_order_status_display()}
        
        Items:
        """
        
        for item in order.items.all():
            message += f"\n- {item.product_name} x {item.quantity} = ₹{item.subtotal}"
        
        message += f"\n\nShipping Address:\n{order.shipping_address}"
        message += f"\n\nOrder Date: {order.created_at.strftime('%d %b %Y, %I:%M %p')}"
        message += f"\n\nView Order: http://localhost:8000/admin-panel/orders/{order.id}/"
        
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [admin_email],
            fail_silently=True
        )
    except Exception:
        logger.exception("Admin email notification failed for order_id=%s", order.id)


def export_orders_to_excel(orders):
    """Export orders to Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse
    from datetime import datetime
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"
    
    # Headers
    headers = ['Order Number', 'Product', 'Quantity', 'Price', 'Customer Name', 'Email', 'Phone', 
               'Order Status', 'Payment Status', 'Order Date', 'Tracking Number']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for item in orders:
        ws.append([
            item.order.order_number,
            item.product_name,
            item.quantity,
            float(item.subtotal),
            item.order.user.get_full_name() or item.order.user.username,
            item.order.user.email,
            item.order.user.userprofile.mobile_number if hasattr(item.order.user, 'userprofile') else '',
            item.order.get_order_status_display(),
            item.order.get_payment_status_display(),
            item.order.created_at.strftime('%d-%m-%Y'),
            item.order.tracking_number or ''
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Order {datetime.now().strftime("%m/%Y")}.xlsx"'
    wb.save(response)
    
    return response


@login_required(login_url='login')
@staff_member_required(login_url='login')
@never_cache
def admin_orders(request):
    """Comprehensive Admin Order Management with 24+ features"""
    from django.core.paginator import Paginator
    from django.http import HttpResponse, JsonResponse
    from django.db.models import Sum, Count, Q
    import csv
    from datetime import datetime, timedelta

    # Remove invalid leftovers so stats and delete actions stay in sync.
    _cleanup_orphan_orders()

    # Handle Bulk Actions (POST)
    if request.method == 'POST':
        action = request.POST.get('action')
        order_ids = request.POST.getlist('selected_orders')
        
        if action and order_ids:
            # Normalize order IDs (remove blanks, duplicates, and strip spaces)
            normalized_ids = []
            for oid in order_ids:
                if not oid:
                    continue
                try:
                    normalized_ids.append(int(str(oid).strip()))
                except (ValueError, TypeError):
                    continue
            normalized_ids = list(set(normalized_ids))

            if not normalized_ids:
                messages.error(request, 'No valid orders were selected for bulk action.')
                return redirect('admin_orders')

            orders = Order.objects.filter(id__in=normalized_ids)

            if action in ['PENDING', 'PROCESSING', 'SHIPPED', 'DELIVERED', 'CANCELLED']:
                # Bulk status update
                from django.utils import timezone
                for order in orders:
                    old_status = order.order_status
                    order.order_status = action
                    # Set delivery_date and payment_status when status changes to DELIVERED
                    if action == 'DELIVERED' and old_status != 'DELIVERED':
                        order.delivery_date = timezone.now()
                        order.payment_status = 'PAID'
                    order.save()
                    # Create status history
                    OrderStatusHistory.objects.create(
                        order=order,
                        old_status=old_status,
                        new_status=action,
                        changed_by=request.user,
                        notes=f"Bulk status update by {request.user.username}"
                    )
                    # Send email notification to customer with beautiful template
                    email_sent = send_order_status_update_email(order, old_status, action)
                    if not email_sent:
                        messages.warning(request, f"Status updated for {order.order_number}, but email could not be sent.")
                messages.success(request, f"{orders.count()} orders updated to {action}")

            elif action == 'delete':
                # Bulk delete
                deleted_count, _ = orders.delete()
                _cleanup_orphan_orders()
                messages.success(request, f"{deleted_count} records deleted (orders and related items).")

            return redirect('admin_orders')
    
    # Get all order items (individual products with their order details)
    all_orders = OrderItem.objects.select_related('order', 'order__user', 'order__cancellation_request', 'product').order_by('-order__created_at')
    
    # === FILTERS ===
    # Status filter
    status_filter = request.GET.get('status', '')
    if status_filter:
        all_orders = all_orders.filter(order__order_status=status_filter)
    
    # Payment status filter
    payment_filter = request.GET.get('payment', '')
    if payment_filter:
        all_orders = all_orders.filter(order__payment_status=payment_filter)

    # Payment method filter
    payment_method_filter = request.GET.get('payment_method', '')
    if payment_method_filter:
        all_orders = all_orders.filter(order__payment_method=payment_method_filter)
    
    # Approval status filter
    approval_filter = request.GET.get('approval', '')
    if approval_filter:
        all_orders = all_orders.filter(order__approval_status=approval_filter)
    
    # Suspicious orders filter
    suspicious_filter = request.GET.get('suspicious', '')
    if suspicious_filter == '1':
        all_orders = all_orders.filter(order__is_suspicious=True)
    elif suspicious_filter == '0':
        all_orders = all_orders.filter(order__is_suspicious=False)
    
    # Quick date range filter (Today, Yesterday, Last Week)
    date_range = request.GET.get('date_range', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if date_range:
        from django.utils import timezone
        today = timezone.now().date()
        
        if date_range == 'today':
            all_orders = all_orders.filter(order__created_at__date=today)
        elif date_range == 'yesterday':
            yesterday = today - timedelta(days=1)
            all_orders = all_orders.filter(order__created_at__date=yesterday)
        elif date_range == 'last_week':
            last_week = today - timedelta(days=7)
            all_orders = all_orders.filter(order__created_at__date__gte=last_week)
    else:
        # Manual date range filter (only if quick filter not used)
        if date_from:
            all_orders = all_orders.filter(order__created_at__gte=date_from)
        if date_to:
            all_orders = all_orders.filter(order__created_at__lte=date_to)
    
    # Search filter (Order ID, Customer name, Phone, Email, Product name)
    search_query = request.GET.get('search', '')
    if search_query:
        all_orders = all_orders.filter(
            Q(order__order_number__icontains=search_query) |
            Q(order__user__username__icontains=search_query) |
            Q(order__user__email__icontains=search_query) |
            Q(order__user__first_name__icontains=search_query) |
            Q(order__user__last_name__icontains=search_query) |
            Q(order__user__userprofile__mobile_number__icontains=search_query) |
            Q(order__tracking_number__icontains=search_query) |
            Q(product_name__icontains=search_query)
        )
    
    # Sorting
    sort_by = request.GET.get('sort', '-order__created_at')
    # Fix sorting for fields that belong to Order model (since we're querying OrderItem)
    if sort_by == 'total_amount':
        sort_by = 'order__total_amount'
    elif sort_by == '-total_amount':
        sort_by = '-order__total_amount'
    elif sort_by == 'order_status':
        sort_by = 'order__order_status'
    elif sort_by == '-order_status':
        sort_by = '-order__order_status'
    elif sort_by == 'payment_status':
        sort_by = 'order__payment_status'
    elif sort_by == '-payment_status':
        sort_by = '-order__payment_status'
    all_orders = all_orders.order_by(sort_by)
    
    # === STATISTICS ===
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(order_status='PENDING').count()
    processing_orders = Order.objects.filter(order_status='PROCESSING').count()
    shipped_orders = Order.objects.filter(order_status='SHIPPED').count()
    delivered_orders = Order.objects.filter(order_status='DELIVERED').count()
    cancelled_orders = Order.objects.filter(order_status='CANCELLED').count()
    
    # Approval stats
    pending_approval_orders = Order.objects.filter(approval_status='PENDING_APPROVAL').count()
    approved_orders = Order.objects.filter(approval_status='APPROVED').count()
    rejected_orders = Order.objects.filter(approval_status='REJECTED').count()
    suspicious_orders = Order.objects.filter(is_suspicious=True).count()
    
    # Revenue stats (use Order model, not OrderItem)
    from decimal import Decimal
    paid_orders = Order.objects.filter(payment_status='PAID')
    total_revenue = paid_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    today_revenue = paid_orders.filter(created_at__date=datetime.now().date()).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    this_month_revenue = paid_orders.filter(created_at__month=datetime.now().month).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
    
    # Profit calculation (margin-based)
    total_profit_items = _order_item_profit(OrderItem.objects.filter(order__in=paid_orders))
    total_profit = total_profit_items + _return_fee_profit()
    today_profit_items = _order_item_profit(
        OrderItem.objects.filter(order__in=paid_orders, order__created_at__date=datetime.now().date())
    )
    today_profit = today_profit_items + _return_fee_profit(start_date=datetime.now().date(), end_date=datetime.now().date())
    month_start = datetime.now().replace(day=1).date()
    this_month_profit_items = _order_item_profit(
        OrderItem.objects.filter(order__in=paid_orders, order__created_at__date__gte=month_start)
    )
    this_month_profit = this_month_profit_items + _return_fee_profit(start_date=month_start)
    
    # Today's orders
    today_orders = Order.objects.filter(created_at__date=datetime.now().date()).count()
    
    # Payment methods list
    payment_methods = list(
        Order.objects.exclude(payment_method__isnull=True)
        .exclude(payment_method__exact='')
        .values_list('payment_method', flat=True)
        .distinct()
    )

    # === EXPORT TO CSV ===
    if request.GET.get('export') == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="Order {datetime.now().strftime("%m/%Y")}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Order Number', 'Product', 'Quantity', 'Price', 'Customer', 'Email', 'Phone', 'Order Status', 'Payment Status', 'Date'])
        
        for item in all_orders:
            writer.writerow([
                item.order.order_number,
                item.product_name,
                item.quantity,
                item.subtotal,
                item.order.user.get_full_name() or item.order.user.username,
                item.order.user.email,
                item.order.user.userprofile.mobile_number if hasattr(item.order.user, 'userprofile') else '',
                item.order.get_order_status_display(),
                item.order.get_payment_status_display(),
                item.order.created_at.strftime('%d-%m-%Y')
            ])
        
        return response
    
    # === EXPORT TO EXCEL ===
    if request.GET.get('export') == 'excel':
        return export_orders_to_excel(all_orders)
    
    # === PAGINATION ===
    page_size = request.GET.get('page_size', '20')
    paginator = Paginator(all_orders, int(page_size))
    page_number = request.GET.get('page', 1)
    orders = paginator.get_page(page_number)
    
    context = {
        'orders': orders,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'payment_method_filter': payment_method_filter,
        'approval_filter': approval_filter,
        'suspicious_filter': suspicious_filter,
        'search_query': search_query,
        'date_range': date_range,
        'date_from': date_from,
        'date_to': date_to,
        'sort_by': sort_by,
        'page_size': page_size,
        
        # Statistics
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        
        # Approval stats
        'pending_approval_orders': pending_approval_orders,
        'approved_orders': approved_orders,
        'rejected_orders': rejected_orders,
        'suspicious_orders': suspicious_orders,
        
        # Revenue
        'total_revenue': total_revenue,
        'today_revenue': today_revenue,
        'this_month_revenue': this_month_revenue,
        'today_orders': today_orders,

        # Payment methods
        'payment_methods': payment_methods,
        
        # Profit
        'total_profit': total_profit,
        'today_profit': today_profit,
        'this_month_profit': this_month_profit,
    }
    
    return render(request, 'admin_panel/orders.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_order_details(request, order_id):
    """Order Details with timeline, tracking, notes, and actions"""
    order = get_object_or_404(Order, id=order_id)
    from django.utils import timezone
    
    # Handle POST actions
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_status':
            from django.utils import timezone
            old_status = order.order_status
            new_status = request.POST.get('new_status')
            order.order_status = new_status
            # Set delivery_date and payment_status when status changes to DELIVERED
            if new_status == 'DELIVERED' and old_status != 'DELIVERED':
                order.delivery_date = timezone.now()
                order.payment_status = 'PAID'
            order.save()
            
            # Create status history
            OrderStatusHistory.objects.create(
                order=order,
                old_status=old_status,
                new_status=new_status,
                changed_by=request.user,
                notes=request.POST.get('status_notes', '')
            )
            
            # Send email notification with beautiful template
            email_sent = send_order_status_update_email(order, old_status, new_status)
            if email_sent:
                messages.success(request, f'Order status updated to {new_status}. Email sent to customer.')
            else:
                messages.warning(request, f'Order status updated to {new_status}, but email could not be sent. Please check email settings/logs.')
            
        elif action == 'add_tracking':
            order.tracking_number = request.POST.get('tracking_number')
            order.courier_name = request.POST.get('courier_name')
            order.save()
            messages.success(request, 'Tracking information updated')
            
        elif action == 'update_payment':
            old_payment_status = order.payment_status
            new_payment_status = request.POST.get('new_payment_status') or request.POST.get('payment_status')
            order.payment_status = new_payment_status
            order.save()
            messages.success(request, f'Payment status updated to {order.get_payment_status_display()}')
            
        elif action == 'add_note':
            order.admin_notes = request.POST.get('admin_notes')
            order.save()
            messages.success(request, 'Admin notes saved')
        elif action == 'create_rto_case':
            existing_rto = getattr(order, 'rto_case', None)
            if existing_rto:
                messages.info(request, 'RTO case already exists for this order.')
                return redirect('admin_rto_detail', rto_id=existing_rto.id)

            reason = (request.POST.get('rto_reason') or 'OTHER').strip()
            reason_notes = (request.POST.get('rto_reason_notes') or '').strip()
            rto_case = RTOCase.objects.create(
                order=order,
                reason=reason if reason in dict(RTOCase.REASON_CHOICES) else 'OTHER',
                reason_notes=reason_notes,
                courier_name=order.courier_name or '',
                tracking_number=order.tracking_number or '',
                last_attempted_at=timezone.now(),
            )
            _log_rto_history(rto_case, '', rto_case.status, request.user, 'RTO case created from order detail.')
            messages.success(request, 'RTO case created.')
            return redirect('admin_rto_detail', rto_id=rto_case.id)
        elif action in ['approve_cancel', 'reject_cancel']:
            cancel_request = getattr(order, 'cancellation_request', None)
            if not cancel_request or cancel_request.status != 'REQUESTED':
                messages.error(request, 'No pending cancellation request found.')
                return redirect('admin_order_details', order_id=order.id)

            now = timezone.now()
            admin_notes = request.POST.get('cancel_notes', '').strip()

            if action == 'approve_cancel':
                old_status = order.order_status
                order.order_status = 'CANCELLED'
                order.save(update_fields=['order_status'])

                for item in order.items.select_related('product'):
                    if item.product:
                        item.product.stock = F('stock') + item.quantity
                        item.product.save(update_fields=['stock'])

                if not cancel_request.refund_method and order.payment_method != 'COD' and order.payment_status == 'PAID':
                    cancel_request.refund_method = 'WALLET'

                refund_notes = ''
                if order.payment_method != 'COD' and order.payment_status == 'PAID':
                    refund_success, refund_notes = _process_cancellation_refund(order, cancel_request)
                    if not refund_success:
                        refund_notes = refund_notes or 'Refund failed'

                cancel_request.status = 'APPROVED'
                cancel_request.processed_at = now
                cancel_request.processed_by = request.user
                if admin_notes:
                    cancel_request.notes = f"{admin_notes}\n{cancel_request.notes}".strip()
                cancel_request.save()

                status_note = 'Cancellation approved.'
                if refund_notes:
                    status_note = f"{status_note} {refund_notes}"

                OrderStatusHistory.objects.create(
                    order=order,
                    old_status=old_status,
                    new_status='CANCELLED',
                    changed_by=request.user,
                    notes=status_note
                )
                if refund_notes and 'failed' in refund_notes.lower():
                    messages.warning(request, f'Cancellation approved and order cancelled, but refund issue: {refund_notes}')
                else:
                    messages.success(request, 'Cancellation approved and order cancelled.')
            else:
                cancel_request.status = 'REJECTED'
                cancel_request.processed_at = now
                cancel_request.processed_by = request.user
                if admin_notes:
                    cancel_request.notes = f"{admin_notes}\n{cancel_request.notes}".strip()
                cancel_request.save()

                OrderStatusHistory.objects.create(
                    order=order,
                    old_status=order.order_status,
                    new_status='CANCEL_REJECTED',
                    changed_by=request.user,
                    notes=admin_notes or 'Cancellation rejected.'
                )
                messages.success(request, 'Cancellation request rejected.')
        
        return redirect('admin_order_details', order_id=order.id)
    
    # Get status history
    status_history = order.status_history.all()
    
    context = {
        'order': order,
        'status_history': status_history,
        'cancel_request': getattr(order, 'cancellation_request', None),
        'rto_case': getattr(order, 'rto_case', None),
        'rto_reason_choices': RTOCase.REASON_CHOICES,
    }
    return render(request, 'admin_panel/order_details.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_api_search_orders(request):
    """Dynamic order search for admin navbar (order ID / number)."""
    from django.http import JsonResponse
    from django.db.models import Q
    from django.urls import reverse

    query = (request.GET.get('q') or '').strip()
    if not query:
        return JsonResponse({'results': []})

    orders = Order.objects.select_related('user').filter(
        Q(order_number__icontains=query)
    )

    if query.isdigit():
        orders = orders | Order.objects.select_related('user').filter(id=int(query))

    orders = orders.order_by('-created_at')[:10]

    results = []
    for order in orders:
        customer_name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
        results.append({
            'id': order.id,
            'order_number': order.order_number,
            'customer': customer_name,
            'total_amount': float(order.total_amount),
            'order_status': order.order_status,
            'payment_status': order.payment_status,
            'created_at': order.created_at.strftime('%d %b %Y, %I:%M %p'),
            'detail_url': reverse('admin_order_details', args=[order.id]),
        })

    return JsonResponse({'results': results})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_api_search_orders(request):
    """Dynamic order search for admin navbar (order ID / number)."""
    from django.http import JsonResponse
    from django.db.models import Q
    from django.urls import reverse

    query = (request.GET.get('q') or '').strip()
    if not query:
        return JsonResponse({'results': []})

    orders = Order.objects.select_related('user').filter(
        Q(order_number__icontains=query)
    )

    if query.isdigit():
        orders = orders | Order.objects.select_related('user').filter(id=int(query))

    orders = orders.order_by('-created_at')[:10]

    results = []
    for order in orders:
        customer_name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
        results.append({
            'id': order.id,
            'order_number': order.order_number,
            'customer': customer_name,
            'total_amount': float(order.total_amount),
            'order_status': order.order_status,
            'payment_status': order.payment_status,
            'created_at': order.created_at.strftime('%d %b %Y, %I:%M %p'),
            'detail_url': reverse('admin_order_details', args=[order.id]),
        })

    return JsonResponse({'results': results})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_approve_order(request, order_id):
    """Approve a pending order"""
    order = get_object_or_404(Order, id=order_id)
    
    # Log current status
    logger.info(f"Approve Order #{order.order_number} - Current approval_status: {order.approval_status}")
    
    if order.approval_status != 'PENDING_APPROVAL':
        messages.warning(request, f'Order is not pending approval. Current status: {order.approval_status}')
        return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))
    
    # Update approval status
    from django.utils import timezone
    order.approval_status = 'APPROVED'
    order.approved_by = request.user
    order.approved_at = timezone.now()
    order.order_status = 'PROCESSING'
    
    # Add approval notes if provided
    if request.method == 'POST':
        notes = request.POST.get('approval_notes', '')
        if notes:
            order.approval_notes = notes
    else:
        order.approval_notes = f'Manually approved by {request.user.username}'
    
    order.save()
    
    logger.info(f"Order approved! New approval_status: {order.approval_status}, order_status: {order.order_status}")
    
    # Send approval email to customer using proper email utility
    from .email_utils import send_order_approval_email
    email_sent = send_order_approval_email(order, request=request, approved_by=request.user)
    
    if email_sent:
        logger.info(f"Approval email notification queued for order {order.order_number}")
    else:
        logger.warning(f"Failed to send approval email for order {order.order_number}. Check EmailLog for details.")
    
    messages.success(request, f'Order {order.order_number} approved successfully!')
    return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reject_order(request, order_id):
    """Reject a pending order"""
    order = get_object_or_404(Order, id=order_id)
    
    if order.approval_status != 'PENDING_APPROVAL':
        messages.warning(request, 'Order is not pending approval.')
        return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))
    
    # Get rejection reason from POST
    if request.method == 'POST':
        rejection_reason = request.POST.get('rejection_reason', '')
        additional_notes = request.POST.get('additional_notes', '')
        
        if not rejection_reason:
            messages.error(request, 'Please select a rejection reason.')
            return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))
        
        # Combine reason and notes
        full_rejection_message = rejection_reason
        if additional_notes:
            full_rejection_message += f"\n\nAdditional Notes: {additional_notes}"
    else:
        full_rejection_message = 'Order rejected by admin'
    
    # Update approval status
    from django.utils import timezone
    order.approval_status = 'REJECTED'
    order.approved_by = request.user
    order.approved_at = timezone.now()
    order.order_status = 'CANCELLED'
    order.approval_notes = f'Rejected by {request.user.username}.\n\nReason: {full_rejection_message}'
    order.save()
    
    # Send rejection email to customer
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        
        subject = f'Order Rejected - {order.order_number}'
        
        # Format reason for email (remove emojis)
        reason_display = rejection_reason
        if '🚫' in rejection_reason or '📦' in rejection_reason or '🚨' in rejection_reason:
            reason_display = rejection_reason.split(' ', 1)[-1] if ' ' in rejection_reason else rejection_reason
        
        # Render HTML email
        html_content = render_to_string('emails/order_rejected.html', {
            'order': order,
            'rejection_reason': reason_display,
            'additional_notes': additional_notes if request.method == 'POST' else '',
        })
        
        # Plain text version
        text_content = f'''
Dear {order.user.get_full_name() or order.user.username},

We regret to inform you that your order {order.order_number} could not be processed and has been cancelled.

Reason: {reason_display}
{f"Additional Information: {additional_notes}" if additional_notes else ""}

Order Details:
- Order Number: {order.order_number}
- Total Amount: ₹{order.total_amount}
- Status: Cancelled

If you have any questions or concerns, please contact our customer support.

We apologize for any inconvenience caused and hope to serve you better in the future.

Thank you for your understanding.

Best regards,
FashioHub Team
        '''
        
        # Send email
        email = EmailMultiAlternatives(subject, text_content, _get_from_email(), [order.user.email])
        email.attach_alternative(html_content, "text/html")
        email.send()
        
    except Exception:
        logger.exception("Email sending failed while rejecting order_id=%s", order.id)
    
    messages.warning(request, f'Order {order.order_number} rejected. Reason: {rejection_reason}')
    return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_order(request, order_id):
    """Delete an order permanently"""
    order = get_object_or_404(Order, id=order_id)
    
    # Store order details for message
    order_number = order.order_number
    customer_name = f"{order.user.first_name} {order.user.last_name}".strip() or order.user.username
    
    # Delete the order (this will cascade delete related OrderItems and OrderStatusHistory)
    order.delete()
    _cleanup_orphan_orders()
    
    messages.success(request, f'Order {order_number} by {customer_name} has been permanently deleted.')
    return redirect(request.META.get('HTTP_REFERER', 'admin_orders'))


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_invoices(request):
    """Admin Invoices List - Using Orders as Invoices"""
    all_invoices = Order.objects.select_related('user').prefetch_related('items').order_by('-created_at')
    
    # Status filter
    status_filter = request.GET.get('status', '')
    search_query = request.GET.get('search', '')
    
    # Filter by status if provided
    if status_filter:
        if status_filter == 'PAID':
            all_invoices = all_invoices.filter(payment_status='PAID')
        elif status_filter == 'PENDING':
            all_invoices = all_invoices.filter(payment_status='PENDING')
        elif status_filter == 'UNPAID':
            all_invoices = all_invoices.filter(payment_status='FAILED')
    
    # Filter by search query
    if search_query:
        all_invoices = all_invoices.filter(
            Q(order_number__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Get statistics
    total_invoices = Order.objects.count()
    total_clients = Order.objects.values('user').distinct().count()
    total_paid = Order.objects.filter(payment_status='PAID').aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    total_unpaid = Order.objects.filter(payment_status__in=['PENDING', 'FAILED']).aggregate(Sum('total_amount'))['total_amount__sum'] or 0
    
    # Pagination
    items_per_page = request.GET.get('per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except (ValueError, TypeError):
        items_per_page = 10
    
    paginator = Paginator(all_invoices, items_per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    # Build query string for pagination
    query_params = {}
    if search_query:
        query_params['search'] = search_query
    if status_filter:
        query_params['status'] = status_filter
    
    query_string = '&' + urlencode(query_params) if query_params else ''
    
    context = {
        'invoices': page_obj.object_list,
        'page_obj': page_obj,
        'total_invoices': total_invoices,
        'total_clients': total_clients,
        'total_paid': total_paid,
        'total_unpaid': total_unpaid,
        'total_count': all_invoices.count(),
        'status_filter': status_filter,
        'search_query': search_query,
        'items_per_page': items_per_page,
        'query_string': query_string,
    }
    
    return render(request, 'admin_panel/invoices.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_invoice_inventory(request):
    """Unified Invoice & Inventory management dashboard for admins"""

    invoices_qs = (
        Order.objects
        .select_related('user')
        .prefetch_related('items')
        .order_by('-created_at')
    )

    status_filter = request.GET.get('status', 'all')
    payment_filter = request.GET.get('payment', 'all')
    search_query = request.GET.get('search', '').strip()
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if status_filter and status_filter.lower() != 'all':
        invoices_qs = invoices_qs.filter(order_status=status_filter)

    if payment_filter and payment_filter.lower() != 'all':
        invoices_qs = invoices_qs.filter(payment_status=payment_filter)

    if search_query:
        invoices_qs = invoices_qs.filter(
            Q(order_number__icontains=search_query)
            | Q(user__username__icontains=search_query)
            | Q(user__email__icontains=search_query)
            | Q(user__first_name__icontains=search_query)
            | Q(user__last_name__icontains=search_query)
        )

    if date_from:
        invoices_qs = invoices_qs.filter(created_at__date__gte=date_from)

    if date_to:
        invoices_qs = invoices_qs.filter(created_at__date__lte=date_to)

    items_per_page = request.GET.get('per_page', 10)
    try:
        items_per_page = int(items_per_page)
    except (TypeError, ValueError):
        items_per_page = 10

    paginator = Paginator(invoices_qs, items_per_page)
    page_number = request.GET.get('page', 1)

    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    paid_orders_qs = Order.objects.filter(payment_status='PAID')
    outstanding_orders_qs = Order.objects.filter(payment_status__in=['PENDING', 'FAILED'])

    total_revenue = paid_orders_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    outstanding_amount = outstanding_orders_qs.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    average_order_value = paid_orders_qs.aggregate(avg=Avg('total_amount'))['avg'] or Decimal('0.00')
    overdue_invoices = outstanding_orders_qs.filter(order_status__in=['PENDING', 'PROCESSING']).count()

    low_stock_threshold = 10
    critical_stock_threshold = 3

    inventory_value_expr = ExpressionWrapper(
        F('stock') * F('price'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )

    active_products_qs = Product.objects.filter(is_active=True)

    total_skus = active_products_qs.count()
    total_units = active_products_qs.aggregate(total=Sum('stock'))['total'] or 0
    total_inventory_value = active_products_qs.annotate(value=inventory_value_expr).aggregate(total=Sum('value'))['total'] or Decimal('0.00')

    low_stock_qs = (
        active_products_qs
        .filter(stock__gt=0, stock__lte=low_stock_threshold)
        .annotate(value=inventory_value_expr)
        .order_by('stock')
    )
    low_stock_products = list(low_stock_qs[:12])

    out_of_stock_qs = (
        Product.objects
        .filter(stock__lte=0)
        .annotate(value=inventory_value_expr)
        .order_by('name')
    )
    out_of_stock_products = list(out_of_stock_qs[:12])

    low_stock_count = low_stock_qs.count()
    out_of_stock_count = out_of_stock_qs.count()

    top_movers = list(
        active_products_qs
        .annotate(total_sold=Sum('orderitem__quantity'))
        .filter(total_sold__gt=0)
        .order_by('-total_sold')[:8]
    )

    recent_restock = []
    if any(field.name == 'updated_at' for field in Product._meta.get_fields()):
        recent_restock = list(
            active_products_qs
            .filter(updated_at__isnull=False)
            .order_by('-updated_at')[:6]
        )

    # Inventory manager list (all products)
    inventory_qs = Product.objects.all().order_by('name')
    product_search_query = request.GET.get('product_search', '').strip()

    if product_search_query:
        inventory_qs = inventory_qs.filter(
            Q(name__icontains=product_search_query)
            | Q(sku__icontains=product_search_query)
            | Q(brand__icontains=product_search_query)
        )

    product_items_per_page = request.GET.get('product_per_page', 15)
    try:
        product_items_per_page = int(product_items_per_page)
    except (TypeError, ValueError):
        product_items_per_page = 15

    product_paginator = Paginator(inventory_qs, product_items_per_page)
    product_page_number = request.GET.get('product_page', 1)

    try:
        product_page_obj = product_paginator.page(product_page_number)
    except PageNotAnInteger:
        product_page_obj = product_paginator.page(1)
    except EmptyPage:
        product_page_obj = product_paginator.page(product_paginator.num_pages)

    query_params = request.GET.copy()
    query_params.pop('page', None)
    querystring = query_params.urlencode()

    product_query_params = request.GET.copy()
    product_query_params.pop('product_page', None)
    product_querystring = product_query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'invoices': page_obj.object_list,
        'items_per_page': items_per_page,
        'status_filter': status_filter,
        'payment_filter': payment_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'querystring': querystring,
        'product_querystring': product_querystring,
        'current_path': request.get_full_path(),

        'invoice_total_count': invoices_qs.count(),
        'paid_invoice_count': paid_orders_qs.count(),
        'outstanding_invoice_count': outstanding_orders_qs.count(),
        'total_revenue': total_revenue,
        'outstanding_amount': outstanding_amount,
        'average_order_value': average_order_value,
        'overdue_invoices': overdue_invoices,

        'low_stock_products': low_stock_products,
        'out_of_stock_products': out_of_stock_products,
        'low_stock_threshold': low_stock_threshold,
        'critical_stock_threshold': critical_stock_threshold,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
        'total_skus': total_skus,
        'total_units': total_units,
        'total_inventory_value': total_inventory_value,
        'top_movers': top_movers,
        'recent_restock': recent_restock,
        'status_choices': Order.ORDER_STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_STATUS_CHOICES,
        'product_page_obj': product_page_obj,
        'inventory_products': product_page_obj.object_list,
        'product_items_per_page': product_items_per_page,
        'product_search_query': product_search_query,
    }

    return render(request, 'admin_panel/invoice_inventory.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_update_inventory(request):
    """Adjust product stock from the Inventory dashboard."""
    product_id = request.POST.get('product_id')
    action = request.POST.get('action', 'save_stock')
    redirect_url = request.POST.get('next') or reverse('admin_invoice_inventory')

    try:
        product = Product.objects.get(id=product_id)
    except Product.DoesNotExist:
        messages.error(request, 'Product not found')
        return redirect(redirect_url)

    try:
        previous_stock = product.stock
        if action == 'mark_out':
            product.stock = 0
            # Keep the product visible with an out-of-stock tag
            product.is_active = True
            product.save(update_fields=['stock', 'is_active'])
            messages.success(request, f'Marked {product.name} as out of stock (visible in shop).')
        else:
            raw_stock = request.POST.get('stock', '0')
            new_stock = max(0, int(raw_stock))
            product.stock = new_stock

            # Reactivate when restocked; keep visible even at 0
            product.is_active = True if new_stock >= 0 else product.is_active
            product.save(update_fields=['stock', 'is_active'])
            messages.success(request, f'Updated stock for {product.name} to {new_stock}.')

        # Notify subscribers when restocked from zero
        if previous_stock <= 0 and product.stock > 0:
            notifications = ProductStockNotification.objects.filter(product=product, is_sent=False)
            sent_count = 0
            failed_count = 0
            last_error = ''
            for note in notifications:
                try:
                    send_mail(
                        subject=f"{product.name} is back in stock",
                        message=f"Good news! {product.name} is available again. Visit the product page to purchase.",
                        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
                        recipient_list=[note.email],
                        fail_silently=False,
                    )
                    note.mark_sent()
                    sent_count += 1
                except Exception as exc:
                    failed_count += 1
                    last_error = str(exc)
            if notifications.exists():
                messages.info(request, f"Restock notifications attempted: sent {sent_count}, failed {failed_count}.")
                if failed_count and last_error:
                    messages.warning(request, f"Email error: {last_error}")
            else:
                messages.info(request, "No pending restock notifications for this product.")
    except ValueError:
        messages.error(request, 'Invalid stock value')
    except Exception as exc:
        messages.error(request, f'Could not update stock: {exc}')

    return redirect(redirect_url)


@login_required
def admin_customers(request):
    """Admin view to manage customers"""
    if not request.user.is_staff:
        messages.error(request, 'Unauthorized access')
        return redirect('index')
    
    # Handle POST requests (block/unblock, segment update)
    if request.method == 'POST':
        action = request.POST.get('action')
        customer_ids = request.POST.getlist('customer_ids')
        
        if action and customer_ids:
            users = User.objects.filter(id__in=customer_ids)
            
            if action == 'block':
                UserProfile.objects.filter(user__in=users).update(is_blocked=True)
                messages.success(request, f'Successfully blocked {len(customer_ids)} customer(s)')
            elif action == 'unblock':
                UserProfile.objects.filter(user__in=users).update(is_blocked=False)
                messages.success(request, f'Successfully unblocked {len(customer_ids)} customer(s)')
            elif action in ['NEW', 'REGULAR', 'VIP', 'ADMIN']:
                UserProfile.objects.filter(user__in=users).update(customer_segment=action)
                users.update(is_staff=(action == 'ADMIN'))
                messages.success(request, f'Successfully updated {len(customer_ids)} customer(s) to {action}')
            elif action == 'delete':
                users.delete()
                messages.success(request, f'Successfully deleted {len(customer_ids)} customer(s)')
            
            return redirect('admin_customers')
    
    # Get filter parameters
    segment_filter = request.GET.get('segment', 'all')
    status_filter = request.GET.get('status', 'all')
    search_query = request.GET.get('search', '')
    
    # Base queryset - include all users
    customers = User.objects.all().select_related('userprofile')
    
    # Apply filters
    if segment_filter != 'all':
        customers = customers.filter(userprofile__customer_segment=segment_filter)
    
    if status_filter == 'blocked':
        customers = customers.filter(userprofile__is_blocked=True)
    elif status_filter == 'active':
        customers = customers.filter(userprofile__is_blocked=False)
    
    if search_query:
        customers = customers.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    # Always show ADMIN segment users first
    customers = customers.annotate(
        admin_order=Case(
            When(userprofile__customer_segment='ADMIN', then=Value(0)),
            default=Value(1),
            output_field=IntegerField(),
        )
    ).order_by('admin_order', 'id')
    
    # Calculate stats
    total_customers = User.objects.all().count()  # Include all users
    active_customers = UserProfile.objects.filter(is_blocked=False).count()
    blocked_customers = UserProfile.objects.filter(is_blocked=True).count()
    new_customers = UserProfile.objects.filter(customer_segment='NEW').count()
    regular_customers = UserProfile.objects.filter(customer_segment='REGULAR').count()
    vip_customers = UserProfile.objects.filter(customer_segment='VIP').count()
    admin_customers = UserProfile.objects.filter(customer_segment='ADMIN').count()
    
    # Pagination
    paginator = Paginator(customers, 20)
    page_number = request.GET.get('page', 1)
    
    try:
        customers_page = paginator.page(page_number)
    except PageNotAnInteger:
        customers_page = paginator.page(1)
    except EmptyPage:
        customers_page = paginator.page(paginator.num_pages)
    
    context = {
        'customers': customers_page,
        'total_customers': total_customers,
        'active_customers': active_customers,
        'blocked_customers': blocked_customers,
        'new_customers': new_customers,
        'regular_customers': regular_customers,
        'vip_customers': vip_customers,
        'admin_customers': admin_customers,
        'segment_filter': segment_filter,
        'status_filter': status_filter,
        'search_query': search_query,
    }
    return render(request, 'admin_panel/customers.html', context)


@login_required
def admin_customer_details(request, customer_id):
    """Admin view to see comprehensive customer details with real-time data"""
    if not request.user.is_staff:
        messages.error(request, 'Unauthorized access')
        return redirect('index')
    
    customer = get_object_or_404(User, id=customer_id)
    profile, created = UserProfile.objects.get_or_create(user=customer)
    
    # Handle POST requests (update segment, block/unblock, admin role)
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'toggle_block':
            profile.is_blocked = not profile.is_blocked
            profile.save()
            status = 'blocked' if profile.is_blocked else 'unblocked'
            messages.success(request, f'Customer {status} successfully')
        elif action == 'update_segment':
            new_segment = request.POST.get('segment')
            if new_segment in ['NEW', 'REGULAR', 'VIP', 'ADMIN']:
                profile.customer_segment = new_segment
                customer.is_staff = (new_segment == 'ADMIN')
                customer.save()
                profile.save()
                messages.success(request, f'Customer segment updated to {new_segment}')
        elif action == 'toggle_admin':
            customer.is_staff = not customer.is_staff
            customer.save()
            if customer.is_staff:
                profile.customer_segment = 'ADMIN'
            else:
                profile.customer_segment = 'REGULAR'
            profile.save()
            status = 'granted' if customer.is_staff else 'revoked'
            messages.success(request, f'Admin access {status} successfully')
        
        return redirect('admin_customer_details', customer_id=customer_id)
    
    # Wishlist data
    wishlist_items = Wishlist.objects.filter(user=customer).select_related('product')
    total_wishlist = wishlist_items.count()
    
    # Cart data
    cart_items = Cart.objects.filter(user=customer).select_related('product')
    total_cart_value = sum(item.get_total_price() for item in cart_items)
    
    # Reviews data
    reviews = ProductReview.objects.filter(user=customer).select_related('product').order_by('-created_at')
    total_reviews = reviews.count()
    approved_reviews = reviews.filter(is_approved=True).count()
    
    # Orders data
    orders = Order.objects.filter(user=customer).order_by('-created_at')
    total_orders = orders.count()
    
    # Calculate total spent (from PAID orders)
    paid_orders = orders.filter(payment_status='PAID')
    total_spent = paid_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0.00')
    
    # Update profile total_spent
    if profile.total_spent != total_spent:
        profile.total_spent = total_spent
        profile.save(update_fields=['total_spent'])
    
    # Calculate profit (margin-based)
    profit_orders_qs = _paid_orders_qs().filter(user=customer)
    profit_items = _order_item_profit(OrderItem.objects.filter(order__in=profit_orders_qs))
    total_profit = profit_items + _return_fee_profit(user=customer)
    
    # Order statistics
    pending_orders = orders.filter(order_status='PENDING').count()
    processing_orders = orders.filter(order_status='PROCESSING').count()
    shipped_orders = orders.filter(order_status='SHIPPED').count()
    delivered_orders = orders.filter(order_status='DELIVERED').count()
    cancelled_orders = orders.filter(order_status='CANCELLED').count()
    
    # Payment statistics
    paid_orders_count = orders.filter(payment_status='PAID').count()
    pending_payments = orders.filter(payment_status='PENDING').count()
    
    # Products purchased
    order_items = OrderItem.objects.filter(order__user=customer).select_related('product', 'order')
    unique_products = order_items.values('product').distinct().count()
    total_items_purchased = order_items.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Top purchased products
    from collections import Counter
    product_purchases = Counter()
    for item in order_items:
        if item.product:
            product_purchases[item.product] += item.quantity
    top_products = [{'product': product, 'count': count} for product, count in product_purchases.most_common(5)]
    
    # Recent activity logs - Build comprehensive activity timeline
    from datetime import timedelta
    recent_activities = []
    
    # Add all individual wishlist additions
    for wishlist_item in wishlist_items.order_by('-added_at')[:10]:
        recent_activities.append({
            'icon': 'bx bx-heart',
            'color': 'danger',
            'title': 'Added to Wishlist',
            'description': f'{wishlist_item.product.name}',
            'timestamp': wishlist_item.added_at
        })
    
    # Add all individual cart additions
    for cart_item in cart_items.order_by('-added_at')[:10]:
        recent_activities.append({
            'icon': 'bx bx-cart-add',
            'color': 'info',
            'title': 'Added to Cart',
            'description': f'{cart_item.product.name} (Qty: {cart_item.quantity})',
            'timestamp': cart_item.added_at
        })
    
    # Add all individual reviews
    for review in reviews[:10]:
        recent_activities.append({
            'icon': 'bx bx-message-square-detail',
            'color': 'warning',
            'title': 'Submitted Review',
            'description': f'{review.product.name} - {review.rating} stars {"✓ Approved" if review.is_approved else "⏳ Pending"}',
            'timestamp': review.created_at
        })
    
    # Add all orders with detailed status
    for order in orders[:15]:
        status_icons = {
            'PENDING': 'bx bx-time-five',
            'PROCESSING': 'bx bx-loader-alt',
            'SHIPPED': 'bx bx-package',
            'DELIVERED': 'bx bx-check-circle',
            'CANCELLED': 'bx bx-x-circle',
        }
        status_colors = {
            'PENDING': 'warning',
            'PROCESSING': 'info',
            'SHIPPED': 'primary',
            'DELIVERED': 'success',
            'CANCELLED': 'danger',
        }
        recent_activities.append({
            'icon': status_icons.get(order.order_status, 'bx bx-shopping-bag'),
            'color': status_colors.get(order.order_status, 'primary'),
            'title': f'Order #{order.order_number}',
            'description': f'₹{order.total_amount} - {order.get_order_status_display()} ({order.get_payment_status_display()})',
            'timestamp': order.created_at
        })
    
    # Account created
    recent_activities.append({
        'icon': 'bx bx-user-plus',
        'color': 'success',
        'title': 'Account Created',
        'description': 'Joined VibeMall',
        'timestamp': customer.date_joined
    })
    
    # Last login/activity
    if profile.last_activity and profile.last_activity > customer.date_joined:
        recent_activities.append({
            'icon': 'bx bx-log-in',
            'color': 'info',
            'title': 'Last Activity',
            'description': 'User active on site',
            'timestamp': profile.last_activity
        })
    
    # Sort activities by timestamp (newest first)
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:25]  # Limit to 25 most recent
    
    context = {
        'customer': customer,
        'profile': profile,
        
        # Stats cards
        'total_spent': total_spent,
        'total_wishlist': total_wishlist,
        'total_cart_value': total_cart_value,
        'total_reviews': total_reviews,
        'approved_reviews': approved_reviews,
        
        # Tab content
        'wishlist_items': wishlist_items,
        'cart_items': cart_items,
        'reviews': reviews,
        'recent_activities': recent_activities,
        
        # Orders & Profit section
        'orders': orders[:10],  # Latest 10 orders
        'total_orders': total_orders,
        'total_profit': total_profit,
        'pending_orders': pending_orders,
        'processing_orders': processing_orders,
        'shipped_orders': shipped_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'paid_orders_count': paid_orders_count,
        'pending_payments': pending_payments,
        
        # Products purchased
        'unique_products': unique_products,
        'total_items_purchased': total_items_purchased,
        'top_products': top_products,
        'order_items': order_items[:10],  # Latest 10 items
        
        # Segment choices for dropdown
        'segment_choices': UserProfile.CUSTOMER_SEGMENT_CHOICES,
    }
    return render(request, 'admin_panel/customer_details.html', context)


# ===== PUBLIC VIEWS =====

def spa_home(request):
    """Render the new React/Vite-style home page (integrated from vibemall---home/dist)."""
    return render(request, 'vibemall_spa_home.html')


def index(request):
    homepage_cache_key = 'homepage_public_context_v2'
    public_context = cache.get(homepage_cache_key)

    if public_context is None:
        sliders = list(Slider.objects.filter(is_active=True).order_by('order', 'id'))
        features = list(Feature.objects.filter(is_active=True).order_by('order', 'id'))
        banners = list(Banner.objects.filter(is_active=True).order_by('order', 'id'))

        categories = list(CategoryIcon.objects.filter(is_active=True).order_by('order', 'id'))
        if not categories:
            excluded_categories = {'TOP_DEALS', 'TOP_SELLING', 'TOP_FEATURED', 'RECOMMENDED'}
            categories = [
                {
                    'name': label,
                    'category_key': key,
                    'background_gradient': 'linear-gradient(135deg, #e0f7ff 0%, #b3e5fc 100%)',
                    'card_image': None,
                    'icon_image': None,
                    'icon_class': None,
                    'icon_color': '#0288d1',
                    'icon_size': 48,
                }
                for key, label in Product.CATEGORY_CHOICES
                if key not in excluded_categories
            ]

        category_fallback_images = {}
        category_keys = []
        for category in categories:
            if hasattr(category, 'category_key'):
                key = (category.category_key or '').strip()
            else:
                key = (category.get('category_key') or '').strip()
            if key and key not in category_keys:
                category_keys.append(key)

        if category_keys:
            fallback_products = (
                Product.objects
                .filter(is_active=True, category__in=category_keys)
                .exclude(image='')
                .only('category', 'image')
                .order_by('category', '-id')
            )
            for product in fallback_products:
                key = (product.category or '').strip()
                if key and key not in category_fallback_images and product.image:
                    category_fallback_images[key] = product.image.url

        from .models import MainPageProduct, Reel, MainPageSubCategoryBanner, MainPageBanner

        category_buckets = {
            'category1': [],
            'category2': [],
            'category3': [],
            'category4': [],
        }
        main_page_products = list(
            MainPageProduct.objects
            .filter(category__in=tuple(category_buckets.keys()))
            .select_related('product')
            .order_by('category', 'order', 'id')
        )
        for item in main_page_products:
            bucket = category_buckets.get(item.category)
            if bucket is not None and len(bucket) < 10:
                bucket.append(item.product)

        latest_products = list(Product.objects.filter(is_active=True).order_by('-id')[:10])
        top_deals = category_buckets['category1'] or latest_products
        top_selling = category_buckets['category2'] or latest_products
        top_featured = category_buckets['category3'] or latest_products
        recommended = category_buckets['category4'] or latest_products

        subcategory_banner_records = list(
            MainPageSubCategoryBanner.objects
            .filter(is_active=True, sub_category__is_active=True)
            .select_related('sub_category')
            .order_by('order', 'id')
        )
        subcategory_banners = [
            {
                'title': item.display_title,
                'image_url': item.image.url if item.image else '',
                'category_key': item.sub_category.category_key,
                'sub_category': item.sub_category.name,
            }
            for item in subcategory_banner_records
        ]

        banner_records = list(
            MainPageBanner.objects
            .filter(banner_area__in=['first', 'second'], is_active=True)
            .order_by('banner_area', 'order', 'id')
        )
        first_banners = []
        second_banners = []
        for item in banner_records:
            banner_data = {
                'badge_text': item.badge_text,
                'title': item.title,
                'description': item.description,
                'image_url': item.image.url if item.image else '',
                'link_url': item.link_url,
            }
            if item.banner_area == 'first' and len(first_banners) < 3:
                first_banners.append(banner_data)
            elif item.banner_area == 'second' and len(second_banners) < 2:
                second_banners.append(banner_data)

        product_ids = {p.id for p in top_deals + top_selling + top_featured + recommended}
        stats_qs = (
            ProductReview.objects
            .filter(product_id__in=product_ids, is_approved=True)
            .values('product_id')
            .annotate(avg_rating=Avg('rating'), review_count=Count('id'))
        )
        product_stats = {
            item['product_id']: {
                'avg_rating': round(item['avg_rating'] or 0, 1),
                'review_count': item['review_count'],
            }
            for item in stats_qs
        }

        countdown = DealCountdown.objects.filter(is_active=True).first()
        watch_shop_reels = list(
            Reel.objects
            .filter(
                is_published=True,
                product__isnull=False,
                product__is_active=True,
                video_file__isnull=False,
            )
            .exclude(video_file='')
            .select_related('product')
            .order_by('order', '-created_at')[:6]
        )

        public_context = {
            'sliders': sliders,
            'features': features,
            'banners': banners,
            'categories': categories,
            'category_fallback_images': category_fallback_images,
            'top_deals': top_deals,
            'top_selling': top_selling,
            'top_featured': top_featured,
            'recommended': recommended,
            'subcategory_banners': subcategory_banners,
            'first_banners': first_banners,
            'second_banners': second_banners,
            'countdown': countdown,
            'watch_shop_reels': watch_shop_reels,
            'product_stats': product_stats,
        }
        cache.set(homepage_cache_key, public_context, 300)

    liked_reel_ids = []
    for raw_id in request.session.get('reel_liked_ids', []):
        try:
            liked_reel_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    # Get wishlist product IDs for logged-in user
    wishlist_product_ids = []
    cart_product_ids = []
    delivered_orders = []
    
    current_user = getattr(request, 'user', None)
    if current_user and current_user.is_authenticated:
        wishlist_product_ids = list(Wishlist.objects.filter(user=current_user).values_list('product_id', flat=True))
        cart_product_ids = list(Cart.objects.filter(user=current_user).values_list('product_id', flat=True))
        
        # Check for recently delivered orders (delivered in last 7 days without reviews)
        from datetime import timedelta
        from django.utils import timezone
        seven_days_ago = timezone.now() - timedelta(days=7)
        
        # Get delivered orders that don't have review notification shown
        shown_review_notifications = request.session.get('shown_review_notifications', [])
        
        delivered_orders = list(Order.objects.filter(
            user=current_user,
            order_status='DELIVERED',
            delivery_date__gte=seven_days_ago
        ).exclude(
            order_number__in=shown_review_notifications
        ).prefetch_related('items__product')[:3])  # Show up to 3 recent delivered orders
        
        # Mark these orders as shown
        if delivered_orders:
            for order in delivered_orders:
                shown_review_notifications.append(order.order_number)
            request.session['shown_review_notifications'] = shown_review_notifications
            request.session.modified = True

    context = dict(public_context)
    context.update({
        'wishlist_product_ids': wishlist_product_ids,
        'cart_product_ids': cart_product_ids,
        'liked_reel_ids': liked_reel_ids,
        'delivered_orders': delivered_orders,
    })

    return render(request, 'index.html', context)


@require_POST
def reel_track_view(request, reel_id):
    """Track reel view once per session and return latest count."""
    from Hub.models import Reel

    reel = get_object_or_404(Reel, id=reel_id, is_published=True)

    viewed_ids = set()
    for raw_id in request.session.get('reel_viewed_ids', []):
        try:
            viewed_ids.add(int(raw_id))
        except (TypeError, ValueError):
            continue

    if reel_id not in viewed_ids:
        Reel.objects.filter(id=reel.id).update(view_count=F('view_count') + 1)
        viewed_ids.add(reel_id)
        request.session['reel_viewed_ids'] = sorted(viewed_ids)
        request.session.modified = True

    reel.refresh_from_db(fields=['view_count'])
    return JsonResponse({
        'success': True,
        'view_count': int(reel.view_count or 0),
    })


@require_POST
def reel_set_like(request, reel_id):
    """Set like status for a reel (per session) and return latest like count."""
    from Hub.models import Reel

    reel = get_object_or_404(Reel, id=reel_id, is_published=True)

    liked_param = str(request.POST.get('liked', '')).strip().lower()
    if liked_param not in {'1', '0', 'true', 'false', 'on', 'off', 'yes', 'no'}:
        return JsonResponse({
            'success': False,
            'message': 'Invalid like state',
        }, status=400)

    should_like = liked_param in {'1', 'true', 'on', 'yes'}

    liked_ids = set()
    for raw_id in request.session.get('reel_liked_ids', []):
        try:
            liked_ids.add(int(raw_id))
        except (TypeError, ValueError):
            continue

    already_liked = reel_id in liked_ids

    if should_like and not already_liked:
        Reel.objects.filter(id=reel.id).update(like_count=F('like_count') + 1)
        liked_ids.add(reel_id)
    elif not should_like and already_liked:
        Reel.objects.filter(id=reel.id).update(
            like_count=Case(
                When(like_count__gt=0, then=F('like_count') - 1),
                default=Value(0),
                output_field=IntegerField(),
            )
        )
        liked_ids.remove(reel_id)

    request.session['reel_liked_ids'] = sorted(liked_ids)
    request.session.modified = True

    reel.refresh_from_db(fields=['like_count'])
    return JsonResponse({
        'success': True,
        'liked': should_like,
        'like_count': int(reel.like_count or 0),
    })










def about(request):
    site_settings = SiteSettings.get_settings()

    product_count = Product.objects.filter(is_active=True).count()
    region_count = Address.objects.exclude(country__isnull=True).exclude(country__exact='').values('country').distinct().count()
    avg_rating = ProductReview.objects.filter(is_approved=True).aggregate(value=Avg('rating')).get('value') or 0

    if product_count >= 1000:
        products_display = f"{product_count / 1000:.1f}k+"
    else:
        products_display = str(product_count or 0)

    context = {
        'site_settings': site_settings,
        'about_subtitle': (
            site_settings.tagline
            if site_settings and site_settings.tagline
            else 'VibeMall blends trend-aware shopping, dependable service, and a modern marketplace feel into one destination made for everyday discovery.'
        ),
        'about_hero': {
            'eyebrow': 'Our Heritage',
            'title': 'Curation as a Form of Art.',
            'line_1': 'VibeMall was founded on the belief that shopping should be a narrative experience. We curate collections that balance trend, utility, and long-term style value.',
            'line_2': 'Every piece in our marketplace is selected to make discovery easier, checkout smoother, and every order more satisfying.',
        },
        'about_stats': {
            'products': products_display,
            'regions': str(region_count or 0),
            'rating': f"{avg_rating:.1f}" if avg_rating else '0.0',
            'artisans': '450+',
            'clients': '85k',
            'awards': '14',
        },
        'about_concierge': {
            'title': 'The Concierge Service',
            'styling_title': 'Personal Styling',
            'styling_copy': 'Book a private digital session with our stylists to build looks that fit your routine and personality.',
            'delivery_title': 'White Glove Delivery',
            'delivery_copy': 'Priority shipping with secure and premium packaging.',
            'auth_title': 'Authenticity Promise',
            'auth_copy': 'Carefully sourced products from trusted brands and partners.',
        },
        'about_why': [
            {
                'index': '01',
                'title': 'Ethical Sourcing',
                'copy': 'We evaluate partners for quality standards, fair practices, and responsible sourcing.',
            },
            {
                'index': '02',
                'title': 'Curated Rarity',
                'copy': 'Exclusive drops and limited collections that keep the catalog fresh and exciting.',
            },
            {
                'index': '03',
                'title': 'Timeless Quality',
                'copy': 'Pieces selected for repeat wear, durability, and style relevance beyond a single season.',
            },
        ],
    }

    return render(request, 'about.html', context)
def blog(request): return render(request, 'blog.html')
def blog_details(request): return render(request, 'blog-details.html')


@require_POST
@csrf_exempt
def subscribe_newsletter(request):
    """Capture newsletter subscribers from CTA forms."""
    email = (request.POST.get('email') or '').strip().lower()
    source_page = (request.POST.get('source_page') or '').strip()[:120]
    is_ajax = (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'
    )

    redirect_to = request.META.get('HTTP_REFERER') or reverse('index')
    parsed_redirect = urlparse(redirect_to)
    if parsed_redirect.netloc and parsed_redirect.netloc != request.get_host():
        redirect_to = reverse('index')

    if not email:
        message = 'Please enter a valid email address.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': message}, status=400)
        messages.error(request, message)
        return redirect(redirect_to)

    try:
        validate_email(email)
    except ValidationError:
        message = 'Please enter a valid email address.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': message}, status=400)
        messages.error(request, message)
        return redirect(redirect_to)

    subscriber, created = NewsletterSubscription.objects.get_or_create(
        email=email,
        defaults={
            'source_page': source_page,
            'is_active': True,
        }
    )

    status = 'subscribed'
    if created:
        message = 'Thanks for subscribing. You will receive our latest offers by email.'
    elif subscriber.is_active:
        status = 'already_subscribed'
        message = 'This email is already subscribed to our newsletter.'
    else:
        status = 'resubscribed'
        subscriber.is_active = True
        subscriber.unsubscribed_at = None
        if source_page:
            subscriber.source_page = source_page
            subscriber.save(update_fields=['is_active', 'unsubscribed_at', 'source_page', 'updated_at'])
        else:
            subscriber.save(update_fields=['is_active', 'unsubscribed_at', 'updated_at'])
        message = 'Welcome back. Your newsletter subscription is active again.'

    # Send a welcome email when the user subscribes or resubscribes
    if created or status == 'resubscribed':
        from_email = getattr(settings, 'DEFAULT_FROM_EMAIL', 'no-reply@vibemall.com')
        subject = 'Welcome to VibeMall!'
        plain_message = 'Thank you for joining VibeMall. We will notify you when we are live.'
        try:
            send_mail(
                subject,
                plain_message,
                from_email,
                [email],
                fail_silently=True,
            )
        except Exception as e:
            logger.exception('Failed to send newsletter welcome email to %s: %s', email, e)

    if is_ajax:
        return JsonResponse({
            'success': True,
            'status': status,
            'message': message,
        })

    if status == 'subscribed':
        messages.success(request, message)
    elif status == 'already_subscribed':
        messages.info(request, message)
    else:
        messages.success(request, message)

    return redirect(redirect_to)


@login_required(login_url='login')
def cart(request: HttpRequest) -> HttpResponse:
    cart_items = Cart.objects.filter(user=request.user).select_related('product')
    total_price = sum(item.get_total_price() for item in cart_items)
    cart_count = cart_items.count()
    
    return render(request, 'cart.html', {
        'cart_items': cart_items,
        'total_price': total_price,
        'cart_count': cart_count
    })



def buy_now(request, product_id):
    """Buy Now - Redirect directly to checkout with product"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        quantity = int(request.POST.get('quantity', 1))
        
        if quantity < 1:
            return JsonResponse({
                'success': False,
                'message': 'Invalid quantity'
            }, status=400)
        
        # Check stock
        if product.stock < quantity:
            return JsonResponse({
                'success': False,
                'message': f'Only {product.stock} items available in stock'
            }, status=400)
        
        # Store in session for checkout
        request.session['buy_now_item'] = {
            'product_id': product.id,
            'quantity': quantity,
            'price': str(product.price)
        }
        
        return JsonResponse({
            'success': True,
            'message': f'{product.name} added to checkout!',
            'redirect_url': '/checkout/'
        })
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)


@login_required(login_url='login')
def checkout(request: HttpRequest) -> HttpResponse:
    """Dynamic Checkout Page"""
    cart_items, buy_now_item, total_price = _get_checkout_items(request)
    base_cart_total = Decimal(str(total_price))
    total_item_qty = _get_checkout_total_quantity(cart_items)
    
    # Handle form submission
    if request.method == 'POST':
        # Get form data
        first_name = (request.POST.get('first_name') or '').strip()
        last_name = (request.POST.get('last_name') or '').strip()
        email = (request.POST.get('email') or '').strip()
        phone = (request.POST.get('phone') or '').strip()
        country = (request.POST.get('country') or '').strip()
        address = (request.POST.get('address') or '').strip()
        city = (request.POST.get('city') or '').strip()
        state = (request.POST.get('state') or '').strip()
        postcode = (request.POST.get('postcode') or '').strip()
        pincode_valid = request.POST.get('pincode_valid', '')
        order_type = request.POST.get('order_type', 'for_self')
        payment_method = request.POST.get('payment_method', 'COD')
        customer_notes = request.POST.get('customer_notes', '')
        is_resell = (order_type == 'for_resell') or (request.POST.get('is_resell') == 'on')
        user_selected_resell = is_resell
        explicit_self_selected = order_type == 'for_self'
        use_default_address = request.POST.get('use_default_address') == 'on'
        set_default_address = request.POST.get('set_default_address') == 'on'
        resell_margin_raw = normalize_decimal_input(request.POST.get('resell_margin', '0'))
        try:
            resell_margin_per_unit = Decimal(resell_margin_raw or '0')
        except Exception:
            resell_margin_per_unit = Decimal('0')
        
        # Loyalty points redemption
        redeem_points = request.POST.get('redeem_points') == 'on'
        points_to_redeem = 0
        if redeem_points:
            points_to_redeem = int(request.POST.get('points_to_redeem', 0))
        
        # Coupon handling
        coupon_id = request.POST.get('coupon_id', '')
        
        # Log selected payment method
        logger.info(f"Payment method selected: {payment_method}")
        if not payment_method or payment_method not in ['COD', 'RAZORPAY']:
            messages.error(request, 'Please select a valid payment method.')
            return redirect('checkout')
        
        # Resell FROM details
        from_name = (request.POST.get('from_name') or '').strip()
        from_phone = (request.POST.get('from_phone') or '').strip()

        if use_default_address:
            default_address = Address.objects.filter(user=request.user, is_default=True).first()
            if not default_address:
                messages.error(request, 'Default address not found. Please enter address manually.')
                return redirect('checkout')
            # Reusing an existing default address should not trigger new default-save intent.
            set_default_address = False
            first_name, last_name = _split_full_name(default_address.full_name)
            phone = default_address.mobile_number
            address = default_address.address_line1
            city = default_address.city
            state = default_address.state
            postcode = default_address.pincode
            country = default_address.country

        # Resell link orders are always resell and use link margin
        resell_link = None
        resell_link_id = request.session.get('resell_link_id')
        if resell_link_id:
            try:
                from .models import ResellLink
                resell_link = ResellLink.objects.select_related('reseller', 'product').get(
                    id=resell_link_id,
                    is_active=True
                )
                matched_link_qty = _get_resell_link_matching_quantity(cart_items, resell_link)
                if matched_link_qty > 0 and not explicit_self_selected:
                    is_resell = True
                    order_type = 'for_resell'
                    resell_margin_per_unit = Decimal(str(resell_link.margin_amount or 0))
                    if not from_name:
                        from_name = resell_link.reseller.get_full_name() or resell_link.reseller.username
                    if not from_phone and hasattr(resell_link.reseller, 'userprofile'):
                        from_phone = resell_link.reseller.userprofile.mobile_number or ''
                elif explicit_self_selected:
                    # User explicitly switched to self-order, so clear stale resell session.
                    request.session.pop('resell_link_id', None)
                    request.session.pop('resell_code', None)
                    resell_link = None
                    is_resell = False
                    order_type = 'for_self'
                    resell_margin_per_unit = Decimal('0')
                else:
                    # Link product no longer in checkout (cart changed), so disable link mode.
                    request.session.pop('resell_link_id', None)
                    request.session.pop('resell_code', None)
                    resell_link = None
                    is_resell = user_selected_resell
                    order_type = 'for_resell' if user_selected_resell else 'for_self'
            except ResellLink.DoesNotExist:
                request.session.pop('resell_link_id', None)
                request.session.pop('resell_code', None)
                resell_link = None
        
        # Validation
        if not all([first_name, last_name, email, phone, address, city, state, postcode]):
            messages.error(request, 'Please fill all required fields')
            return redirect('checkout')

        if (country or '').strip().lower() == 'india':
            if not re.fullmatch(r"[0-9]{6}", postcode or ''):
                messages.error(request, 'Please enter a valid 6-digit pincode.')
                return redirect('checkout')
            if pincode_valid != '1' and not _validate_indian_pincode(postcode):
                messages.error(request, 'Pincode not found. Please enter a valid pincode.')
                return redirect('checkout')
        
        # Validate resell FROM details if resell order
        if is_resell and (not from_name or not from_phone):
            messages.error(request, 'Please provide FROM details for resell order')
            return redirect('checkout')
        if is_resell and resell_margin_per_unit < 0:
            messages.error(request, 'Margin cannot be negative.')
            return redirect('checkout')
        if is_resell and resell_link is None and resell_margin_per_unit > 0 and base_cart_total > 0:
            min_unit_price = _get_checkout_min_unit_price(cart_items)
            max_margin = min_unit_price * Decimal('0.50')
            if resell_margin_per_unit > max_margin:
                messages.error(request, f'Margin cannot exceed 50% of item price (max ₹{max_margin:.2f} per item).')
                return redirect('checkout')

        request.session['checkout_form'] = {
            'first_name': first_name,
            'last_name': last_name,
            'email': email,
            'phone': phone,
            'country': country,
            'address': address,
            'city': city,
            'state': state,
            'postcode': postcode,
            'payment_method': payment_method,
            'customer_notes': customer_notes,
            'is_resell': is_resell,
            'order_type': order_type,
            'resell_margin': str(resell_margin_per_unit),
            'from_name': from_name,
            'from_phone': from_phone,
            'redeem_points': redeem_points,
            'points_to_redeem': points_to_redeem,
            'use_default_address': use_default_address,
            'set_default_address': set_default_address,
            'coupon_id': coupon_id,  # Save coupon ID
        }
        return redirect('checkout_confirm')
    
    # GET request - Show checkout form
    user_profile = UserProfile.objects.filter(user=request.user).first()
    default_address = Address.objects.filter(user=request.user, is_default=True).first()
    checkout_form = request.session.get('checkout_form', {})
    selected_order_type = checkout_form.get('order_type', 'for_self')
    manual_margin_raw = normalize_decimal_input(checkout_form.get('resell_margin', '0'))
    try:
        manual_margin_per_unit = Decimal(manual_margin_raw or '0')
    except Exception:
        manual_margin_per_unit = Decimal('0')
    if manual_margin_per_unit < 0:
        manual_margin_per_unit = Decimal('0')
    if selected_order_type == 'for_resell' and manual_margin_per_unit > 0:
        min_unit_price = _get_checkout_min_unit_price(cart_items)
        max_margin = min_unit_price * Decimal('0.50')
        if max_margin >= 0 and manual_margin_per_unit > max_margin:
            manual_margin_per_unit = max_margin
    
    # Check for resell link in session
    resell_link = None
    matched_link_qty = 0
    resell_link_id = request.session.get('resell_link_id')
    if resell_link_id:
        if order_type == 'for_self':
            request.session.pop('resell_link_id', None)
            request.session.pop('resell_code', None)
            resell_link_id = None
    
    if resell_link_id:
        try:
            from .models import ResellLink
            resell_link = ResellLink.objects.select_related('reseller', 'product').get(
                id=resell_link_id,
                is_active=True
            )
            matched_link_qty = _get_resell_link_matching_quantity(cart_items, resell_link)
            if matched_link_qty <= 0:
                request.session.pop('resell_link_id', None)
                request.session.pop('resell_code', None)
                resell_link = None
        except ResellLink.DoesNotExist:
            # Clear invalid resell link from session
            request.session.pop('resell_link_id', None)
            request.session.pop('resell_code', None)
    
    if resell_link:
        selected_order_type = 'for_resell'
        checkout_form['order_type'] = 'for_resell'
        checkout_form['is_resell'] = True

    # Get loyalty account
    loyalty_account = None
    if request.user.is_authenticated:
        try:
            loyalty_account = LoyaltyPoints.objects.get(user=request.user)
        except LoyaltyPoints.DoesNotExist:
            pass
    
    # Calculate totals with resell margins
    manual_margin_total = Decimal('0')
    display_subtotal = base_cart_total

    if resell_link:
        link_margin_total = Decimal(str(resell_link.margin_amount or 0)) * Decimal(str(matched_link_qty))
        display_subtotal = base_cart_total + link_margin_total
        manual_margin_per_unit = Decimal(str(resell_link.margin_amount or 0))
        manual_margin_total = link_margin_total
    elif selected_order_type == 'for_resell' and manual_margin_per_unit > 0 and total_item_qty > 0:
        manual_margin_total = manual_margin_per_unit * Decimal(str(total_item_qty))
        display_subtotal = base_cart_total + manual_margin_total
    
    tax_amount = display_subtotal * Decimal('0.05')
    shipping_cost = Decimal('0.00') if display_subtotal > 500 else Decimal('50.00')
    final_total = display_subtotal + tax_amount + shipping_cost

    context = {
        'cart_items': cart_items,
        'total_price': display_subtotal,
        'cart_base_total': base_cart_total,
        'total_item_qty': total_item_qty,
        'buy_now_item': buy_now_item,
        'user_profile': user_profile,
        'default_address': default_address,
        'checkout_form': checkout_form,
        'selected_order_type': selected_order_type,
        'loyalty_account': loyalty_account,
        'resell_link': resell_link,  # Add resell link to context
        'resell_link_matching_qty': matched_link_qty if resell_link else 0,
        'manual_resell_margin_per_unit': manual_margin_per_unit,
        'manual_resell_margin_total': manual_margin_total,
        'tax_amount': tax_amount,
        'shipping_cost': shipping_cost,
        'final_total': final_total,
    }
    
    return render(request, 'checkout.html', context)


@login_required(login_url='login')
def checkout_confirm(request):
    checkout_form = request.session.get('checkout_form')
    if not checkout_form:
        return redirect('checkout')

    cart_items, buy_now_item, total_price = _get_checkout_items(request)
    base_cart_total = Decimal(str(total_price))
    total_item_qty = _get_checkout_total_quantity(cart_items)

    if not cart_items:
        messages.error(request, 'Your cart is empty.')
        return redirect('cart')

    order_type = checkout_form.get('order_type', 'for_self')
    is_resell = (order_type == 'for_resell') or (checkout_form.get('is_resell') is True)
    margin_raw = normalize_decimal_input(checkout_form.get('resell_margin', '0'))
    try:
        resell_margin_per_unit = Decimal(margin_raw or '0')
    except Exception:
        resell_margin_per_unit = Decimal('0')
    if resell_margin_per_unit < 0:
        resell_margin_per_unit = Decimal('0')
    if is_resell and resell_margin_per_unit > 0:
        min_unit_price = _get_checkout_min_unit_price(cart_items)
        max_margin = min_unit_price * Decimal('0.50')
        if max_margin >= 0 and resell_margin_per_unit > max_margin:
            resell_margin_per_unit = max_margin

    # Check for resell link in session
    resell_link = None
    resell_link_id = request.session.get('resell_link_id')
    total_margin = Decimal('0')
    base_amount = base_cart_total
    matched_link_qty = 0

    if resell_link_id:
        try:
            from .models import ResellLink
            resell_link = ResellLink.objects.select_related('reseller', 'product').get(
                id=resell_link_id,
                is_active=True
            )
            matched_link_qty = _get_resell_link_matching_quantity(cart_items, resell_link)
            if matched_link_qty > 0:
                total_margin = Decimal(str(resell_link.margin_amount or 0)) * Decimal(str(matched_link_qty))
                resell_margin_per_unit = Decimal(str(resell_link.margin_amount or 0))
                is_resell = True
                order_type = 'for_resell'
            else:
                request.session.pop('resell_link_id', None)
                request.session.pop('resell_code', None)
                resell_link = None
                is_resell = False
                order_type = 'for_self'
                resell_margin_per_unit = Decimal('0')
        except ResellLink.DoesNotExist:
            # Clear invalid resell link from session
            request.session.pop('resell_link_id', None)
            request.session.pop('resell_code', None)
            is_resell = False
            order_type = 'for_self'
            resell_margin_per_unit = Decimal('0')
    elif is_resell and resell_margin_per_unit > 0 and total_item_qty > 0:
        total_margin = resell_margin_per_unit * Decimal(str(total_item_qty))

    subtotal = base_amount + total_margin
    tax = subtotal * Decimal('0.05')
    shipping_cost = Decimal('0.00') if subtotal > 500 else Decimal('50.00')

    # Coupon discount
    coupon_discount = Decimal('0')
    applied_coupon = None
    coupon_id = checkout_form.get('coupon_id')
    if coupon_id:
        try:
            from .models import Coupon, CouponUsage
            applied_coupon = Coupon.objects.get(id=coupon_id)
            # Validate coupon is still valid
            if applied_coupon.is_valid():
                # Check if not already used
                already_used = CouponUsage.objects.filter(
                    coupon=applied_coupon,
                    user=request.user
                ).exists()
                if not already_used:
                    # Calculate discount
                    cart_total = subtotal + tax + shipping_cost
                    coupon_discount = Decimal(str(applied_coupon.get_discount_amount(float(cart_total))))
        except Coupon.DoesNotExist:
            applied_coupon = None

    points_discount = Decimal('0')
    points_to_redeem = int(checkout_form.get('points_to_redeem') or 0)
    redeem_points = checkout_form.get('redeem_points') is True
    if redeem_points and points_to_redeem > 0:
        try:
            loyalty_account = LoyaltyPoints.objects.get(user=request.user)
            if points_to_redeem <= loyalty_account.points_available:
                points_discount = Decimal(str(points_to_redeem)) * Decimal('0.03')
        except LoyaltyPoints.DoesNotExist:
            redeem_points = False

    total_amount = subtotal + tax + shipping_cost - coupon_discount - points_discount
    if total_amount < 0:
        total_amount = Decimal('0')

    if request.method == 'POST':
        first_name = checkout_form.get('first_name')
        last_name = checkout_form.get('last_name')
        email = checkout_form.get('email')
        phone = checkout_form.get('phone')
        country = checkout_form.get('country')
        address = checkout_form.get('address')
        city = checkout_form.get('city')
        state = checkout_form.get('state')
        postcode = checkout_form.get('postcode')
        payment_method = checkout_form.get('payment_method', 'COD')
        customer_notes = checkout_form.get('customer_notes', '')
        from_name = checkout_form.get('from_name', '')
        from_phone = checkout_form.get('from_phone', '')
        set_default_address = checkout_form.get('set_default_address') is True

        if not all([first_name, last_name, email, phone, address, city, state, postcode]):
            messages.error(request, 'Please fill all required fields')
            return redirect('checkout')

        if (country or '').strip().lower() == 'india':
            if not re.fullmatch(r"[0-9]{6}", postcode or ''):
                messages.error(request, 'Please enter a valid 6-digit pincode.')
                return redirect('checkout')
            if not _validate_indian_pincode(postcode):
                messages.error(request, 'Pincode not found. Please enter a valid pincode.')
                return redirect('checkout')

        try:
            shipping_address = f"{first_name} {last_name}\n{address}\n{city}, {state} {postcode}\n{country}"
            
            # Check for resell link in session
            resell_link_id = request.session.get('resell_link_id')
            resell_link = None
            
            if resell_link_id:
                try:
                    from .models import ResellLink
                    from .resell_services import ResellOrderProcessor
                    
                    resell_link = ResellLink.objects.select_related('reseller', 'product').get(
                        id=resell_link_id,
                        is_active=True
                    )
                    
                    # Use ResellOrderProcessor for resell orders
                    order = ResellOrderProcessor.create_resell_order(
                        cart_items=cart_items,
                        resell_link=resell_link,
                        customer=request.user,
                        shipping_address=shipping_address,
                        billing_address=shipping_address,
                        payment_method=payment_method,
                        tax=tax,
                        shipping_cost=shipping_cost,
                        coupon_discount=coupon_discount,
                        points_discount=points_discount,
                        coupon=applied_coupon,
                        payment_status='PENDING',
                    )
                    
                    # Add customer notes if provided
                    if customer_notes:
                        order.customer_notes = customer_notes
                    if not order.resell_from_name:
                        order.resell_from_name = from_name or (resell_link.reseller.get_full_name() or resell_link.reseller.username)
                    if not order.resell_from_phone and hasattr(resell_link.reseller, 'userprofile'):
                        order.resell_from_phone = from_phone or (resell_link.reseller.userprofile.mobile_number or '')
                    order.save(update_fields=['customer_notes', 'resell_from_name', 'resell_from_phone'])
                    
                except ResellLink.DoesNotExist:
                    messages.error(request, 'Resell link is no longer valid.')
                    return redirect('checkout')
                except Exception as e:
                    messages.error(request, f'Error processing resell order: {str(e)}')
                    return redirect('checkout')
            else:
                # Regular order and manual resell order
                order_kwargs = {
                    'user': request.user,
                    'subtotal': subtotal,
                    'tax': tax,
                    'shipping_cost': shipping_cost,
                    'coupon': applied_coupon,  # Save coupon reference
                    'coupon_discount': coupon_discount,  # Save discount amount
                    'total_amount': total_amount,
                    'shipping_address': shipping_address,
                    'billing_address': shipping_address,
                    'payment_method': payment_method,
                    'customer_notes': customer_notes,
                    'is_resell': is_resell,
                    'resell_from_name': from_name if is_resell else '',
                    'resell_from_phone': from_phone if is_resell else '',
                }
                if is_resell:
                    order_kwargs.update({
                        'reseller': request.user,
                        'base_amount': base_amount,
                        'total_margin': total_margin,
                    })
                order = Order.objects.create(**order_kwargs)

            # Create CouponUsage record if coupon was applied
            if applied_coupon and coupon_discount > 0:
                from .models import CouponUsage
                CouponUsage.objects.create(
                    coupon=applied_coupon,
                    user=request.user,
                    order=order,
                    discount_amount=coupon_discount
                )
                order.admin_notes += f"\nCoupon Applied: {applied_coupon.code} (₹{coupon_discount} discount)"
                order.save()

            if redeem_points and points_to_redeem > 0:
                order.admin_notes += f"\nLoyalty Points Redeemed: {points_to_redeem} points (₹{points_discount} discount)"
                order.save()

            # Only create order items if NOT a resell order (ResellOrderProcessor already created them)
            if not resell_link:
                if buy_now_item:
                    product = buy_now_item['product']
                    product_image = ''
                    if product.image:
                        image_url = product.image.url
                        if not image_url.startswith('http'):
                            site_url = request.build_absolute_uri('/').rstrip('/')
                            product_image = f"{site_url}{image_url}"
                        else:
                            product_image = image_url

                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        product_name=product.name,
                        product_price=Decimal(str(buy_now_item['price'])) + (resell_margin_per_unit if is_resell else Decimal('0')),
                        product_image=product_image,
                        quantity=buy_now_item['quantity'],
                        base_price=Decimal(str(buy_now_item['price'])),
                        margin_amount=resell_margin_per_unit if is_resell else (product.margin if product else Decimal('0'))
                    )
                    del request.session['buy_now_item']
                else:
                    for item in cart_items:
                        product_image = ''
                        if item.product.image:
                            image_url = item.product.image.url
                            if not image_url.startswith('http'):
                                site_url = request.build_absolute_uri('/').rstrip('/')
                                product_image = f"{site_url}{image_url}"
                            else:
                                product_image = image_url

                        item_base_price = Decimal(str(item.product.price))
                        item_margin = resell_margin_per_unit if is_resell else Decimal(str(item.product.margin or 0))
                        OrderItem.objects.create(
                            order=order,
                            product=item.product,
                            product_name=item.product.name,
                            product_price=item_base_price + item_margin,
                            product_image=product_image,
                            quantity=item.quantity,
                            base_price=item_base_price,
                            margin_amount=item_margin
                        )

            # If no actual order items were created, delete the order and abort.
            if not OrderItem.objects.filter(order=order).exists():
                order.delete()
                messages.error(request, 'Checkout failed: cart was empty or items could not be added. Please try again.')
                return redirect('cart')

            if is_resell and not resell_link and total_margin > 0:
                from .models import ResellerProfile, ResellerEarning
                reseller_profile, _ = ResellerProfile.objects.get_or_create(
                    user=request.user,
                    defaults={
                        'is_reseller_enabled': True,
                        'business_name': request.user.get_full_name() or request.user.username,
                    }
                )
                if not reseller_profile.is_reseller_enabled:
                    reseller_profile.is_reseller_enabled = True
                    reseller_profile.save(update_fields=['is_reseller_enabled'])

                if not hasattr(order, 'reseller_earning'):
                    ResellerEarning.objects.create(
                        reseller=request.user,
                        order=order,
                        resell_link=None,
                        margin_amount=total_margin,
                        status='PENDING'
                    )

            if redeem_points and points_to_redeem > 0:
                loyalty_account = LoyaltyPoints.objects.get(user=request.user)
                loyalty_account.redeem_points(points_to_redeem, f"Order #{order.order_number} - ₹{points_discount} discount")

            if set_default_address:
                Address.objects.create(
                    user=request.user,
                    full_name=f"{first_name} {last_name}".strip(),
                    mobile_number=phone,
                    address_line1=address,
                    address_line2='',
                    city=city,
                    state=state,
                    pincode=postcode,
                    country=country or 'India',
                    address_type='HOME',
                    is_default=True
                )

            request.session.pop('checkout_form', None)
            
            # Clear resell link from session after order is created
            if resell_link:
                request.session.pop('resell_link_id', None)
                request.session.pop('resell_code', None)

            if payment_method == 'RAZORPAY':
                return redirect('razorpay_payment', order_id=order.id)

            order.payment_status = 'PENDING'
            order.order_status = 'PENDING'
            order.save()

            order.auto_process_approval()

            # Clear cart for both regular and resell orders
            if not buy_now_item:
                Cart.objects.filter(user=request.user).delete()

            try:
                email_sent = send_order_confirmation_email(order)
                if not email_sent:
                    logger.warning(f"Order confirmation email failed silently for order {order.order_number}")
                    messages.warning(request, 'Order placed successfully, but confirmation email could not be delivered. Please check your email address or contact support.')
            except Exception as email_exc:
                logger.error(f"Order confirmation email exception for order {order.order_number}: {email_exc}", exc_info=True)
                messages.warning(request, 'Order placed successfully, but confirmation email failed to send. Please contact support.')

            try:
                send_admin_order_notification(order, request)
            except Exception as admin_email_exc:
                logger.error(f"Admin notification email exception for order {order.order_number}: {admin_email_exc}", exc_info=True)

            if order.approval_status == 'PENDING_APPROVAL':
                messages.warning(request, f'Order placed successfully! Order #: {order.order_number}. Your order is pending approval due to security checks.')
            else:
                messages.success(request, f'Order placed successfully! Order #: {order.order_number}')

            return redirect('order_confirmation', order_id=order.id)

        except Exception as exc:
            messages.error(request, f'Error creating order: {str(exc)}')
            return redirect('checkout')

    context = {
        'checkout_form': checkout_form,
        'cart_items': cart_items,
        'subtotal': subtotal,
        'base_amount': base_amount,
        'total_margin': total_margin,
        'resell_margin_per_unit': resell_margin_per_unit,
        'total_item_qty': total_item_qty,
        'resell_margin_item_qty': matched_link_qty if resell_link else total_item_qty,
        'is_resell': is_resell,
        'order_type': order_type,
        'from_name': checkout_form.get('from_name', ''),
        'from_phone': checkout_form.get('from_phone', ''),
        'tax': tax,
        'shipping_cost': shipping_cost,
        'coupon_discount': coupon_discount,
        'applied_coupon': applied_coupon,
        'points_discount': points_discount,
        'total_amount': total_amount,
        'resell_link': resell_link,  # Add resell link to context
        'payment_method': checkout_form.get('payment_method', 'COD'),
        'address_text': f"{checkout_form.get('first_name', '')} {checkout_form.get('last_name', '')}\n{checkout_form.get('address', '')}\n{checkout_form.get('city', '')}, {checkout_form.get('state', '')} {checkout_form.get('postcode', '')}\n{checkout_form.get('country', '')}".strip(),
    }
    return render(request, 'checkout_confirm.html', context)
def contact(request):
    """Contact page with form submission"""
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        
        # Validation
        if not all([name, email, subject, message]):
            messages.error(request, 'Please fill all required fields.')
            return render(request, 'contact.html')
        
        # Email validation
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            messages.error(request, 'Please enter a valid email address.')
            return render(request, 'contact.html')
        
        try:
            # Send email to admin
            from django.core.mail import send_mail
            from django.conf import settings
            
            email_subject = f"Contact Form: {subject}"
            email_message = f"""
New Contact Form Submission

Name: {name}
Email: {email}
Phone: {phone if phone else 'Not provided'}
Subject: {subject}

Message:
{message}

---
This message was sent from VibeMall Contact Form
            """
            
            send_mail(
                email_subject,
                email_message,
                settings.DEFAULT_FROM_EMAIL,
                ['info.vibemall@gmail.com'],
                fail_silently=False,
            )
            
            # Send confirmation email to user
            confirmation_subject = "Thank you for contacting VibeMall"
            confirmation_message = f"""
Dear {name},

Thank you for reaching out to VibeMall! We have received your message and our team will get back to you within 24 hours.

Your Message Details:
Subject: {subject}
Message: {message}

If you have any urgent concerns, please feel free to call us at +91 123 456 7890.

Best regards,
VibeMall Team
info.vibemall@gmail.com
            """
            
            send_mail(
                confirmation_subject,
                confirmation_message,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=True,
            )
            
            messages.success(request, 'Thank you for your message! We will get back to you within 24 hours.')
            return redirect('contact')
            
        except Exception as e:
            messages.error(request, 'Sorry, there was an error sending your message. Please try again or email us directly at info.vibemall@gmail.com')
            return render(request, 'contact.html')
    
    return render(request, 'contact.html')
def faq(request):
    active_faqs = FAQ.objects.filter(is_active=True).order_by('sort_order', 'question')
    categories = (
        FAQCategory.objects
        .filter(is_active=True)
        .prefetch_related(Prefetch('faqs', queryset=active_faqs))
        .order_by('sort_order', 'name')
    )

    faq_groups = []
    for index, category in enumerate(categories, start=1):
        category_faqs = list(category.faqs.all())
        if not category_faqs:
            continue

        anchor = re.sub(r'[^a-z0-9]+', '-', (category.name or '').lower()).strip('-')
        faq_groups.append({
            'id': category.id,
            'name': category.name,
            'description': category.description,
            'anchor': anchor or f'faq-category-{index}',
            'faqs': category_faqs,
        })

    if not faq_groups:
        faq_groups = [
            {
                'id': 0,
                'name': 'Orders & Delivery',
                'description': 'Everything about placing orders, shipping timelines, and tracking your parcel.',
                'anchor': 'orders-delivery',
                'faqs': [
                    {
                        'question': 'How do I place an order on VibeMall?',
                        'answer': 'Browse any category or use Search to find the product you want. Select your size or variant, click Add to Cart, and proceed to Checkout. You can pay by card, UPI, net banking, or Cash on Delivery.',
                    },
                    {
                        'question': 'How long does delivery take?',
                        'answer': 'Standard orders are delivered within 5 to 7 business days. Delivery timelines may vary based on your pincode and the product category. Estimated delivery is shown at checkout before you confirm the order.',
                    },
                    {
                        'question': 'How can I track my order?',
                        'answer': 'Once your order is shipped, a tracking link is sent to your registered email and mobile number. You can also check the live status from My Orders inside your account.',
                    },
                    {
                        'question': 'Can I change my delivery address after placing an order?',
                        'answer': 'Address changes can be requested only before the order is dispatched. Please contact us at info.vibemall@gmail.com immediately with your order number and the updated address.',
                    },
                    {
                        'question': 'Do you deliver across India?',
                        'answer': 'Yes, VibeMall delivers to most serviceable pin codes across India. You can check delivery availability by entering your pin code on the product page.',
                    },
                ],
            },
            {
                'id': 1,
                'name': 'Returns & Refunds',
                'description': 'Details about return eligibility, the return process, and how refunds are processed.',
                'anchor': 'returns-refunds',
                'faqs': [
                    {
                        'question': 'What is the return policy at VibeMall?',
                        'answer': 'Most products are eligible for return within 7 days of delivery, provided the item is unused, unwashed, and returned in its original packaging with all tags intact. The exact return window is shown on each product page.',
                    },
                    {
                        'question': 'How do I initiate a return?',
                        'answer': 'Go to My Orders in your account, select the order, and click Request Return. Fill in the reason and submit. Our team will review and arrange a pickup within 48 hours of approval.',
                    },
                    {
                        'question': 'When will I receive my refund?',
                        'answer': 'Refunds are processed within 5 to 7 business days after we receive and inspect the returned item. The amount is credited to your original payment method or as VibeMall wallet credit, depending on the refund type.',
                    },
                    {
                        'question': 'Are there items that cannot be returned?',
                        'answer': 'Innerwear, customised products, perishable goods, and items marked as final sale are not eligible for return. This is mentioned clearly on the respective product pages.',
                    },
                    {
                        'question': 'What should I do if I received a damaged or wrong product?',
                        'answer': 'We apologise for the inconvenience. Please email us at info.vibemall@gmail.com within 48 hours of delivery with your order number and clear photos of the item. We will arrange an immediate replacement or full refund.',
                    },
                ],
            },
            {
                'id': 2,
                'name': 'Payments',
                'description': 'Accepted payment methods, EMI options, coupon usage, and transaction concerns.',
                'anchor': 'payments',
                'faqs': [
                    {
                        'question': 'What payment methods are accepted?',
                        'answer': 'We accept all major credit and debit cards, UPI (GPay, PhonePe, Paytm), net banking, and Cash on Delivery for eligible orders.',
                    },
                    {
                        'question': 'Is it safe to use my card or UPI on VibeMall?',
                        'answer': 'Yes. All transactions are processed through a secure and encrypted payment gateway. VibeMall does not store your card details. Look for the padlock icon in your browser address bar during checkout.',
                    },
                    {
                        'question': 'My payment failed but the amount was deducted. What should I do?',
                        'answer': 'If a payment fails after deduction, the amount is automatically reversed to your bank or card within 5 to 7 business days. If it is not credited, please contact us at info.vibemall@gmail.com with your order ID and transaction reference.',
                    },
                    {
                        'question': 'How do I apply a coupon or promo code?',
                        'answer': 'Enter your coupon code in the Promo Code field on the checkout page and click Apply. The discount will be reflected in your order total before payment.',
                    },
                    {
                        'question': 'Is Cash on Delivery available for all orders?',
                        'answer': 'Cash on Delivery is available for select pin codes and order values. You will see the COD option at checkout if it is available for your location and order.',
                    },
                ],
            },
            {
                'id': 3,
                'name': 'Products & Collections',
                'description': 'Size guidance, product authenticity, availability, and how we curate our catalogue.',
                'anchor': 'products-collections',
                'faqs': [
                    {
                        'question': 'How do I find the right size?',
                        'answer': 'Each product page includes a Size Guide specific to that category. We recommend checking the guide before ordering, especially for clothing and footwear, as sizing may vary by brand.',
                    },
                    {
                        'question': 'Are the products on VibeMall authentic?',
                        'answer': 'Yes. All products are sourced from verified brand partners and authorised suppliers. We do not list counterfeit or duplicate goods. Quality is reviewed before listings are approved.',
                    },
                    {
                        'question': 'What if a product I want is out of stock?',
                        'answer': 'You can tap the Notify Me button on the product page to receive an alert when the item is restocked. We restock popular items regularly based on demand.',
                    },
                    {
                        'question': 'Do product images accurately represent the actual item?',
                        'answer': 'We try our best to display accurate images. However, slight variations in colour may occur due to screen settings and photography lighting. The product description provides additional details about the actual item.',
                    },
                    {
                        'question': 'Can I buy products in bulk or for resale?',
                        'answer': 'Bulk and resale purchases may be available for selected categories. Please reach out to us at info.vibemall@gmail.com with your requirements for guidance.',
                    },
                ],
            },
            {
                'id': 4,
                'name': 'Account & Profile',
                'description': 'Managing your account, saved addresses, order history, and profile settings.',
                'anchor': 'account-profile',
                'faqs': [
                    {
                        'question': 'How do I create a VibeMall account?',
                        'answer': 'Click on the Profile icon and select Sign Up. Enter your name, email address, and a password. You can also register using your mobile number with OTP verification.',
                    },
                    {
                        'question': 'I forgot my password. How do I reset it?',
                        'answer': 'On the login page, click Forgot Password and enter your registered email address. A reset link will be sent to your inbox. Follow the link to set a new password.',
                    },
                    {
                        'question': 'How do I save or update a delivery address?',
                        'answer': 'Go to My Account and select Manage Addresses. You can add new addresses, edit existing ones, and set a default address for faster checkout.',
                    },
                    {
                        'question': 'Where can I see my past orders?',
                        'answer': 'All your orders are listed under My Orders in your account. You can view order status, tracking details, and download invoices from there.',
                    },
                    {
                        'question': 'Can I cancel my order?',
                        'answer': 'Orders can be cancelled before they are dispatched. Go to My Orders, select the order, and choose Cancel. If the order has already shipped, you can initiate a return once it is delivered.',
                    },
                ],
            },
            {
                'id': 5,
                'name': 'Support & Contact',
                'description': 'How to reach our team and what to expect when you contact us.',
                'anchor': 'support-contact',
                'faqs': [
                    {
                        'question': 'How can I contact VibeMall customer support?',
                        'answer': 'You can reach us by emailing info.vibemall@gmail.com or by using the Contact Us form on our website. We aim to respond to all queries within 24 hours on business days.',
                    },
                    {
                        'question': 'What are the customer support hours?',
                        'answer': 'Our support team is available Monday to Saturday, 10 AM to 7 PM IST. For urgent concerns outside these hours, please email us and we will get back to you on the next business day.',
                    },
                    {
                        'question': 'How do I report a problem with an order or the website?',
                        'answer': 'Email us at info.vibemall@gmail.com with a clear description of the issue and your order number if applicable. Screenshots or photos help us resolve the issue faster.',
                    },
                    {
                        'question': 'I have a suggestion or feedback. How do I share it?',
                        'answer': 'We love hearing from our customers. Please use the Contact Us page or email us directly at info.vibemall@gmail.com. Your feedback helps us improve the VibeMall experience.',
                    },
                ],
            },
        ]

    total_faqs = sum(len(group['faqs']) for group in faq_groups)
    context = {
        'faq_groups': faq_groups,
        'total_faqs': total_faqs,
        'total_categories': len(faq_groups),
    }
    return render(request, 'faq.html', context)


def coming_soon(request):
    launch_date = timezone.now() + timedelta(days=21)
    context = {
        'launch_date_iso': launch_date.strftime('%Y-%m-%dT%H:%M:%SZ'),
        'launch_date_display': launch_date.strftime('%d %b %Y'),
    }
    return render(request, 'coming_soon.html', context)


def launch_animation(request):
    redirect_url = '/' if not getattr(settings, 'COMING_SOON_MODE', True) else '/coming-soon/'
    return render(request, 'launch_animation.html', {'redirect_url': redirect_url})


def login_view(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        username = (request.POST.get("username") or "").strip()
        password = request.POST.get("password")
        remember_me = request.POST.get("remember_me")
        next_url = request.POST.get("next") or request.GET.get("next")

        login_username = username
        if "@" in username:
            try:
                matched_user = User.objects.filter(email__iexact=username).only('username').first()
                if matched_user:
                    login_username = matched_user.username
            except Exception:
                login_username = username

        user = authenticate(request, username=login_username, password=password)
        if user:
            login(request, user)
            # Set session expiry based on remember me checkbox
            if remember_me:
                request.session.set_expiry(30 * 24 * 60 * 60)  # 30 days
            else:
                request.session.set_expiry(0)  # Session expires when browser closes

            if next_url and next_url.startswith('/') and not next_url.startswith('//'):
                return redirect(next_url)

            if user.is_staff:
                return redirect("admin_dashboard")

            return redirect("index")
        else:
            messages.error(request, "Invalid username or password. Please check your credentials and try again.")

    return render(request, "login.html", {
        "next": request.GET.get("next", "")
    })




def my_account(request): return render(request, 'profile.html')
def product(request): return render(request, 'product.html')


def product_detail(request: HttpRequest, slug: str) -> HttpResponse:
    product = get_object_or_404(Product, slug=slug, is_active=True)
    return redirect('product-details', product_id=product.id)
def product_details(request: HttpRequest, product_id: Optional[int] = None) -> HttpResponse:
    """Display product details with dynamic data and enhanced reviews"""
    if product_id:
        try:
            product = Product.objects.get(id=product_id, is_active=True)
            in_wishlist = False
            if request.user.is_authenticated:
                in_wishlist = Wishlist.objects.filter(user=request.user, product=product).exists()

            # Auto-generate reviews for existing products if missing
            if product.review_count and product.rating:
                existing_qs = ProductReview.objects.filter(product=product, is_approved=True)
                existing_count = existing_qs.count()
                if existing_count < product.review_count:
                    existing_sum = existing_qs.aggregate(total=Sum('rating'))['total'] or 0
                    desired_sum = int(round(float(product.rating) * int(product.review_count)))
                    missing_n = int(product.review_count) - existing_count
                    if missing_n > 0:
                        missing_sum = desired_sum - existing_sum
                        missing_avg = missing_sum / missing_n if missing_n else float(product.rating)
                        if missing_avg < 1:
                            missing_avg = 1
                        if missing_avg > 5:
                            missing_avg = 5
                        generate_auto_reviews(product, missing_n, missing_avg, request.user if request.user.is_authenticated else None)
            
            # Get filter and sort parameters
            rating_filter = request.GET.get('rating', 'all')
            sort_by = request.GET.get('sort', 'recent')
            
            # Get approved customer reviews for this product (exclude auto-generated)
            approved_reviews = ProductReview.objects.filter(
                product=product, 
                is_approved=True,
                is_auto_generated=False
            ).select_related('user').prefetch_related('images', 'votes')
            
            # Apply rating filter
            if rating_filter != 'all':
                try:
                    rating_value = int(rating_filter)
                    approved_reviews = approved_reviews.filter(rating=rating_value)
                except ValueError:
                    pass
            
            # Apply sorting
            if sort_by == 'recent':
                approved_reviews = approved_reviews.order_by('-created_at')
            elif sort_by == 'rating_high':
                approved_reviews = approved_reviews.order_by('-rating', '-created_at')
            elif sort_by == 'rating_low':
                approved_reviews = approved_reviews.order_by('rating', '-created_at')
            elif sort_by == 'helpful':
                approved_reviews = approved_reviews.order_by('-helpful_count', '-created_at')
            
            review_count = approved_reviews.count()
            
            # Calculate rating breakdown (include ALL approved reviews for stats)
            total_reviews = ProductReview.objects.filter(product=product, is_approved=True).count()
            rating_breakdown = {
                5: ProductReview.objects.filter(product=product, is_approved=True, rating=5).count(),
                4: ProductReview.objects.filter(product=product, is_approved=True, rating=4).count(),
                3: ProductReview.objects.filter(product=product, is_approved=True, rating=3).count(),
                2: ProductReview.objects.filter(product=product, is_approved=True, rating=2).count(),
                1: ProductReview.objects.filter(product=product, is_approved=True, rating=1).count(),
            }
            
            # Calculate percentages
            rating_percentages = {}
            for rating, count in rating_breakdown.items():
                rating_percentages[rating] = int((count / total_reviews * 100)) if total_reviews > 0 else 0
            
            # Calculate average rating (include ALL approved reviews for stats)
            avg_rating = ProductReview.objects.filter(
                product=product, 
                is_approved=True
            ).aggregate(Avg('rating'))['rating__avg'] or 0
            
            # Get product questions
            approved_questions = ProductQuestion.objects.filter(
                product=product,
                is_approved=True
            ).select_related('user', 'answered_by').order_by('-answered_at', '-created_at')
            
            total_questions = approved_questions.count()
            
            # Get user's votes if authenticated
            user_votes = {}
            if request.user.is_authenticated:
                votes = ReviewVote.objects.filter(
                    user=request.user,
                    review__in=approved_reviews
                ).values_list('review_id', 'is_helpful')
                user_votes = dict(votes)
            
            return render(request, 'product-details.html', {
                'product': product,
                'in_wishlist': in_wishlist,
                'reviews': approved_reviews,
                'review_count': review_count,
                'total_reviews': total_reviews,
                'avg_rating': round(avg_rating, 1),
                'rating_breakdown': rating_breakdown,
                'rating_percentages': rating_percentages,
                'approved_questions': approved_questions,
                'total_questions': total_questions,
                'user_votes': user_votes,
                'rating_filter': rating_filter,
                'sort_by': sort_by,
            })
        except Product.DoesNotExist:
            return render(request, '404.html', status=404)
    else:
        return render(request, '404.html', status=404)
def shop(request):
    products = Product.objects.filter(is_active=True)
    
    # Get banners for shop page
    banners = Banner.objects.filter(is_active=True).filter(
        Q(page_type='SHOP') | Q(page_type='BOTH')
    ).order_by('order')

    # Filters from query params
    search_query = request.GET.get('q', '').strip()
    min_price = request.GET.get('min_price')
    max_price = request.GET.get('max_price')
    min_rating = request.GET.get('min_rating')
    selected_sort = request.GET.get('sort', 'curated').strip().lower() or 'curated'
    selected_category = request.GET.get('category', '').strip()
    selected_sub_category = request.GET.get('sub_category', '').strip()
    selected_brands = []
    seen_brands = set()
    for raw_brand in request.GET.getlist('brand'):
        brand_name = (raw_brand or '').strip()
        brand_key = brand_name.casefold()
        if brand_name and brand_key not in seen_brands:
            selected_brands.append(brand_name)
            seen_brands.add(brand_key)

    def parse_decimal_filter(raw_value):
        if raw_value in (None, ''):
            return None
        try:
            return Decimal(str(raw_value).strip())
        except (InvalidOperation, TypeError, ValueError):
            return None

    if search_query:
        products = products.filter(
            Q(name__icontains=search_query) |
            Q(sku__icontains=search_query) |
            Q(brand__icontains=search_query) |
            Q(tags__icontains=search_query)
        )

    special_offers = Product.objects.filter(is_active=True).order_by('-discount_percent', '-id')[:5]

    # Get categories from CategoryIcon (dynamic admin-managed categories)
    category_icons = CategoryIcon.objects.filter(is_active=True).order_by('order', 'id')

    def normalize_category_key(raw_key):
        return (raw_key or '').strip()

    def normalize_for_compare(raw_key):
        return normalize_category_key(raw_key).casefold()

    category_keys = {normalize_for_compare(cat.category_key) for cat in category_icons}

    selected_category = normalize_category_key(selected_category)
    selected_category_norm = normalize_for_compare(selected_category)
    if selected_category_norm and selected_category_norm in category_keys:
        products = products.annotate(
            _category_norm=Lower(Trim(Coalesce('category', Value(''))))
        ).filter(_category_norm=selected_category_norm)
    else:
        selected_category = ''
        selected_category_norm = ''

    if selected_sub_category:
        selected_sub_category = normalize_category_key(selected_sub_category)
        selected_sub_category_norm = normalize_for_compare(selected_sub_category)
        valid_sub_category_qs = Product.objects.filter(is_active=True)
        if selected_category_norm:
            valid_sub_category_qs = valid_sub_category_qs.annotate(
                _category_norm=Lower(Trim(Coalesce('category', Value(''))))
            ).filter(_category_norm=selected_category_norm)

        valid_sub_categories = set(
            normalize_for_compare(item)
            for item in valid_sub_category_qs.exclude(sub_category='').values_list('sub_category', flat=True)
            if normalize_category_key(item)
        )

        if selected_sub_category_norm in valid_sub_categories:
            products = products.annotate(
                _sub_category_norm=Lower(Trim(Coalesce('sub_category', Value(''))))
            ).filter(_sub_category_norm=selected_sub_category_norm)
        else:
            selected_sub_category = ''
    else:
        selected_sub_category = ''

    sidebar_scope = products

    designer_brands = list(
        sidebar_scope.annotate(
            brand_label=Trim(Coalesce('brand', Value('')))
        )
        .exclude(brand_label='')
        .values('brand_label')
        .annotate(product_count=Count('id'))
        .order_by('-product_count', 'brand_label')[:5]
    )

    if selected_brands:
        brand_query = Q()
        for brand_name in selected_brands:
            brand_query |= Q(brand__iexact=brand_name)
        products = products.filter(brand_query)

    rating_scope = products
    rating_options = []
    for threshold in (4, 3, 2):
        match_count = rating_scope.filter(rating__gte=threshold).count()
        if match_count:
            rating_options.append({
                'value': str(threshold),
                'label': f'{threshold} Stars & Up',
                'count': match_count,
            })

    selected_rating = ''
    if min_rating:
        try:
            parsed_rating = str(int(float(min_rating)))
            if parsed_rating in {'2', '3', '4', '5'}:
                selected_rating = parsed_rating
                products = products.filter(rating__gte=int(parsed_rating))
        except (TypeError, ValueError):
            selected_rating = ''

    price_scope = products
    price_bounds = price_scope.aggregate(min_price=Min('price'), max_price=Max('price'))
    filter_min_price = int(price_bounds['min_price'] or 0)
    filter_max_price = int(price_bounds['max_price'] or filter_min_price or 0)
    range_input_max = filter_max_price if filter_max_price > filter_min_price else filter_min_price + 1

    parsed_min_price = parse_decimal_filter(min_price)
    parsed_max_price = parse_decimal_filter(max_price)

    if parsed_min_price is not None:
        products = products.filter(price__gte=parsed_min_price)
    if parsed_max_price is not None:
        products = products.filter(price__lte=parsed_max_price)

    category_counts_qs = Product.objects.filter(is_active=True).values('category').annotate(total=Count('id'))
    category_counts = {row['category']: row['total'] for row in category_counts_qs}
    category_counts_norm = {
        normalize_category_key(key): total
        for key, total in category_counts.items()
    }
    category_data = [
        (
            normalize_category_key(cat.category_key),
            cat.name,
            category_counts_norm.get(normalize_category_key(cat.category_key), 0),
        )
        for cat in category_icons
    ]

    sub_category_data = []
    if selected_category_norm:
        sub_category_qs = Product.objects.filter(is_active=True).annotate(
            _category_norm=Lower(Trim(Coalesce('category', Value(''))))
        ).filter(_category_norm=selected_category_norm)
        sub_category_counts_qs = (
            sub_category_qs.exclude(sub_category='')
            .values('sub_category')
            .annotate(total=Count('id'))
            .order_by('sub_category')
        )
        sub_category_data = [
            (row['sub_category'], row['sub_category'], row['total'])
            for row in sub_category_counts_qs
        ]

    wishlist_product_ids = set()
    cart_product_ids = set()
    if request.user.is_authenticated:
        wishlist_product_ids = set(
            Wishlist.objects.filter(user=request.user).values_list('product_id', flat=True)
        )
        cart_product_ids = set(
            Cart.objects.filter(user=request.user).values_list('product_id', flat=True)
        )

    if selected_sort == 'name_asc':
        products = products.order_by('name', '-id')
    elif selected_sort == 'name_desc':
        products = products.order_by('-name', '-id')
    elif selected_sort == 'price_asc':
        products = products.order_by('price', '-id')
    elif selected_sort == 'price_desc':
        products = products.order_by('-price', '-id')
    else:
        selected_sort = 'curated'
        products = products.order_by('-id')

    # Pagination - 9 products per page (3 columns x 3 rows)
    paginator = Paginator(products, 9)
    page = request.GET.get('page', 1)
    
    try:
        products_page = paginator.page(page)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)

    def build_page_links(current, total):
        if total <= 1:
            return [1]
        if total <= 7:
            return list(range(1, total + 1))

        pages = {1, 2, total - 1, total, current - 1, current, current + 1}
        pages = [p for p in pages if 1 <= p <= total]
        pages = sorted(set(pages))
        output = []
        last = None
        for p in pages:
            if last is not None and p - last > 1:
                output.append(None)
            output.append(p)
            last = p
        return output

    page_links = build_page_links(products_page.number, paginator.num_pages)

    query_params = {}
    if selected_category:
        query_params['category'] = selected_category
    if min_price:
        query_params['min_price'] = min_price
    if max_price:
        query_params['max_price'] = max_price
    if selected_rating:
        query_params['min_rating'] = selected_rating
    if selected_sub_category:
        query_params['sub_category'] = selected_sub_category
    if search_query:
        query_params['q'] = search_query
    if selected_brands:
        query_params['brand'] = selected_brands
    if selected_sort != 'curated':
        query_params['sort'] = selected_sort

    page_query = f"&{urlencode(query_params, doseq=True)}" if query_params else ''
    sort_query_params = {key: value for key, value in query_params.items() if key != 'sort'}
    sort_base_query = urlencode(sort_query_params, doseq=True) if sort_query_params else ''

    selected_max_price_value = filter_max_price
    if parsed_max_price is not None:
        selected_max_price_value = int(
            max(
                Decimal(str(filter_min_price)),
                min(parsed_max_price, Decimal(str(range_input_max)))
            )
        )

    product_ids = [p.id for p in products_page.object_list]
    stats_qs = ProductReview.objects.filter(product_id__in=product_ids, is_approved=True).values('product_id').annotate(
        avg_rating=Avg('rating'),
        review_count=Count('id')
    )
    product_stats = {
        item['product_id']: {
            'avg_rating': round(item['avg_rating'] or 0, 1),
            'review_count': item['review_count']
        }
        for item in stats_qs
    }

    return render(request, 'shop.html', {
        'products': products_page,
        'banners': banners,
        'special_offers': special_offers,
        'category_icons': category_icons,
        'category_data': category_data,
        'selected_category': selected_category,
        'selected_sub_category': selected_sub_category,
        'min_price': min_price or '',
        'max_price': max_price or '',
        'selected_rating': selected_rating,
        'selected_sort': selected_sort,
        'selected_brands': selected_brands,
        'designer_brands': designer_brands,
        'rating_options': rating_options,
        'filter_min_price': filter_min_price,
        'filter_max_price': filter_max_price,
        'range_input_max': range_input_max,
        'selected_max_price_value': selected_max_price_value,
        'sub_category_data': sub_category_data,
        'search_query': search_query,
        'wishlist_product_ids': wishlist_product_ids,
        'cart_product_ids': cart_product_ids,
        'product_stats': product_stats,
        'page_links': page_links,
        'page_query': page_query,
        'sort_base_query': sort_base_query,
    })


def product_search_api(request):
    """Return top product matches for the header search dropdown."""
    query = request.GET.get('q', '').strip()
    results = []

    if len(query) >= 2:
        products = Product.objects.filter(is_active=True).filter(
            Q(name__icontains=query) |
            Q(sku__icontains=query) |
            Q(brand__icontains=query) |
            Q(tags__icontains=query)
        ).order_by('name')[:8]

        for product in products:
            results.append({
                'id': product.id,
                'name': product.name,
                'price_display': f"{product.price:.2f}",
                'image_url': product.image.url if product.image else None,
                'url': reverse('product-details', args=[product.id]),
            })

    return JsonResponse({'results': results})
def shop_details(request): return render(request, 'shop-details.html')


@login_required(login_url='login')
def wishlist(request: HttpRequest) -> HttpResponse:
    wishlist_items = list(
        Wishlist.objects.filter(user=request.user)
        .select_related('product')
        .order_by('-added_at')
    )
    wishlist_count = len(wishlist_items)

    wishlist_product_ids = [item.product_id for item in wishlist_items]
    preferred_brands = sorted({
        (item.product.brand or '').strip()
        for item in wishlist_items
        if getattr(item.product, 'brand', '')
    })
    preferred_categories = sorted({
        item.product.category
        for item in wishlist_items
        if getattr(item.product, 'category', '')
    })

    tablet_wishlist_items = []
    for item in wishlist_items:
        product = item.product
        subtitle = (
            (product.brand or '').strip()
            or (product.get_category_display() if product.category else '')
            or ('In Stock' if product.stock > 0 else 'Out of Stock')
        )
        status_label = 'In Stock' if product.stock > 0 else 'Out of Stock'
        tablet_wishlist_items.append({
            'wishlist_id': item.id,
            'product_id': product.id,
            'name': product.name,
            'subtitle': subtitle,
            'price': product.price,
            'image_url': product.image.url if product.image else '',
            'detail_url': reverse('product-details', args=[product.id]),
            'remove_url': reverse('remove_from_wishlist', args=[item.id]),
            'move_to_cart_url': reverse('move_wishlist_to_cart', args=[item.id]),
            'is_in_stock': product.stock > 0,
            'status_label': status_label,
        })

    recommendation_pool = Product.objects.filter(is_active=True).exclude(id__in=wishlist_product_ids)
    prioritized_recommendations = recommendation_pool.none()
    if preferred_brands or preferred_categories:
        prioritized_recommendations = recommendation_pool.filter(
            Q(brand__in=preferred_brands) | Q(category__in=preferred_categories)
        ).order_by('-is_top_deal', '-rating', '-review_count', '-id')

    selected_recommendations = []
    selected_recommendation_ids = set()

    for product in prioritized_recommendations[:6]:
        selected_recommendations.append(product)
        selected_recommendation_ids.add(product.id)

    if len(selected_recommendations) < 6:
        fallback_recommendations = (
            recommendation_pool.exclude(id__in=selected_recommendation_ids)
            .order_by('-is_top_deal', '-rating', '-review_count', '-id')[:6 - len(selected_recommendations)]
        )
        for product in fallback_recommendations:
            selected_recommendations.append(product)

    wishlist_recommendations = [{
        'id': product.id,
        'name': product.name,
        'eyebrow': (
            (product.brand or '').strip()
            or (product.get_category_display() if product.category else '')
            or 'VibeMall Edit'
        ),
        'image_url': product.image.url if product.image else '',
        'detail_url': reverse('product-details', args=[product.id]),
    } for product in selected_recommendations]

    return render(request, 'wishlist.html', {
        'wishlist_items': wishlist_items,
        'wishlist_count': wishlist_count,
        'tablet_wishlist_items': tablet_wishlist_items,
        'wishlist_recommendations': wishlist_recommendations,
    })

def page_404(request): return render(request, '404.html', status=404)

def custom_404(request, exception=None):
    """Custom 404 error handler"""
    return render(request, '404.html', status=404)

def custom_500(request):
    """Custom 500 error handler"""
    return render(request, '500.html', status=500)

def order_tracking(request):return render(request, 'order-tracking.html')



def register_view(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        name = request.POST.get("name")
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")
        country_code = request.POST.get("country_code")
        mobile_number = request.POST.get("mobile_number")
        profile_image = request.FILES.get("profile_image")
        terms_accepted = request.POST.get("terms_accepted")

        # Validate terms acceptance
        if not terms_accepted:
            messages.error(request, "You must accept the Terms and Conditions to register.")
            return redirect("register")

        # Validate password strength
        if len(password) < 8:
            messages.error(request, "Password must be at least 8 characters long")
            return redirect("register")

        if password != confirm_password:
            messages.error(request, "Passwords do not match. Please ensure both passwords are identical.")
            return redirect("register")

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username already taken. Please choose a different username.")
            return redirect("register")

        if User.objects.filter(email=email).exists():
            messages.error(request, "Email already registered. Please use a different email or login.")
            return redirect("register")

        try:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password
            )
            user.first_name = name
            user.is_active = False
            user.save()

            profile = user.userprofile
            profile.country_code = country_code
            profile.mobile_number = mobile_number
            if profile_image:
                profile.profile_image = profile_image
            profile.save()

            # Send verification email
            send_verification_email(user, request)
            
            # Send welcome email with Terms & Conditions PDF
            from .email_utils import send_welcome_email_with_terms
            try:
                send_welcome_email_with_terms(user, request)
            except Exception as email_error:
                # Log error but don't fail registration
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Failed to send welcome email: {str(email_error)}")
            
            messages.success(request, "Account created! Please verify your email to activate your account.", extra_tags='success')
            return redirect("verify_email_sent")
        except Exception as e:
            messages.error(request, f"An error occurred during registration: {str(e)}")
            return redirect("register")

    return render(request, "register.html")


def send_verification_email(user, request):
    """Send email verification link"""
    uid = urlsafe_base64_encode(force_bytes(user.pk))
    token = default_token_generator.make_token(user)
    verify_url = request.build_absolute_uri(
        f"/verify-email/{uid}/{token}/"
    )

    subject = "Verify your VibeMall account"
    message = render_to_string('emails/verify_email.html', {
        'user': user,
        'verify_url': verify_url,
    })

    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email],
        fail_silently=False,
        html_message=message,
    )


def verify_email_sent(request):
    """Show verification sent page"""
    return render(request, 'verify_email_sent.html')


def verify_email(request, uidb64, token):
    """Verify user email and activate account"""
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)

        if default_token_generator.check_token(user, token):
            user.is_active = True
            user.save()
            messages.success(request, "Email verified successfully! You can log in now.")
            return redirect('login')

        messages.error(request, "Verification link is invalid or expired.")
        return redirect('register')
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, "Invalid verification link.")
        return redirect('register')

    return render(request, "register.html")


def logout_view(request):
    logout(request)
    return redirect("index")


# ===== PASSWORD RESET VIEWS =====

def password_reset_view(request):
    """Display password reset request form"""
    if request.method == 'POST':
        email = request.POST.get('email')
        from .models import PasswordResetLog
        ip = request.META.get('REMOTE_ADDR')
        try:
            user = User.objects.filter(email=email).first()
            if user is None:
                raise User.DoesNotExist
            # Generate token and uid
            from django.contrib.auth.tokens import default_token_generator
            from django.utils.encoding import force_bytes
            from django.utils.http import urlsafe_base64_encode
            uid = urlsafe_base64_encode(force_bytes(user.pk))
            token = default_token_generator.make_token(user)
            reset_url = request.build_absolute_uri(
                f"/password_reset_confirm/{uid}/{token}/"
            )
            request.session['reset_url'] = reset_url
            request.session['reset_email'] = email
            PasswordResetLog.objects.create(user=user, email=email, ip_address=ip, status='requested')
            # Send email
            from django.core.mail import send_mail
            subject = 'Password Reset Request'
            message = f"You requested a password reset. Click the link below to reset your password:\n{reset_url}\n\nIf you did not request this, please ignore this email."
            send_mail(subject, message, None, [email], fail_silently=False)
            messages.success(request, f"If an account exists, a reset link will be sent to your email.")
            return redirect('password_reset_done')
        except User.DoesNotExist:
            PasswordResetLog.objects.create(user=None, email=email, ip_address=ip, status='failed', reason='No account found')
            messages.error(request, "If an account exists, a reset link will be sent to your email.")
    
    return render(request, 'password_reset.html')

    
    return render(request, 'password_reset.html')


def password_reset_done_view(request):
    """Display message after password reset request"""
    email = request.session.get('reset_email', '')
    return render(request, 'password_reset_done.html', {'email': email})


def password_reset_confirm_view(request, uidb64, token):
    """Handle password reset with token verification"""
    from django.contrib.auth.tokens import default_token_generator
    from django.utils.http import urlsafe_base64_decode
    
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk=uid)
        
        from .models import PasswordResetLog
        ip = request.META.get('REMOTE_ADDR')
        if not default_token_generator.check_token(user, token):
            PasswordResetLog.objects.create(user=user, email=user.email, ip_address=ip, status='failed', reason='Token expired/invalid')
            messages.error(request, "Password reset link has expired or is invalid")
            return redirect('login')
        
        if request.method == 'POST':
            password1 = request.POST.get('new_password1')
            password2 = request.POST.get('new_password2')
            
            if not password1 or not password2:
                messages.error(request, "Both password fields are required")
                return render(request, 'password_reset_confirm.html')
            
            if password1 != password2:
                messages.error(request, "Passwords do not match")
                return render(request, 'password_reset_confirm.html')
            
            if len(password1) < 8:
                messages.error(request, "Password must be at least 8 characters")
                return render(request, 'password_reset_confirm.html')

                import re
                if not re.search(r'[A-Z]', password1):
                    messages.error(request, "Password must contain at least one uppercase letter")
                    return render(request, 'password_reset_confirm.html')
                if not re.search(r'[0-9]', password1):
                    messages.error(request, "Password must contain at least one number")
                    return render(request, 'password_reset_confirm.html')
                if not re.search(r'[^A-Za-z0-9]', password1):
                    messages.error(request, "Password must contain at least one special character")
                    return render(request, 'password_reset_confirm.html')
            
            user.set_password(password1)
            user.save()
            PasswordResetLog.objects.create(user=user, email=user.email, ip_address=ip, status='success', reason='Password reset')
            messages.success(request, "Password reset successfully!")
            return redirect('password_reset_complete')
        
        return render(request, 'password_reset_confirm.html', {'uidb64': uidb64, 'token': token})
    
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        messages.error(request, "Invalid password reset link")
        return redirect('login')


def password_reset_complete_view(request):
    """Display success message after password reset"""
    return render(request, 'password_reset_complete.html')


# ===== CART MANAGEMENT VIEWS =====

@login_required(login_url='login')
@require_POST
def add_to_cart(request: HttpRequest) -> JsonResponse:
    """Add product to cart or increase quantity"""
    product_id = request.POST.get('product_id')
    quantity = int(request.POST.get('quantity', 1))
    
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    
    try:
        product = Product.objects.get(id=product_id, is_active=True)

        if product.stock <= 0:
            if is_ajax:
                return JsonResponse({
                    'success': False,
                    'message': 'This item is out of stock'
                }, status=400)
            messages.error(request, f"{product.name} is out of stock")
            return redirect('product-details', product_id=product.id)

        cart_item, created = Cart.objects.get_or_create(
            user=request.user,
            product=product,
            defaults={'quantity': quantity}
        )
        
        if not created:
            new_qty = cart_item.quantity + quantity
            if new_qty > product.stock:
                new_qty = product.stock
            cart_item.quantity = new_qty
            cart_item.save()
        
        # Get total cart count
        cart_count = Cart.objects.filter(user=request.user).count()
        
        if is_ajax:
            return JsonResponse({
                'success': True,
                'message': f'{product.name} added to cart!',
                'cart_count': cart_count,
                'product_name': product.name
            })
        else:
            messages.success(request, f"{product.name} added to cart!")
            return redirect('cart')
    except Product.DoesNotExist:
        if is_ajax:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            }, status=404)
        else:
            messages.error(request, "Product not found")
            return redirect('shop')


@login_required(login_url='login')
def cart_summary(request: HttpRequest) -> JsonResponse:
    cart_items = Cart.objects.filter(user=request.user).select_related('product')
    cart_count = cart_items.count()
    cart_total_value = sum(item.get_total_price() for item in cart_items)
    cart_total = f"{cart_total_value:.2f}"
    mini_cart_html = render_to_string('partials/mini_cart.html', {
        'cart_items': cart_items,
        'cart_count': cart_count,
        'cart_total': cart_total,
    }, request=request)
    return JsonResponse({
        'cart_count': cart_count,
        'cart_total': cart_total,
        'mini_cart_html': mini_cart_html,
    })


@require_POST
def request_stock_notification(request, product_id):
    """Capture email to notify when a product is restocked."""
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'
    product = get_object_or_404(Product, id=product_id)

    email = request.POST.get('email') or (request.user.email if request.user.is_authenticated else '')
    if not email:
        message = 'Please provide an email so we can notify you.'
        if is_ajax:
            return JsonResponse({'success': False, 'message': message}, status=400)
        messages.error(request, message)
        return redirect('product-details', product_id=product.id)

    notification, created = ProductStockNotification.objects.get_or_create(
        product=product,
        email=email,
        defaults={'user': request.user if request.user.is_authenticated else None}
    )

    if not created and request.user.is_authenticated and notification.user is None:
        notification.user = request.user
        notification.save(update_fields=['user'])

    message = 'We will email you as soon as this item is back in stock.' if created else 'You are already subscribed for a restock alert.'

    if is_ajax:
        return JsonResponse({'success': True, 'message': message})

    messages.success(request, message)
    next_url = request.META.get('HTTP_REFERER') or reverse('product-details', args=[product.id])
    return redirect(next_url)


@login_required(login_url='login')
def remove_from_cart(request, cart_id):
    """Remove product from cart"""
    try:
        cart_item = Cart.objects.get(id=cart_id, user=request.user)
        product_name = cart_item.product.name
        cart_item.delete()
        messages.success(request, f"{product_name} removed from cart!")
    except Cart.DoesNotExist:
        messages.error(request, "Item not found in cart")
    
    return redirect('cart')


@login_required(login_url='login')
@require_POST
def ajax_toggle_cart(request, product_id):
    """AJAX endpoint to toggle product in/out of cart"""
    try:
        product = Product.objects.get(id=product_id)
        if product.stock <= 0:
            return JsonResponse({
                'success': False,
                'message': 'This item is out of stock',
                'in_cart': False
            }, status=400)
        cart_item = Cart.objects.filter(user=request.user, product=product).first()
        
        if cart_item:
            # Remove from cart
            cart_item.delete()
            in_cart = False
            action = 'removed'
            message = f'{product.name} removed from cart'
        else:
            # Add to cart
            Cart.objects.create(user=request.user, product=product, quantity=1)
            in_cart = True
            action = 'added'
            message = f'{product.name} added to cart!'
        
        cart_count = Cart.objects.filter(user=request.user).count()
        
        return JsonResponse({
            'success': True,
            'in_cart': in_cart,
            'action': action,
            'message': message,
            'cart_count': cart_count
        })
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)
    except Exception:
        logger.exception("Failed to toggle cart item for user_id=%s product_id=%s", request.user.id, product_id)
        return JsonResponse({
            'success': False,
            'message': 'Unable to update cart right now. Please try again.'
        }, status=500)


@login_required(login_url='login')
@require_POST
def update_cart_quantity(request, cart_id):
    """Update cart item quantity"""
    try:
        quantity = int(request.POST.get('quantity', 1))
        cart_item = Cart.objects.get(id=cart_id, user=request.user)
        
        if quantity <= 0:
            cart_item.delete()
            messages.success(request, "Item removed from cart")
        else:
            cart_item.quantity = quantity
            cart_item.save()
            messages.success(request, "Quantity updated")
    except (Cart.DoesNotExist, ValueError):
        messages.error(request, "Error updating quantity")
    
    return redirect('cart')


# ===== WISHLIST MANAGEMENT VIEWS =====

@login_required(login_url='login')
@require_POST
def add_to_wishlist(request: HttpRequest) -> JsonResponse:
    """Add product to wishlist"""
    product_id = request.POST.get('product_id')
    
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        wishlist_item, created = Wishlist.objects.get_or_create(
            user=request.user,
            product=product
        )
        
        if created:
            messages.success(request, f"{product.name} added to wishlist!")
        else:
            messages.info(request, f"{product.name} is already in your wishlist")
        
        return redirect('wishlist')
    except Product.DoesNotExist:
        messages.error(request, "Product not found")
        return redirect('shop')


@login_required(login_url='login')
def remove_from_wishlist(request, wishlist_id):
    """Remove product from wishlist"""
    try:
        wishlist_item = Wishlist.objects.get(id=wishlist_id, user=request.user)
        product_name = wishlist_item.product.name
        wishlist_item.delete()
        messages.success(request, f"{product_name} removed from wishlist!")
    except Wishlist.DoesNotExist:
        messages.error(request, "Item not found in wishlist")
    
    return redirect('wishlist')


@login_required(login_url='login')
def move_wishlist_to_cart(request, wishlist_id):
    """Move item from wishlist to cart"""
    try:
        wishlist_item = Wishlist.objects.get(id=wishlist_id, user=request.user)
        product = wishlist_item.product
        
        # Add to cart
        cart_item, created = Cart.objects.get_or_create(
            user=request.user,
            product=product,
            defaults={'quantity': 1}
        )
        
        # Remove from wishlist
        wishlist_item.delete()
        
        messages.success(request, f"{product.name} moved to cart!")
    except Wishlist.DoesNotExist:
        messages.error(request, "Item not found in wishlist")
    
    return redirect('cart')


@login_required(login_url='login')
@require_POST
def submit_review(request, product_id):
    """Submit a product review with images (requires admin approval)"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        
        rating = int(request.POST.get('rating', 5))
        comment = request.POST.get('comment', '').strip()
        
        # Auto-fill user details if logged in
        if request.user.is_authenticated:
            name = request.user.get_full_name() or request.user.username
            email = request.user.email
        else:
            name = request.POST.get('name', '').strip()
            email = request.POST.get('email', '').strip()
        
        # Check if user has purchased this product (verified purchase)
        is_verified_purchase = False
        order_id = request.POST.get('order_id')
        if request.user.is_authenticated and order_id:
            is_verified_purchase = Order.objects.filter(
                id=order_id,
                user=request.user,
                order_status='DELIVERED',
                items__product=product
            ).exists()
        
        # Create review (not approved by default)
        review = ProductReview.objects.create(
            product=product,
            user=request.user if request.user.is_authenticated else None,
            rating=rating,
            name=name,
            email=email,
            comment=comment,
            is_approved=False,  # Admin must approve
            is_verified_purchase=is_verified_purchase
        )
        
        # Handle image uploads (up to 5 images)
        images = request.FILES.getlist('review_images')
        for i, image in enumerate(images[:5]):  # Limit to 5 images
            ReviewImage.objects.create(review=review, image=image)
        
        messages.success(request, "Thank you for your review! It will be visible after admin approval.")
        
    except Product.DoesNotExist:
        messages.error(request, "Product not found")
    except ValueError:
        messages.error(request, "Invalid rating value")
    
    # Redirect to referrer or homepage
    return redirect(request.META.get('HTTP_REFERER', 'index'))


@login_required
@require_POST
def vote_review(request, review_id):
    """AJAX endpoint to vote on review helpfulness"""
    try:
        review = get_object_or_404(ProductReview, id=review_id, is_approved=True)
        is_helpful = request.POST.get('is_helpful') == 'true'
        
        # Check if user already voted
        existing_vote = ReviewVote.objects.filter(review=review, user=request.user).first()
        
        if existing_vote:
            # Update existing vote
            if existing_vote.is_helpful != is_helpful:
                # Decrement old count
                if existing_vote.is_helpful:
                    review.helpful_count = max(0, review.helpful_count - 1)
                else:
                    review.not_helpful_count = max(0, review.not_helpful_count - 1)
                
                # Increment new count
                if is_helpful:
                    review.helpful_count += 1
                else:
                    review.not_helpful_count += 1
                
                existing_vote.is_helpful = is_helpful
                existing_vote.save()
                review.save()
                
                return JsonResponse({
                    'success': True,
                    'message': 'Vote updated',
                    'helpful_count': review.helpful_count,
                    'not_helpful_count': review.not_helpful_count
                })
            else:
                return JsonResponse({'success': True, 'message': 'Already voted'})
        else:
            # Create new vote
            ReviewVote.objects.create(review=review, user=request.user, is_helpful=is_helpful)
            
            # Update counts
            if is_helpful:
                review.helpful_count += 1
            else:
                review.not_helpful_count += 1
            review.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Vote recorded',
                'helpful_count': review.helpful_count,
                'not_helpful_count': review.not_helpful_count
            })
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)


@login_required(login_url='login')
@require_POST
def mobile_review_prompt_dismiss(request):
    """Dismiss mobile delivered-order review prompt for the active login session."""
    request.session['mobile_review_prompt_seen'] = True
    request.session.modified = True
    return JsonResponse({'success': True})


@login_required(login_url='login')
@require_POST
def mobile_review_prompt_submit(request, product_id):
    """Submit quick mobile review for delivered product and mark prompt as seen for session."""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
    except Product.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Product not found.'}, status=404)

    delivered_match = OrderItem.objects.filter(
        order__user=request.user,
        order__order_status='DELIVERED',
        product=product,
    ).exists()

    if not delivered_match:
        return JsonResponse({'success': False, 'message': 'This review is only available for delivered products.'}, status=403)

    # Prevent duplicate review submissions from this one-time prompt.
    if ProductReview.objects.filter(product=product, user=request.user).exists():
        request.session['mobile_review_prompt_seen'] = True
        request.session.modified = True
        return JsonResponse({'success': True, 'message': 'Review already submitted for this product.'})

    raw_rating = request.POST.get('rating', '5')
    comment = (request.POST.get('comment') or '').strip()

    try:
        rating = int(raw_rating)
    except (TypeError, ValueError):
        rating = 5

    if rating < 1:
        rating = 1
    if rating > 5:
        rating = 5

    ProductReview.objects.create(
        product=product,
        user=request.user,
        rating=rating,
        name=request.user.get_full_name().strip() or request.user.username,
        email=request.user.email or f'{request.user.username}@example.com',
        comment=comment,
        is_approved=True,
        is_verified_purchase=True,
    )

    request.session['mobile_review_prompt_seen'] = True
    request.session.modified = True

    product.refresh_from_db(fields=['review_count', 'rating'])
    return JsonResponse({
        'success': True,
        'message': 'Thanks for sharing your experience.',
        'review_count': int(product.review_count or 0),
        'rating': float(product.rating or 0),
    })


@login_required
@require_POST
@login_required
def submit_question(request, product_id):
    """Submit a product question - requires admin to answer before showing on page"""
    try:
        product = get_object_or_404(Product, id=product_id, is_active=True)
        question_text = request.POST.get('question', '').strip()
        
        if not question_text:
            messages.error(request, "Please enter a question.")
            return redirect('product-details', product_id=product_id)
        
        ProductQuestion.objects.create(
            product=product,
            user=request.user,
            question=question_text,
            is_answered=False,
            is_approved=False  # Requires admin approval and answer
        )
        
        messages.success(request, "Your question has been submitted. It will appear once our team answers it!")
        
    except Exception as e:
        messages.error(request, f"Error submitting question: {str(e)}")
    
    return redirect('product-details', product_id=product_id)



from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import UserProfile




from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import UserProfile

@login_required(login_url='login')
def profile_view(request):

    if request.method == "POST":
        profile, created = UserProfile.objects.get_or_create(user=request.user)

        request.user.first_name = request.POST.get("first_name")
        request.user.last_name = request.POST.get("last_name")
        request.user.email = request.POST.get("email")

        profile.country_code = request.POST.get("country_code")
        profile.mobile_number = request.POST.get("mobile_number")

        if request.FILES.get("profile_image"):
            profile.profile_image = request.FILES.get("profile_image")

        request.user.save()
        profile.save()

        messages.success(request, "Profile saved successfully")

        # 🔥 THIS LINE IS REQUIRED
        return redirect('profile')

    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Get loyalty points and transactions
    loyalty_account, _ = LoyaltyPoints.objects.get_or_create(user=request.user)
    points_transactions = PointsTransaction.objects.filter(user=request.user).order_by('-created_at')[:10]
    
    return render(request, 'profile.html', {
        'profile': profile,
        'loyalty_account': loyalty_account,
        'points_transactions': points_transactions
    })


@login_required(login_url='login')
def api_profile_stats(request):
    """Return live profile stats and recent orders."""
    from django.http import JsonResponse
    from django.urls import reverse

    user = request.user
    orders = Order.objects.filter(user=user).prefetch_related('items__product')

    total_orders = orders.count()
    delivered_orders = orders.filter(order_status='DELIVERED').count()
    cancelled_orders = orders.filter(order_status='CANCELLED').count()
    pending_orders = orders.filter(order_status__in=['PENDING', 'PROCESSING', 'SHIPPED']).count()

    def to_absolute_url(url_value):
        if not url_value:
            return ''
        if url_value.startswith('http://') or url_value.startswith('https://'):
            return url_value
        if not url_value.startswith('/'):
            url_value = f'/{url_value}'
        return request.build_absolute_uri(url_value)

    recent_orders = orders.order_by('-created_at')[:5]
    recent_list = []
    for order in recent_orders:
        order_items = list(order.items.all())
        first_item = order_items[0] if order_items else None
        total_qty = sum(item.quantity for item in order_items) if order_items else 0

        product_name = first_item.product_name if first_item else 'Order Item'
        item_summary = product_name
        if total_qty > 1:
            item_summary = f"{product_name} +{total_qty - 1} more item(s)"

        product_image = ''
        if first_item:
            if first_item.product_image:
                product_image = to_absolute_url(first_item.product_image)
            elif first_item.product and first_item.product.image:
                product_image = to_absolute_url(first_item.product.image.url)

        recent_list.append({
            'order_number': order.order_number,
            'status': order.order_status,
            'status_display': order.get_order_status_display(),
            'total_amount': float(order.total_amount),
            'created_at': order.created_at.strftime('%d %b %Y'),
            'created_at_full': order.created_at.strftime('%d %b %Y, %I:%M %p'),
            'detail_url': reverse('order_details', args=[order.order_number]),
            'product_name': product_name,
            'item_summary': item_summary,
            'product_image': product_image,
            'item_count': total_qty,
        })

    return JsonResponse({
        'total_orders': total_orders,
        'delivered_orders': delivered_orders,
        'pending_orders': pending_orders,
        'cancelled_orders': cancelled_orders,
        'recent_orders': recent_list,
    })





# Add this to the end of views.py

@login_required(login_url='login')
def add_product(request):
    # Check if user is admin (username = 'admin')
    if request.user.username != 'VibeMall':
        messages.error(request, 'Access denied. Only VibeMall user can add products.')
        return redirect('index')
    
    if request.method == 'POST':
        try:
            # Get basic form data
            name = request.POST.get('name')
            price = normalize_decimal_input(request.POST.get('price'))
            old_price = normalize_decimal_input(request.POST.get('old_price'))
            stock = request.POST.get('stock')
            rating = request.POST.get('rating', 0)
            review_count = request.POST.get('review_count', 0)
            is_active = request.POST.get('is_active') == 'on'
            is_top_deal = request.POST.get('is_top_deal') == 'on'
            
            # Get new fields
            category = request.POST.get('category')
            sku = request.POST.get('sku', '')
            brand = request.POST.get('brand', '')
            description = request.POST.get('description', '')
            weight = request.POST.get('weight', '')
            color = request.POST.get('color', '')
            # Handle multiple size selections from checkboxes
            size_list = request.POST.getlist('size')
            size = ', '.join(size_list) if size_list else ''
            
            # Get images
            image = request.FILES.get('image')
            descriptionImage = request.FILES.get('descriptionImage')
            gallery_images = request.FILES.getlist('gallery_images')

            # Process color variant images with optional color labels
            variant_images = []
            variant_colors = request.POST.getlist('variant_colors')
            variant_files = request.FILES.getlist('variant_images')
            for idx, variant_file in enumerate(variant_files, 1):
                color_value = (variant_colors[idx - 1].strip() if idx - 1 < len(variant_colors) else '').strip()
                variant_images.append({
                    'image': variant_file,
                    'color': color_value,
                })
            
            # Create product
            product = Product.objects.create(
                name=name,
                price=price,
                old_price=old_price if old_price else None,
                discount_percent=calc_discount_percent(price, old_price),
                stock=stock,
                rating=rating,
                review_count=review_count,
                is_active=True,
                is_top_deal=False,
                image=image,
                descriptionImage=descriptionImage,
                category=category if category else None,
                sku=sku,
                brand=brand,
                description=description,
                weight=weight,
                color=color,
                size=size
            )
            
            # Add gallery images if provided
            current_order = 0
            for idx, gallery_image in enumerate(gallery_images, start=1):
                current_order = idx
                ProductImage.objects.create(
                    product=product,
                    image=gallery_image,
                    order=idx,
                    is_active=True
                )

            # Add color variant images if provided
            for idx, variant in enumerate(variant_images, start=current_order + 1):
                ProductImage.objects.create(
                    product=product,
                    image=variant['image'],
                    color=variant['color'],
                    order=idx,
                    is_active=True
                )
            
            success_message = f'Product "{product.name}" added successfully with {len(gallery_images)} gallery images.'
            if variant_images:
                success_message += f' {len(variant_images)} color variant image(s) saved.'
            messages.success(request, success_message)
            return redirect('add_product')
            
        except Exception as e:
            messages.error(request, f'Error adding product: {str(e)}')
    
    # Get all products for display
    products = Product.objects.all().order_by('-id')[:30]  # Show last 30 products
    
    return render(request, 'add_product.html', {
        'products': products,
    })

# ===== AJAX WISHLIST VIEW =====

@login_required(login_url='login')
def ajax_add_to_wishlist(request, product_id):
    """AJAX endpoint to toggle product in wishlist (add/remove)"""
    if request.method == 'POST':
        try:
            product = Product.objects.get(id=product_id, is_active=True)
            
            # Check if already in wishlist
            wishlist_item = Wishlist.objects.filter(user=request.user, product=product).first()
            
            if wishlist_item:
                # Remove from wishlist
                wishlist_item.delete()
                return JsonResponse({
                    'success': True,
                    'action': 'removed',
                    'message': f'{product.name} removed from wishlist',
                    'in_wishlist': False
                })
            else:
                # Add to wishlist
                Wishlist.objects.create(user=request.user, product=product)
                return JsonResponse({
                    'success': True,
                    'action': 'added',
                    'message': f'{product.name} added to wishlist!',
                    'in_wishlist': True
                })
                
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            }, status=404)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required(login_url='login')
def check_wishlist(request, product_id):
    """Check if product is in user's wishlist"""
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            product = Product.objects.get(id=product_id)
            in_wishlist = Wishlist.objects.filter(
                user=request.user,
                product=product
            ).exists()
            
            return JsonResponse({
                'success': True,
                'in_wishlist': in_wishlist
            })
        except Product.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': 'Product not found'
            }, status=404)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


@login_required(login_url='login')
@require_POST
def buy_now(request, product_id):
    """Buy Now - Store product in session and redirect to checkout"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)
        quantity = int(request.POST.get('quantity', 1))
        
        if quantity < 1:
            return JsonResponse({
                'success': False,
                'message': 'Invalid quantity'
            }, status=400)
        
        # Check stock
        if product.stock < quantity:
            return JsonResponse({
                'success': False,
                'message': f'Only {product.stock} items available in stock'
            }, status=400)
        
        # Store in session for checkout
        request.session['buy_now_item'] = {
            'product_id': product.id,
            'quantity': quantity,
            'price': float(product.price)
        }
        request.session.modified = True
        
        return JsonResponse({
            'success': True,
            'message': f'{product.name} - redirecting to checkout!',
            'redirect_url': '/checkout/'
        })
        
    except Product.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Product not found'
        }, status=404)
    except ValueError:
        return JsonResponse({
            'success': False,
            'message': 'Invalid quantity value'
        }, status=400)


# ===== BANNER MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_banners(request):
    """Admin Banner List"""
    banners = Banner.objects.all().order_by('order')
    context = {
        'banners': banners,
    }
    return render(request, 'admin_panel/banners.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_banner(request):
    """Admin Add Banner"""
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            subtitle = request.POST.get('subtitle', '')
            badge_text = request.POST.get('badge_text', '')
            button_text = request.POST.get('button_text', '')
            button_style = request.POST.get('button_style', 'none')
            banner_type = request.POST.get('banner_type', 'LARGE')
            page_type = request.POST.get('page_type', 'HOME')
            link_url = request.POST.get('link_url', '#')
            background_color = request.POST.get('background_color', '')
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            image = request.FILES.get('image')
            
            banner = Banner.objects.create(
                title=title,
                subtitle=subtitle,
                badge_text=badge_text,
                button_text=button_text,
                button_style=button_style,
                banner_type=banner_type,
                page_type=page_type,
                link_url=link_url,
                background_color=background_color,
                order=order,
                is_active=is_active,
                image=image
            )
            
            messages.success(request, f'Banner "{banner.title}" added successfully!')
            return redirect('admin_banners')
            
        except Exception as e:
            messages.error(request, f'Error adding banner: {str(e)}')
    
    return render(request, 'admin_panel/add_banner.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_banner(request, banner_id):
    """Admin Edit Banner"""
    banner = get_object_or_404(Banner, id=banner_id)
    
    if request.method == 'POST':
        try:
            banner.title = request.POST.get('title')
            banner.subtitle = request.POST.get('subtitle', '')
            banner.badge_text = request.POST.get('badge_text', '')
            banner.button_text = request.POST.get('button_text', '')
            banner.button_style = request.POST.get('button_style', 'none')
            banner.banner_type = request.POST.get('banner_type', 'LARGE')
            banner.page_type = request.POST.get('page_type', 'HOME')
            banner.link_url = request.POST.get('link_url', '#')
            banner.background_color = request.POST.get('background_color', '')
            banner.order = request.POST.get('order', 0)
            banner.is_active = request.POST.get('is_active') == 'on'
            
            if 'image' in request.FILES:
                banner.image = request.FILES['image']
            
            banner.save()
            
            messages.success(request, f'Banner "{banner.title}" updated successfully!')
            return redirect('admin_banners')
            
        except Exception as e:
            messages.error(request, f'Error updating banner: {str(e)}')
    
    context = {
        'banner': banner,
    }
    return render(request, 'admin_panel/edit_banner.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_banner(request, banner_id):
    """Admin Delete Banner"""
    banner = get_object_or_404(Banner, id=banner_id)
    
    if request.method == 'POST':
        banner_title = banner.title
        banner.delete()
        messages.success(request, f'Banner "{banner_title}" deleted successfully!')
    
    return redirect('admin_banners')


# ===== SLIDER MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_sliders(request):
    """Admin Slider List"""
    sliders = Slider.objects.all().order_by('order', '-id')
    context = {
        'sliders': sliders,
    }
    return render(request, 'admin_panel/sliders.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_slider(request):
    """Admin Add Slider"""
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            subtitle = request.POST.get('subtitle', '')
            description = request.POST.get('description', '')
            top_button_text = request.POST.get('top_button_text', '')
            top_button_url = request.POST.get('top_button_url', '#')
            order = request.POST.get('order', 0)
            is_active = request.POST.get('is_active') == 'on'
            image = request.FILES.get('image')
            
            slider = Slider.objects.create(
                title=title,
                subtitle=subtitle,
                description=description,
                top_button_text=top_button_text,
                top_button_url=top_button_url,
                order=int(order) if order else 0,
                is_active=is_active,
                image=image
            )
            
            messages.success(request, f'Slider "{slider.title}" added successfully!')
            return redirect('admin_sliders')
            
        except Exception as e:
            messages.error(request, f'Error adding slider: {str(e)}')
    
    return render(request, 'admin_panel/add_slider.html')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_slider(request, slider_id):
    """Admin Edit Slider"""
    slider = get_object_or_404(Slider, id=slider_id)
    
    if request.method == 'POST':
        try:
            slider.title = request.POST.get('title')
            slider.subtitle = request.POST.get('subtitle', '')
            slider.description = request.POST.get('description', '')
            slider.top_button_text = request.POST.get('top_button_text', '')
            slider.top_button_url = request.POST.get('top_button_url', '#')
            order = request.POST.get('order', 0)
            slider.order = int(order) if order else 0
            slider.is_active = request.POST.get('is_active') == 'on'
            
            if 'image' in request.FILES:
                slider.image = request.FILES['image']
            
            slider.save()
            
            messages.success(request, f'Slider "{slider.title}" updated successfully!')
            return redirect('admin_sliders')
            
        except Exception as e:
            messages.error(request, f'Error updating slider: {str(e)}')
    
    context = {
        'slider': slider,
    }
    return render(request, 'admin_panel/edit_slider.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_slider(request, slider_id):
    """Admin Delete Slider"""
    slider = get_object_or_404(Slider, id=slider_id)
    
    if request.method == 'POST':
        slider_title = slider.title
        slider.delete()
        messages.success(request, f'Slider "{slider_title}" deleted successfully!')
    
    return redirect('admin_sliders')


# ===== QUESTIONS MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_questions(request):
    """Admin Questions List"""
    questions = ProductQuestion.objects.select_related('product', 'user').order_by('-created_at')
    
    # Filter by status if provided
    status = request.GET.get('status')
    if status == 'pending':
        questions = questions.filter(is_approved=False, answer__isnull=True)
    elif status == 'answered':
        questions = questions.filter(is_approved=False, answer__isnull=False)
    elif status == 'approved':
        questions = questions.filter(is_approved=True)
    
    context = {
        'questions': questions,
    }
    return render(request, 'admin_panel/questions.html', context)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_approve_question(request, question_id):
    """Admin Approve/Answer Question"""
    question = get_object_or_404(ProductQuestion, id=question_id)
    
    if request.method == 'POST':
        answer = request.POST.get('answer', '')
        
        if answer:
            question.answer = answer
            question.answered_by = request.user
            question.answered_at = timezone.now()
            question.is_approved = True
            question.save()
            
            messages.success(request, 'Question answered and approved!')
        else:
            messages.error(request, 'Please provide an answer')
    
    return redirect('admin_questions')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_question(request, question_id):
    """Admin Delete Question"""
    question = get_object_or_404(ProductQuestion, id=question_id)
    
    if request.method == 'POST':
        question.delete()
        messages.success(request, 'Question deleted successfully!')
    
    return redirect('admin_questions')

@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_adjust_rating(request, product_id):
    """Admin Adjust Product Rating (increment/decrement)"""
    try:
        product = get_object_or_404(Product, id=product_id)
        action = request.POST.get('action')  # 'increment' or 'decrement'
        
        if action == 'increment':
            if product.rating < 5:
                product.rating = min(5, product.rating + 0.1)
                product.save()
                messages.success(request, f'Rating increased to {product.rating:.1f}')
            else:
                messages.warning(request, 'Rating is already at maximum (5.0)')
        elif action == 'decrement':
            if product.rating > 0:
                product.rating = max(0, product.rating - 0.1)
                product.save()
                messages.success(request, f'Rating decreased to {product.rating:.1f}')
            else:
                messages.warning(request, 'Rating is already at minimum (0.0)')
        
        return JsonResponse({
            'success': True,
            'rating': round(product.rating, 1)
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)

@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_approve_review(request, review_id):
    """Admin Approve Review"""
    from django.http import JsonResponse
    from django.db.models import Avg

    def refresh_product_rating(product):
        """Recalculate product rating and review count from approved reviews"""
        approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
        if approved_reviews.exists():
            product.review_count = approved_reviews.count()
            avg_rating = approved_reviews.aggregate(Avg('rating'))['rating__avg']
            product.rating = round(avg_rating, 1) if avg_rating else 0
        else:
            product.review_count = 0
            product.rating = 0
        product.save(update_fields=['review_count', 'rating'])
    
    review = get_object_or_404(ProductReview, id=review_id)
    
    # If POST request with edited data (from modal)
    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        import json
        try:
            data = json.loads(request.body)
            edited_rating = data.get('rating')
            edited_comment = data.get('comment')
            
            # Validate and update rating
            if edited_rating and 1 <= int(edited_rating) <= 5:
                review.rating = int(edited_rating)
            
            # Update comment if provided
            if edited_comment and edited_comment.strip():
                review.comment = edited_comment.strip()
            
            review.is_approved = True
            review.save()
            
            refresh_product_rating(review.product)
            
            return JsonResponse({
                'success': True,
                'message': 'Review approved successfully!'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            }, status=400)
    
    # Regular approval without editing
    review.is_approved = True
    review.save()
    
    refresh_product_rating(review.product)
    
    messages.success(request, 'Review approved successfully!')
    return redirect('admin_reviews')

@login_required(login_url='login')
@staff_member_required(login_url='login')
@require_POST
def admin_delete_review(request, review_id):
    """Admin Delete Review"""
    review = get_object_or_404(ProductReview, id=review_id)
    product = review.product
    review.delete()
    
    from django.db.models import Avg
    approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
    if approved_reviews.exists():
        product.review_count = approved_reviews.count()
        avg_rating = approved_reviews.aggregate(Avg('rating'))['rating__avg']
        product.rating = round(avg_rating, 1) if avg_rating else 0
    else:
        product.review_count = 0
        product.rating = 0
    product.save(update_fields=['review_count', 'rating'])
    
    messages.success(request, 'Review deleted successfully!')
    return redirect('admin_reviews')

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_review_details(request, review_id):
    """Get Review Details as JSON"""
    from django.http import JsonResponse
    
    try:
        review = get_object_or_404(ProductReview, id=review_id)
        
        # Get review images
        review_images = []
        for img in review.images.all():
            review_images.append(request.build_absolute_uri(img.image.url))
        
        # Build response data
        data = {
            'id': review.id,
            'product': {
                'id': review.product.id,
                'name': review.product.name,
                'image': request.build_absolute_uri(review.product.image.url) if review.product.image else None,
            },
            'user': {
                'name': review.name,
                'email': review.email,
            },
            'rating': review.rating,
            'comment': review.comment,
            'is_verified_purchase': review.is_verified_purchase,
            'is_approved': review.is_approved,
            'created_at': review.created_at.strftime('%B %d, %Y at %I:%M %p'),
            'images': review_images,
            'helpful_count': review.helpful_count,
            'not_helpful_count': review.not_helpful_count,
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_review(request, product_id):
    """Admin Add Review to Product"""
    from django.http import JsonResponse
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            # Get admin review details
            rating = int(request.POST.get('rating', 5))
            comment = request.POST.get('comment', '').strip()
            name = request.POST.get('name', request.user.get_full_name() or request.user.username)
            email = request.POST.get('email', request.user.email)
            
            # Validate
            if not 1 <= rating <= 5:
                return JsonResponse({'success': False, 'message': 'Rating must be between 1 and 5'}, status=400)
            
            # Create review (admin reviews are auto-approved)
            review = ProductReview.objects.create(
                product=product,
                user=request.user,
                rating=rating,
                comment=comment,
                name=name,
                email=email,
                is_approved=True,
                is_verified_purchase=False  # Admin reviews are not verified purchases
            )

            review_images = request.FILES.getlist('review_images')
            for image in review_images:
                ReviewImage.objects.create(review=review, image=image)
            
            # Update product review count and rating
            product.review_count = ProductReview.objects.filter(product=product, is_approved=True).count()
            
            # Recalculate average rating
            from django.db.models import Avg
            approved_reviews = ProductReview.objects.filter(product=product, is_approved=True)
            if approved_reviews.exists():
                avg_rating = approved_reviews.aggregate(Avg('rating'))['rating__avg']
                product.rating = round(avg_rating, 1)
            
            product.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Review added successfully!',
                'review_id': review.id
            })
        
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)}, status=500)
    
    return JsonResponse({'success': False, 'message': 'Invalid request'}, status=400)


# ===== ORDER CONFIRMATION & PAYMENT VIEWS =====

@login_required(login_url='login')
def order_confirmation(request, order_id):
    """Order Confirmation Page"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    # If payment method is Razorpay and not paid, redirect to payment page
    if order.payment_method == 'RAZORPAY' and order.payment_status != 'PAID' and order.payment_status != 'PROCESSING':
        messages.warning(request, 'Please complete your payment to confirm the order.')
        return redirect('razorpay_payment', order_id=order.id)
    
    # Calculate loyalty points earned (₹1 = 33 points, 1 point = ₹0.03)
    loyalty_points_earned = int(order.total_amount * 33)

    context = {
        'order': order,
        'order_items': order.items.all(),
        'loyalty_points_earned': loyalty_points_earned,
    }
    return render(request, 'order_confirmation.html', context)

@login_required(login_url='login')
def razorpay_payment(request, order_id):
    """Razorpay Payment Page with OSrder Creation"""
    try:
        import razorpay
    except ModuleNotFoundError as exc:
        messages.error(
            request,
            'Payment service dependency is missing on server. Please contact support and retry shortly.'
        )
        logger.error(f'Razorpay import failed in razorpay_payment: {exc}')
        return redirect('order_confirmation', order_id=order_id)
    
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Get Razorpay keys from settings
    razorpay_key = getattr(settings, 'RAZORPAY_KEY_ID', '')
    razorpay_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
    
    if not razorpay_key or not razorpay_secret:
        messages.error(request, 'Payment gateway not configured')
        return redirect('order_confirmation', order_id=order.id)
    
    try:
        # Create Razorpay client
        client = razorpay.Client(auth=(razorpay_key, razorpay_secret))

        # Amount in paise (ensure int to avoid float issues)
        amount_paise = int(order.total_amount * 100)
        if amount_paise <= 0:
            messages.error(request, 'Invalid order amount for payment')
            return redirect('order_confirmation', order_id=order.id)

        # Create Razorpay order if not already created
        if not order.razorpay_order_id:
            razorpay_order = client.order.create({
                'amount': amount_paise,
                'currency': 'INR',
                'receipt': order.order_number,
                'notes': {
                    'order_id': str(order.id),
                    'customer': order.user.username,
                    'email': order.user.email
                }
            })
            
            # Save Razorpay order ID
            order.razorpay_order_id = razorpay_order['id']
            order.save(update_fields=['razorpay_order_id'])
        
        context = {
            'order': order,
            'razorpay_key': razorpay_key,
            'razorpay_order_id': order.razorpay_order_id,
            'order_amount': amount_paise,
            'first_item': order.items.select_related('product').first(),
        }
        return render(request, 'razorpay_payment.html', context)
        
    except Exception as e:
        messages.error(request, f'Error creating payment: {str(e)}')
        return redirect('order_confirmation', order_id=order.id)


@login_required(login_url='login')
@require_POST
def razorpay_payment_success(request):
    """Handle Razorpay Payment Success"""
    try:
        try:
            import razorpay
        except ModuleNotFoundError as exc:
            logger.error(f'Razorpay import failed in razorpay_payment_success: {exc}')
            return JsonResponse({'success': False, 'message': 'Payment service unavailable. Please try again later.'}, status=503)
        
        order_id = request.POST.get('order_id')
        payment_id = request.POST.get('payment_id')
        signature = request.POST.get('signature')
        
        order = get_object_or_404(Order, id=order_id, user=request.user)
        
        # Verify signature
        razorpay_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
        
        client = razorpay.Client(auth=(getattr(settings, 'RAZORPAY_KEY_ID', ''), razorpay_secret))
        
        params_dict = {
            'razorpay_order_id': order.razorpay_order_id,
            'razorpay_payment_id': payment_id,
            'razorpay_signature': signature
        }
        
        try:
            client.utility.verify_payment_signature(params_dict)
            
            # Payment verified
            order.payment_status = 'PAID'
            order.order_status = 'PROCESSING'
            order.razorpay_payment_id = payment_id
            order.razorpay_signature = signature
            order.save(update_fields=['payment_status', 'order_status', 'razorpay_payment_id', 'razorpay_signature'])
            
            # Auto-process approval (fraud detection & auto-approve)
            order.auto_process_approval()
            
            # Clear cart
            Cart.objects.filter(user=request.user).delete()
            
            # Send order confirmation email
            try:
                email_sent = send_order_confirmation_email(order)
                if not email_sent:
                    logger.warning(f"Razorpay order confirmation email failed silently for order {order.order_number}")
                    messages.warning(request, 'Order placed successfully, but confirmation email could not be delivered. Please check your email address or contact support.')
            except Exception as email_exc:
                logger.error(f"Razorpay order confirmation email exception for order {order.order_number}: {email_exc}", exc_info=True)
                messages.warning(request, 'Order placed successfully, but confirmation email failed to send. Please contact support.')
            
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': True, 'redirect': reverse('order_confirmation', args=[order.id])})
            
            # Check if order needs approval
            if order.approval_status == 'PENDING_APPROVAL':
                messages.warning(request, 'Payment successful! Your order is pending approval due to security checks.')
            else:
                messages.success(request, 'Payment successful! Your order is being processed.')
            
            return redirect('order_confirmation', order_id=order.id)
            
        except Exception as e:
            order.payment_status = 'FAILED'
            order.save(update_fields=['payment_status'])
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                return JsonResponse({'success': False, 'message': 'Payment verification failed'}, status=400)
            messages.error(request, 'Payment verification failed')
            return redirect('checkout')
            
    except Exception as e:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'message': f'Error processing payment: {str(e)}'}, status=400)
        messages.error(request, f'Error processing payment: {str(e)}')
        return redirect('checkout')


@login_required(login_url='login')
def razorpay_payment_cancel(request, order_id):
    """Handle Razorpay Payment Cancellation"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Update order status
    from django.utils import timezone
    order.payment_status = 'FAILED'
    order.order_status = 'CANCELLED'
    order.save()
    
    # Send cancellation email
    try:
        from django.template.loader import render_to_string
        from django.core.mail import EmailMultiAlternatives
        
        subject = f'Order Cancelled - {order.order_number}'
        
        # Get site URL
        site_url = request.build_absolute_uri('/').rstrip('/')
        
        # Render HTML email
        html_content = render_to_string('emails/order_cancelled.html', {
            'order': order,
            'cancelled_at': timezone.now(),
            'site_url': site_url,
        })
        
        # Plain text version
        text_content = f'''
Dear {order.user.get_full_name() or order.user.username},

Your order {order.order_number} has been cancelled as per your request.

Order Details:
- Order Number: {order.order_number}
- Total Amount: ₹{order.total_amount}
- Status: Cancelled

Thank you for your understanding.

Best regards,
FashioHub Team
        '''
        
        # Send email
        email = EmailMultiAlternatives(subject, text_content, _get_from_email(), [order.user.email])
        email.attach_alternative(html_content, "text/html")
        email.send()
        
    except Exception:
        logger.exception("Email sending failed while payment cancellation for order_id=%s", order.id)
    
    messages.warning(request, 'Payment was cancelled. You can try again from your orders.')
    return redirect('checkout')


from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
@require_POST
def razorpay_webhook(request):
    """Handle Razorpay Webhook Events - CSRF exempt because webhook is from external service
    
    Security: CSRF exemption is secure here because:
    1. Webhook signature is verified using RAZORPAY_WEBHOOK_SECRET
    2. External payment gateway cannot send CSRF tokens
    3. Signature verification provides the protection needed
    """
    import razorpay
    import json
    from django.views.decorators.csrf import csrf_exempt
    
    webhook_secret = getattr(settings, 'RAZORPAY_WEBHOOK_SECRET', '')
    webhook_signature = request.headers.get('X-Razorpay-Signature', '')
    webhook_body = request.body
    
    if not webhook_secret:
        return JsonResponse({'status': 'webhook not configured'}, status=400)
    
    # Verify webhook signature
    try:
        # Verify signature
        import hmac
        import hashlib
        
        expected_signature = hmac.new(
            webhook_secret.encode('utf-8'),
            webhook_body,
            hashlib.sha256
        ).hexdigest()
        
        if expected_signature != webhook_signature:
            return JsonResponse({'status': 'invalid signature'}, status=400)
            
    except Exception as e:
        return JsonResponse({'status': 'verification failed'}, status=400)
    
    # Process webhook data
    try:
        data = json.loads(webhook_body)
        event = data.get('event')
        
        if event == 'payment.captured':
            # Payment successful
            payment = data['payload']['payment']['entity']
            order_id = payment['notes'].get('order_id')
            
            if order_id:
                order = Order.objects.get(id=order_id)
                order.payment_status = 'PAID'
                order.order_status = 'PROCESSING'
                order.razorpay_payment_id = payment['id']
                order.save()
                
                # Clear user's cart
                Cart.objects.filter(user=order.user).delete()
                
        elif event == 'payment.failed':
            # Payment failed
            payment = data['payload']['payment']['entity']
            order_id = payment['notes'].get('order_id')
            
            if order_id:
                order = Order.objects.get(id=order_id)
                order.payment_status = 'FAILED'
                order.save()
        
        return JsonResponse({'status': 'ok'})
        
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


@login_required(login_url='login')
@user_passes_test(lambda u: u.is_staff)
def razorpay_refund(request, order_id):
    """Process Razorpay Refund for an Order"""
    order = get_object_or_404(Order, id=order_id)
    
    # Check if order can be refunded
    if order.payment_status != 'PAID':
        messages.error(request, 'Only paid orders can be refunded.')
        return redirect('admin_order_details', order_id=order.id)
    
    if order.payment_method != 'RAZORPAY':
        messages.error(request, 'Only Razorpay orders can be refunded via this method.')
        return redirect('admin_order_details', order_id=order.id)
    
    if not order.razorpay_payment_id:
        messages.error(request, 'No Razorpay payment ID found for this order. Cannot process refund.')
        return redirect('admin_order_details', order_id=order.id)
    
    # Get refund amount (can be partial or full)
    try:
        refund_amount = request.POST.get('refund_amount', '').strip()
        if refund_amount:
            refund_amount_value = Decimal(str(refund_amount))
        else:
            # Full refund
            refund_amount_value = order.total_amount
    except (ValueError, InvalidOperation):
        messages.error(request, 'Invalid refund amount. Please enter a valid number.')
        return redirect('admin_order_details', order_id=order.id)
    
    # Use the centralized refund helper
    refund_notes = {
        'reason': request.POST.get('refund_reason', 'Admin-initiated refund'),
        'order_number': order.order_number,
        'refunded_by': request.user.username
    }
    
    refund_success, refund_error = _create_razorpay_refund(
        order.razorpay_payment_id,
        refund_amount_value,
        notes=refund_notes
    )
    
    if refund_success:
        # Update order status only on successful refund
        order.payment_status = 'REFUNDED'
        order.order_status = 'CANCELLED'
        refund_note_text = f"Refund Amount: ₹{refund_amount_value}\nReason: {refund_notes.get('reason', 'N/A')}"
        if order.admin_notes:
            order.admin_notes = f"{refund_note_text}\n\n{order.admin_notes}"
        else:
            order.admin_notes = refund_note_text
        order.save()
        messages.success(request, f'✓ Refund processed successfully! Amount: ₹{refund_amount_value}')
    else:
        # Refund failed - keep order in PAID status so retry is possible
        messages.error(request, f'Refund failed: {refund_error}')
    
    return redirect('admin_order_details', order_id=order.id)


@login_required(login_url='login')
def resell_order(request, order_id):
    """Create a Resell Order from existing order"""
    try:
        original_order = get_object_or_404(Order, id=order_id, user=request.user)
        
        # Create new order with is_resell=True
        new_order = Order.objects.create(
            user=request.user,
            subtotal=original_order.subtotal,
            tax=original_order.tax,
            shipping_cost=original_order.shipping_cost,
            total_amount=original_order.total_amount,
            shipping_address=original_order.shipping_address,
            billing_address=original_order.billing_address,
            payment_method=original_order.payment_method,
            is_resell=True
        )
        
        # Copy order items
        for item in original_order.items.all():
            OrderItem.objects.create(
                order=new_order,
                product=item.product,
                product_name=item.product_name,
                product_price=item.product_price,
                product_image=item.product_image,
                quantity=item.quantity,
                size=item.size,
                color=item.color,
                margin_amount=item.margin_amount if item.margin_amount is not None else (item.product.margin if item.product else None)
            )
        
        messages.success(request, f'Resell order created! Order #: {new_order.order_number}')
        return redirect('checkout')
        
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('profile')


def order_list(request: HttpRequest) -> HttpResponse:
    """Display list of user's orders"""
    if not request.user.is_authenticated:
        return redirect('login')

    returned_statuses = [
        'REQUESTED',
        'APPROVED',
        'PICKUP_SCHEDULED',
        'RECEIVED',
        'QC_PENDING',
        'QC_PASSED',
        'QC_FAILED',
        'REFUND_PENDING',
        'WRONG_RETURN',
        'REFUNDED',
        'REPLACED',
    ]
    returned_orders_q = Q(return_requests__status__in=returned_statuses) | Q(payment_status='REFUNDED')
    status_definitions = [
        {
            'value': 'ALL',
            'label': 'All Orders',
            'query': None,
        },
        {
            'value': 'PENDING',
            'label': 'Pending',
            'query': Q(order_status__in=['PENDING', 'PROCESSING']) & ~returned_orders_q,
        },
        {
            'value': 'IN_TRANSIT',
            'label': 'In Transit',
            'query': Q(order_status='SHIPPED') & ~returned_orders_q,
        },
        {
            'value': 'DELIVERED',
            'label': 'Delivered',
            'query': Q(order_status='DELIVERED') & ~returned_orders_q,
        },
        {
            'value': 'RETURNED',
            'label': 'Returned',
            'query': returned_orders_q,
        },
    ]
    mobile_status_definitions = [
        {
            'value': 'ALL',
            'label': 'All',
            'query': None,
        },
        {
            'value': 'AWAITING',
            'label': 'Pending',
            'query': Q(order_status='PENDING') & ~returned_orders_q,
        },
        {
            'value': 'PROCESSING',
            'label': 'Processing',
            'query': Q(order_status='PROCESSING') & ~returned_orders_q,
        },
        {
            'value': 'SHIPPED',
            'label': 'Shipped',
            'query': Q(order_status='SHIPPED') & ~returned_orders_q,
        },
        {
            'value': 'DELIVERED',
            'label': 'Delivered',
            'query': Q(order_status='DELIVERED') & ~returned_orders_q,
        },
        {
            'value': 'CANCELLED',
            'label': 'Cancelled',
            'query': Q(order_status='CANCELLED') | Q(payment_status='REFUNDED'),
        },
    ]
    filter_query_map = {item['value']: item['query'] for item in status_definitions}
    filter_query_map.update({item['value']: item['query'] for item in mobile_status_definitions})

    all_orders = (
        Order.objects.filter(user=request.user)
        .prefetch_related('items__product', 'return_requests')
        .order_by('-created_at')
    )

    search_query = request.GET.get('q', '').strip()
    active_status = (request.GET.get('status') or 'ALL').upper()
    valid_statuses = set(filter_query_map.keys())
    if active_status not in valid_statuses:
        active_status = 'ALL'

    filtered_orders = all_orders
    active_status_query = filter_query_map.get(active_status)
    if active_status_query is not None:
        filtered_orders = filtered_orders.filter(active_status_query).distinct()

    if search_query:
        filtered_orders = filtered_orders.filter(
            Q(order_number__icontains=search_query) |
            Q(items__product_name__icontains=search_query)
        ).distinct()

    total_orders_count = all_orders.count()
    filtered_orders_count = filtered_orders.count()
    paginator = Paginator(filtered_orders, 10)
    page_obj = paginator.get_page(request.GET.get('page') or 1)

    def build_orders_url(**updates):
        params = request.GET.copy()
        page_updated = 'page' in updates

        for key, value in updates.items():
            if key == 'status' and value == 'ALL':
                params.pop('status', None)
                continue

            if key == 'page' and str(value) == '1':
                params.pop('page', None)
                continue

            if value in (None, ''):
                params.pop(key, None)
                continue

            params[key] = str(value)

        if not page_updated:
            params.pop('page', None)

        query_string = params.urlencode()
        base_url = reverse('order_list')
        return f'{base_url}?{query_string}' if query_string else base_url

    status_tabs = []
    for status_item in status_definitions:
        query = status_item['query']
        count = total_orders_count if query is None else all_orders.filter(query).distinct().count()
        status_tabs.append({
            'label': status_item['label'],
            'value': status_item['value'],
            'count': count,
            'is_active': active_status == status_item['value'],
            'url': build_orders_url(status=status_item['value'], page=None),
        })

    mobile_status_tabs = []
    for status_item in mobile_status_definitions:
        query = status_item['query']
        count = total_orders_count if query is None else all_orders.filter(query).distinct().count()
        mobile_status_tabs.append({
            'label': status_item['label'],
            'value': status_item['value'],
            'count': count,
            'is_active': active_status == status_item['value'],
            'url': build_orders_url(status=status_item['value'], page=None),
        })

    order_cards = []
    for order in page_obj.object_list:
        order_items = list(order.items.all())
        latest_return = next(iter(order.return_requests.all()), None)
        primary_item = order_items[0] if order_items else None
        image_items = order_items[:3]
        return_in_progress = bool(latest_return and latest_return.status not in ['REJECTED', 'CANCELLED'])
        return_resolved = bool(
            latest_return and latest_return.status in ['REFUNDED', 'REPLACED']
        ) or order.payment_status == 'REFUNDED'

        state_key = 'delivered'
        headline_label = order.get_order_status_display()
        tag_label = 'Completed Order'
        tag_tone = 'neutral'
        amount_label = 'Grand Total'
        amount_value = order.total_amount
        mobile_badge_label = order.get_order_status_display()
        mobile_badge_tone = order.order_status.lower()
        mobile_primary_action_label = 'View Details'
        mobile_primary_action_url = reverse('order_details', args=[order.order_number])
        actions = [{
            'label': 'View Details',
            'url': reverse('order_details', args=[order.order_number]),
            'style': 'primary',
        }]

        if return_resolved:
            state_key = 'returned'
            headline_label = 'Returned'
            tag_label = 'Refund Issued' if order.payment_status == 'REFUNDED' else (latest_return.get_status_display() if latest_return else 'Returned')
            tag_tone = 'soft'
            amount_label = 'Refund Amount'
            amount_value = (
                getattr(latest_return, 'refund_amount_net', None)
                or getattr(latest_return, 'refund_amount', None)
                or order.total_amount
            )
            mobile_badge_label = 'Cancelled' if order.order_status == 'CANCELLED' else 'Returned'
            mobile_badge_tone = 'cancelled'
            mobile_primary_action_label = 'Details'
            reorder_url = reverse('shop')
            if primary_item and primary_item.product and getattr(primary_item.product, 'slug', ''):
                reorder_url = reverse('product_detail', args=[primary_item.product.slug])
            actions = [{
                'label': 'Reorder Items',
                'url': reorder_url,
                'style': 'primary',
            }]
        elif return_in_progress:
            state_key = 'returned'
            headline_label = 'Return in Review'
            tag_label = latest_return.get_status_display()
            tag_tone = 'soft'
            mobile_badge_label = 'Cancelled'
            mobile_badge_tone = 'cancelled'
            mobile_primary_action_label = 'Details'
            actions.append({
                'label': 'Return Status',
                'url': reverse('return_status', args=[latest_return.id]),
                'style': 'secondary',
            })
        elif order.order_status == 'SHIPPED':
            state_key = 'in-transit'
            headline_label = 'In Transit'
            tag_label = 'Tracking Live' if order.tracking_number else 'On the Way'
            tag_tone = 'gold'
            mobile_badge_label = 'Shipped'
            mobile_badge_tone = 'shipped'
            mobile_primary_action_label = 'Track Order'
            mobile_primary_action_url = reverse('order_tracking', args=[order.order_number])
            actions.append({
                'label': 'Track Order',
                'url': reverse('order_tracking', args=[order.order_number]),
                'style': 'secondary',
            })
        elif order.order_status == 'DELIVERED':
            state_key = 'delivered'
            headline_label = 'Delivered'
            tag_label = 'Completed Order'
            tag_tone = 'gold'
            mobile_badge_label = 'Delivered'
            mobile_badge_tone = 'delivered'
            actions.append({
                'label': 'Track Order',
                'url': reverse('order_tracking', args=[order.order_number]),
                'style': 'secondary',
            })
        elif order.order_status == 'PROCESSING':
            state_key = 'processing'
            headline_label = 'Processing'
            tag_label = 'Awaiting Atelier'
            tag_tone = 'rose'
            mobile_badge_label = 'Processing'
            mobile_badge_tone = 'processing'
            actions.append({
                'label': 'Support',
                'url': '#vmOrdersHelp',
                'style': 'secondary',
            })
        elif order.order_status == 'PENDING':
            state_key = 'pending'
            headline_label = 'Pending'
            tag_label = 'Awaiting Payment' if order.payment_status == 'PENDING' else 'Order Received'
            tag_tone = 'rose'
            mobile_badge_label = 'Pending'
            mobile_badge_tone = 'processing'
            actions.append({
                'label': 'Support',
                'url': '#vmOrdersHelp',
                'style': 'secondary',
            })
        elif order.order_status == 'CANCELLED':
            state_key = 'cancelled'
            headline_label = 'Cancelled'
            tag_label = 'Order Cancelled'
            tag_tone = 'soft'
            mobile_badge_label = 'Cancelled'
            mobile_badge_tone = 'cancelled'
            mobile_primary_action_label = 'Details'
            actions.append({
                'label': 'Continue Shopping',
                'url': reverse('shop'),
                'style': 'secondary',
            })

        order_cards.append({
            'order': order,
            'primary_item': primary_item,
            'image_items': image_items,
            'more_items_count': max(len(order_items) - len(image_items), 0),
            'mobile_items_label': (
                '1 item'
                if len(order_items) == 1
                else f"+ {len(order_items)} items"
            ),
            'headline': f"{headline_label} {order.created_at.strftime('%b %d, %Y')}",
            'tag_label': tag_label,
            'tag_tone': tag_tone,
            'amount_label': amount_label,
            'amount_value': amount_value,
            'mobile_badge_label': mobile_badge_label,
            'mobile_badge_tone': mobile_badge_tone,
            'mobile_primary_action_label': mobile_primary_action_label,
            'mobile_primary_action_url': mobile_primary_action_url,
            'state_key': state_key,
            'actions': actions,
            'is_dimmed': state_key == 'returned',
        })

    context = {
        'orders': page_obj.object_list,
        'order_cards': order_cards,
        'page_obj': page_obj,
        'status_tabs': status_tabs,
        'mobile_status_tabs': mobile_status_tabs,
        'active_status': active_status,
        'search_query': search_query,
        'total_orders_count': total_orders_count,
        'filtered_orders_count': filtered_orders_count,
        'has_active_filters': bool(search_query or active_status != 'ALL'),
        'clear_filters_url': reverse('order_list'),
        'clear_search_url': build_orders_url(q='', page=None),
        'previous_page_url': build_orders_url(page=page_obj.previous_page_number()) if page_obj.has_previous() else '',
        'next_page_url': build_orders_url(page=page_obj.next_page_number()) if page_obj.has_next() else '',
    }
    return render(request, 'order_list.html', context)


def _calculate_return_refund_amount(return_request):
    subtotal = Decimal('0')
    total_items_qty = 0
    returned_qty = 0

    for item in return_request.items.select_related('order_item'):
        subtotal += item.order_item.product_price * item.quantity
        returned_qty += item.quantity

    for order_item in return_request.order.items.all():
        total_items_qty += order_item.quantity

    order_subtotal = return_request.order.subtotal or Decimal('0')
    order_tax = return_request.order.tax or Decimal('0')
    order_shipping = return_request.order.shipping_cost or Decimal('0')

    tax_refund = Decimal('0')
    shipping_refund = Decimal('0')

    if order_subtotal > 0:
        tax_refund = (subtotal / order_subtotal) * order_tax

    if returned_qty >= total_items_qty and total_items_qty > 0:
        shipping_refund = order_shipping

    return (subtotal + tax_refund + shipping_refund).quantize(Decimal('0.01'))


def _return_eligibility(order):
    if order.order_status != 'DELIVERED':
        return False, 'Returns are allowed only after delivery.', None, []

    if not order.delivery_date:
        return False, 'Delivery date is not available yet.', None, []

    deadline = order.delivery_date + timedelta(days=RETURN_WINDOW_DAYS)
    if timezone.now() > deadline:
        return False, f'Return window closed on {deadline.strftime("%d %b %Y")}.', deadline, []

    if order.return_requests.count() >= MAX_RETURN_ATTEMPTS:
        return False, 'Maximum return attempts reached for this order.', deadline, []

    eligible_items = []
    for item in order.items.select_related('product'):
        if item.product and item.product.category in NON_RETURNABLE_CATEGORIES:
            continue
        eligible_items.append(item)

    if not eligible_items:
        return False, 'No returnable items found in this order.', deadline, []

    return True, '', deadline, eligible_items


def _paid_orders_qs():
    return Order.objects.filter(
        Q(payment_status='PAID') | Q(payment_method='COD', order_status='DELIVERED')
    )


def _order_item_profit(qs):
    profit_expr = ExpressionWrapper(
        Coalesce('margin_amount', 'product__margin', Value(0)) * F('quantity'),
        output_field=DecimalField(max_digits=12, decimal_places=2)
    )
    return qs.aggregate(total=Sum(profit_expr))['total'] or Decimal('0.00')


def _return_fee_profit(start_date=None, end_date=None, user=None):
    qs = ReturnRequest.objects.filter(status='REFUNDED')
    if user is not None:
        qs = qs.filter(user=user)
    if start_date:
        qs = qs.filter(resolved_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(resolved_at__date__lte=end_date)
    return qs.aggregate(total=Sum('refund_fee'))['total'] or Decimal('0.00')


def _cancel_eligibility(order):
    if order.order_status not in ['PENDING', 'PROCESSING']:
        return False, 'Cancellation is allowed only before shipping.', None

    deadline = order.created_at + timedelta(hours=CANCEL_WINDOW_HOURS)
    if timezone.now() > deadline:
        return False, f'Cancellation window closed at {deadline.strftime("%d %b %Y, %I:%M %p")}.', deadline

    return True, '', deadline


def _log_return_history(return_request, old_status, new_status, user=None, notes=''):
    ReturnHistory.objects.create(
        return_request=return_request,
        old_status=old_status,
        new_status=new_status,
        changed_by=user,
        notes=notes
    )


def _send_return_notification(return_request, subject, message, template_name='emails/return_status_update.html'):
    recipients = [return_request.user.email] if return_request.user.email else []
    if not recipients:
        return

    status_url = ''
    try:
        status_url = reverse('return_status', args=[return_request.id])
    except Exception:
        status_url = ''

    site_url = getattr(settings, 'SITE_URL', 'http://127.0.0.1:8000').rstrip('/')
    full_status_url = f"{site_url}{status_url}" if status_url else site_url

    text_content = f"""{message}

Return Number: {return_request.return_number}
Order Number: {return_request.order.order_number}
Status: {return_request.get_status_display()}

View: {full_status_url}
"""

    try:
        html_content = render_to_string(template_name, {
            'return_request': return_request,
            'message': message,
            'status_label': return_request.get_status_display(),
            'status_url': full_status_url,
        })
        email = EmailMultiAlternatives(subject, text_content, _get_from_email(), recipients)
        email.attach_alternative(html_content, "text/html")
        email.send(fail_silently=True)
    except Exception:
        pass


def _send_admin_return_notification(return_request, request):
    admin_email = _get_admin_chat_email()
    if not admin_email:
        return

    site_url = request.build_absolute_uri('/').rstrip('/')
    admin_return_url = f"{site_url}{reverse('admin_return_detail', args=[return_request.id])}"
    subject = f"New Return Request - {return_request.return_number}"

    bank_lines = ''
    if return_request.refund_method == 'BANK':
        bank_lines = (
            f"Bank Name: {return_request.bank_name}\n"
            f"Account Name: {return_request.bank_account_name}\n"
            f"Account Number: {return_request.bank_account_number}\n"
            f"IFSC: {return_request.bank_ifsc}\n"
        )

    text_content = f"""New return request submitted.

Return Number: {return_request.return_number}
Order Number: {return_request.order.order_number}
Customer: {return_request.user.get_full_name() or return_request.user.username}
Reason: {return_request.get_reason_display()}
Status: {return_request.get_status_display()}
Refund Method: {return_request.refund_method}
{bank_lines}
Open: {admin_return_url}
"""

    try:
        html_content = render_to_string('emails/admin_return_request.html', {
            'return_request': return_request,
            'admin_return_url': admin_return_url,
        })
        email = EmailMultiAlternatives(subject, text_content, _get_from_email(), [admin_email])
        email.attach_alternative(html_content, "text/html")
        email.send(fail_silently=True)
    except Exception:
        pass


def _process_refund(return_request, amount, refund_method=''):
    order = return_request.order
    refund_method = refund_method or return_request.refund_method or order.payment_method
    refund_method = (refund_method or '').strip().upper()
    if refund_method in {'VIBEMALL WALLET', 'VIBEMALL_WALLET', 'WALLET'}:
        refund_method = 'WALLET'

    if amount is None:
        amount = return_request.refund_amount if return_request.refund_amount is not None else _calculate_return_refund_amount(return_request)

    try:
        amount = Decimal(str(amount))
    except Exception:
        amount = Decimal('0.00')

    fee = Decimal('20.00') if amount > 0 else Decimal('0.00')
    net_amount = max(amount - fee, Decimal('0.00'))

    refund_success = False
    refund_notes = ''

    if refund_method == 'RAZORPAY':
        if not order.razorpay_payment_id:
            refund_notes = 'Razorpay payment id missing for this order.'
        else:
            refund_success, refund_notes = _create_razorpay_refund(
                payment_id=order.razorpay_payment_id,
                amount=net_amount,
                notes={
                    'return_id': str(return_request.id),
                    'order_number': order.order_number,
                },
            )
    elif refund_method == 'WALLET':
        profile, _ = UserProfile.objects.get_or_create(user=return_request.user)
        profile.wallet_balance = (profile.wallet_balance or Decimal('0.00')) + net_amount
        profile.save(update_fields=['wallet_balance'])
        refund_success = True
    else:
        refund_success = True

    return_request.refund_amount = amount
    return_request.refund_fee = fee
    return_request.refund_amount_net = net_amount
    return_request.refund_method = refund_method
    if refund_notes:
        return_request.admin_notes = f"{refund_notes}\n{return_request.admin_notes}".strip()

    if refund_success and order.payment_status == 'PAID':
        order.payment_status = 'REFUNDED'
        order.save(update_fields=['payment_status'])

    return refund_success, refund_notes


def _process_cancellation_refund(order, cancel_request):
    refund_method = (cancel_request.refund_method or '').strip().upper()
    if refund_method in {'VIBEMALL WALLET', 'VIBEMALL_WALLET', 'WALLET'}:
        refund_method = 'WALLET'

    if order.payment_method == 'COD' or order.payment_status != 'PAID':
        return False, 'Refund not applicable for COD or unpaid orders.'

    amount = order.total_amount
    refund_success = False
    refund_notes = ''

    if refund_method == 'RAZORPAY' and order.razorpay_payment_id:
        refund_success, refund_notes = _create_razorpay_refund(
            payment_id=order.razorpay_payment_id,
            amount=amount,
            notes={
                'cancel_id': str(cancel_request.id),
                'order_number': order.order_number,
            },
        )
    elif refund_method == 'WALLET':
        profile, _ = UserProfile.objects.get_or_create(user=order.user)
        profile.wallet_balance = (profile.wallet_balance or Decimal('0.00')) + amount
        profile.save(update_fields=['wallet_balance'])
        refund_success = True
    elif refund_method == 'BANK':
        refund_success = True
        refund_notes = 'Bank refund pending'
    else:
        refund_success = True
        refund_notes = 'Manual refund pending'

    if refund_success:
        cancel_request.refund_amount = amount
        order.payment_status = 'REFUNDED'
        order.save(update_fields=['payment_status'])

    return refund_success, refund_notes


def _create_razorpay_refund(payment_id, amount, notes=None):
    """
    Create a Razorpay refund with proper validation and error handling.
    
    Args:
        payment_id: Razorpay payment ID (must start with 'pay_')
        amount: Refund amount as Decimal
        notes: Optional dict of notes to attach to refund
    
    Returns:
        (success: bool, error_message: str)
        - If success=True, error_message is empty ('')
        - If success=False, error_message contains specific reason
    """
    notes = notes or {}

    # Validate payment ID
    payment_id = (payment_id or '').strip()
    if not payment_id:
        return False, 'Razorpay payment ID is missing for this order.'
    
    if not payment_id.startswith('pay_'):
        return False, f'Invalid Razorpay payment ID format: {payment_id}'

    # Validate credentials
    razorpay_key = getattr(settings, 'RAZORPAY_KEY_ID', '')
    razorpay_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
    if not razorpay_key or not razorpay_secret:
        return False, 'Razorpay API keys are not configured in system settings.'

    # Validate amount
    try:
        amount_decimal = Decimal(str(amount or 0))
    except (ValueError, InvalidOperation):
        return False, f'Invalid refund amount: {amount}'

    refund_paisa = int((amount_decimal * Decimal('100')).quantize(Decimal('1')))
    if refund_paisa <= 0:
        return False, 'Refund amount must be greater than zero.'

    try:
        import razorpay
        from razorpay.errors import BadRequestError, NoDataError, ServerError, SignatureVerificationError

        client = razorpay.Client(auth=(razorpay_key, razorpay_secret))
        
        # Fetch payment to check if it exists and get remaining refundable amount
        try:
            payment = client.payment.fetch(payment_id)
        except NoDataError:
            return False, f'Payment {payment_id} not found in Razorpay. Please verify the payment ID.'
        except BadRequestError as e:
            return False, f'Invalid payment ID: {str(e)[:100]}'
        except Exception as e:
            return False, f'Error fetching payment details: {str(e)[:100]}'

        captured_amount = int(payment.get('amount') or 0)
        already_refunded = int(payment.get('amount_refunded') or 0)
        remaining_refundable = max(captured_amount - already_refunded, 0)

        # Check if already fully refunded
        if remaining_refundable <= 0:
            return True, 'Payment is already fully refunded on Razorpay.'

        # Cap refund to remaining refundable amount
        if refund_paisa > remaining_refundable:
            refund_paisa = remaining_refundable

        # Process refund
        try:
            refund_result = client.payment.refund(payment_id, {
                'amount': refund_paisa,
                'notes': notes,
            })
            return True, ''
        except BadRequestError as e:
            error_msg = str(e)
            if 'invalid request' in error_msg.lower():
                return False, 'Invalid refund request. Please check payment amount and try again.'
            elif 'already refunded' in error_msg.lower():
                return False, 'Payment has already been refunded.'
            else:
                return False, f'Refund request failed: {error_msg[:100]}'
        except ServerError as e:
            return False, f'Razorpay server error. Please try again later. ({str(e)[:50]})'
        except Exception as e:
            return False, f'Refund processing error: {str(e)[:100]}'
    
    except ImportError as ie:
        import sys
        import logging
        import os
        import site
        
        logger = logging.getLogger(__name__)
        
        # DEEP DEBUG LOGGING
        debug_info = {
            'executable': sys.executable,
            'version': sys.version,
            'prefix': sys.prefix,
            'base_prefix': getattr(sys, 'base_prefix', 'N/A'),
            'is_venv': hasattr(sys, 'real_prefix') or (hasattr(sys, 'base_prefix') and sys.base_prefix != sys.prefix),
            'site_packages': site.getsitepackages(),
            'user_site': site.getusersitepackages(),
            'cwd': os.getcwd(),
            'sys_path_first_5': sys.path[:5],
        }
        
        # Log everything
        logger.error('='*70)
        logger.error('RAZORPAY IMPORT FAILED - DIAGNOSTIC INFO')
        logger.error('='*70)
        for key, value in debug_info.items():
            logger.error(f'{key}: {value}')
        logger.error(f'ImportError: {ie}')
        logger.error('='*70)
        
        # Try to list site-packages to verify razorpay exists
        if debug_info['site_packages']:
            sp = debug_info['site_packages'][0]
            if os.path.exists(sp):
                razorpay_path = os.path.join(sp, 'razorpay')
                logger.error(f'Razorpay exists in {sp}: {os.path.exists(razorpay_path)}')
                if os.path.exists(razorpay_path):
                    contents = os.listdir(razorpay_path)
                    logger.error(f'Razorpay contents: {contents[:5]}')
        
        # Provide specific instruction
        return False, f'Razorpay SDK not found. Python: {sys.executable[:30]}... Run: pip install razorpay'
    except Exception as exc:
        return False, f'Unexpected error: {str(exc)[:100]}'


@login_required(login_url='login')
@require_POST
def validate_upi_id(request):
    upi_id = (request.POST.get('upi_id') or '').strip().lower()
    if not upi_id:
        return JsonResponse({'valid': False, 'message': 'UPI ID is required.'}, status=400)

    if not re.fullmatch(r"[a-z0-9._-]{2,256}@[a-z]{2,64}", upi_id):
        return JsonResponse({'valid': False, 'message': 'UPI ID format is invalid.'}, status=400)

    valid, name, message = _verify_upi_with_razorpay(upi_id, logger=logger)
    if not valid:
        return JsonResponse({'valid': False, 'message': message or 'UPI ID not found.'}, status=400)

    return JsonResponse({'valid': True, 'name': name})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_razorpay_health(request):
    key_id = getattr(settings, 'RAZORPAY_KEY_ID', '')
    key_secret = getattr(settings, 'RAZORPAY_KEY_SECRET', '')
    return JsonResponse({
        'ok': bool(key_id and key_secret),
        'has_key_id': bool(key_id),
        'has_key_secret': bool(key_secret),
    })


@login_required(login_url='login')
@require_POST
def lookup_ifsc(request):
    ifsc_code = request.POST.get('ifsc', '').strip()
    valid, bank, branch, message = _lookup_ifsc_details(ifsc_code)
    if not valid:
        return JsonResponse({'valid': False, 'message': message or 'IFSC details not found.'}, status=400)
    return JsonResponse({'valid': True, 'bank': bank, 'branch': branch})


def track_order_page(request):
    """Order tracking search page"""
    order_number = ''

    if request.method == 'POST':
        order_number = request.POST.get('order_number', '').strip()
        if order_number:
            query_params = urlencode({'order_number': order_number})
            return redirect(f"{reverse('track_order')}?{query_params}")
        else:
            messages.error(request, 'Please enter an order number')

    if request.method == 'GET':
        order_number = request.GET.get('order_number', '').strip()

    def normalize_order_number(value: str) -> str:
        normalized = re.sub(r'[^A-Za-z0-9]', '', (value or '').strip())
        return normalized.upper()

    if order_number:
        if not request.user.is_authenticated:
            login_query = urlencode({'next': request.get_full_path()})
            return redirect(f"{reverse('login')}?{login_query}")

        try:
            normalized_order_number = normalize_order_number(order_number)
            order = Order.objects.get(order_number__iexact=normalized_order_number, user=request.user)
            order_items = OrderItem.objects.filter(order=order)
            return_request_obj = order.return_requests.order_by('-requested_at').first()
            return_history = None
            if return_request_obj:
                return_history = return_request_obj.history.select_related('changed_by')

            return render(request, 'order_tracking.html', {
                'order': order,
                'order_items': order_items,
                'return_request': return_request_obj,
                'return_history': return_history,
            })
        except Order.DoesNotExist:
            messages.error(request, 'Order not found')

    return render(request, 'track_order.html')


def order_details(request, order_number):
    """Display detailed view of a specific order"""
    if not request.user.is_authenticated:
        return redirect('login')

    normalized_order_number = re.sub(r'[^A-Za-z0-9]', '', (order_number or '').strip()).upper()
    
    try:
        order = Order.objects.get(order_number__iexact=normalized_order_number, user=request.user)
        order_items = OrderItem.objects.filter(order=order)
        is_return_eligible, return_reason, return_deadline, eligible_items = _return_eligibility(order)
        active_return = order.return_requests.order_by('-requested_at').first()
        cancel_request = getattr(order, 'cancellation_request', None)
        cancel_eligible, cancel_reason, cancel_deadline = _cancel_eligibility(order)
        return render(request, 'order_details.html', {
            'order': order,
            'order_items': order_items,
            'is_return_eligible': is_return_eligible,
            'return_reason': return_reason,
            'return_deadline': return_deadline,
            'return_items_eligible': eligible_items,
            'active_return': active_return,
            'cancel_request': cancel_request,
            'cancel_eligible': cancel_eligible,
            'cancel_reason': cancel_reason,
            'cancel_deadline': cancel_deadline,
        })
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('order_list')


def order_tracking(request, order_number):
    """Display beautiful timeline tracking for a specific order"""
    if not request.user.is_authenticated:
        return redirect('login')

    normalized_order_number = re.sub(r'[^A-Za-z0-9]', '', (order_number or '').strip()).upper()
    
    try:
        order = Order.objects.get(order_number__iexact=normalized_order_number, user=request.user)
        order_items = OrderItem.objects.filter(order=order)
        return_request_obj = order.return_requests.order_by('-requested_at').first()
        return_history = None
        if return_request_obj:
            return_history = return_request_obj.history.select_related('changed_by')

        forced_view = (request.GET.get('view') or '').strip().lower()
        if forced_view in {'mobile', 'tablet'}:
            template_name = 'order_tracking_mobile.html'
        else:
            ch_mobile = (request.META.get('HTTP_SEC_CH_UA_MOBILE') or '').strip().lower()
            user_agent = (request.META.get('HTTP_USER_AGENT') or '').lower()
            is_tablet = any(
                token in user_agent
                for token in ['ipad', 'tablet', 'kindle', 'playbook', 'silk', 'sm-t', 'nexus 7', 'nexus 10']
            )
            is_mobile = ch_mobile in {'?1', '1', 'true'} or any(
                token in user_agent for token in ['mobile', 'iphone', 'ipod', 'android mobile', 'opera mini', 'iemobile']
            )
            template_name = 'order_tracking_mobile.html' if (is_mobile or is_tablet) else 'order_tracking.html'

        return render(request, template_name, {
            'order': order,
            'order_items': order_items,
            'return_request': return_request_obj,
            'return_history': return_history,
        })
    except Order.DoesNotExist:
        messages.error(request, 'Order not found')
        return redirect('order_list')


@login_required(login_url='login')
def return_request(request, order_id):
    """Create a return request for a delivered order"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    is_eligible, ineligible_reason, deadline, eligible_items = _return_eligibility(order)
    existing_return = order.return_requests.order_by('-requested_at').first()
    refund_options = [
        ('WALLET', 'VibeMall Wallet'),
        ('BANK', 'Direct Bank Transfer'),
        ('RAZORPAY', 'RazorPay'),
    ]

    if request.method == 'POST':
        if not is_eligible:
            messages.error(request, ineligible_reason)
            return redirect('order_details', order_number=order.order_number)

        selected_items = []
        for item in eligible_items:
            if request.POST.get(f'item_{item.id}_selected') == 'on':
                qty_raw = request.POST.get(f'item_{item.id}_qty', '0')
                try:
                    qty = int(qty_raw)
                except ValueError:
                    qty = 0

                if qty < 1 or qty > item.quantity:
                    messages.error(request, f'Invalid quantity for {item.product_name}.')
                    return redirect('return_request', order_id=order.id)

                condition = request.POST.get(f'item_{item.id}_condition', 'OPENED')
                notes = request.POST.get(f'item_{item.id}_notes', '').strip()
                selected_items.append((item, qty, condition, notes))

        if not selected_items:
            messages.error(request, 'Select at least one item to return.')
            return redirect('return_request', order_id=order.id)

        reason = request.POST.get('reason', 'OTHER')
        reason_notes = request.POST.get('reason_notes', '').strip()
        refund_method = request.POST.get('refund_method', '').strip()
        bank_account_name = request.POST.get('bank_account_name', '').strip()
        bank_account_number = request.POST.get('bank_account_number', '').strip()
        bank_ifsc = request.POST.get('bank_ifsc', '').strip()
        bank_name = request.POST.get('bank_name', '').strip()
        upi_id = request.POST.get('upi_id', '').strip()
        allowed_refunds = {value for value, _ in refund_options}
        if refund_method not in allowed_refunds:
            messages.error(request, 'Please select a valid refund method.')
            return redirect('return_request', order_id=order.id)
        if refund_method == 'BANK':
            if not bank_account_name or not bank_account_number or not bank_ifsc or not bank_name:
                messages.error(request, 'Please enter all bank transfer details.')
                return redirect('return_request', order_id=order.id)
            if not re.fullmatch(r"[0-9]{6,34}", bank_account_number):
                messages.error(request, 'Please enter a valid bank account number.')
                return redirect('return_request', order_id=order.id)
            bank_ifsc = bank_ifsc.upper()
            if not re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", bank_ifsc):
                messages.error(request, 'Please enter a valid IFSC code.')
                return redirect('return_request', order_id=order.id)
        if refund_method == 'RAZORPAY':
            if not upi_id:
                messages.error(request, 'Please enter a valid UPI ID.')
                return redirect('return_request', order_id=order.id)
            upi_id = upi_id.lower()
            if not re.fullmatch(r"[a-z0-9._-]{2,256}@[a-z]{2,64}", upi_id):
                messages.error(request, 'Please enter a valid UPI ID.')
                return redirect('return_request', order_id=order.id)
            valid_upi, upi_name, error_msg = _verify_upi_with_razorpay(upi_id, logger=logger)
            if not valid_upi:
                messages.error(request, error_msg or 'UPI ID not found.')
                return redirect('return_request', order_id=order.id)

        return_request_obj = ReturnRequest.objects.create(
            order=order,
            user=request.user,
            reason=reason,
            reason_notes=reason_notes,
            refund_method=refund_method,
            bank_account_name=bank_account_name if refund_method == 'BANK' else '',
            bank_account_number=bank_account_number if refund_method == 'BANK' else '',
            bank_ifsc=bank_ifsc if refund_method == 'BANK' else '',
            bank_name=bank_name if refund_method == 'BANK' else '',
            upi_id=upi_id if refund_method == 'RAZORPAY' else '',
            upi_name=upi_name if refund_method == 'RAZORPAY' else '',
            request_ip=request.META.get('REMOTE_ADDR'),
            request_user_agent=(request.META.get('HTTP_USER_AGENT') or '')[:255]
        )

        for item, qty, condition, notes in selected_items:
            ReturnItem.objects.create(
                return_request=return_request_obj,
                order_item=item,
                product=item.product,
                quantity=qty,
                condition=condition,
                notes=notes
            )

        for attachment in request.FILES.getlist('attachments'):
            ReturnAttachment.objects.create(
                return_request=return_request_obj,
                file=attachment,
                original_name=attachment.name,
                content_type=getattr(attachment, 'content_type', ''),
                size_bytes=getattr(attachment, 'size', 0)
            )

        _log_return_history(return_request_obj, '', 'REQUESTED', request.user, 'Return requested by customer')
        _send_return_notification(
            return_request_obj,
            f'Return Request Received - {order.order_number}',
            f'We received your return request for order {order.order_number}. We will update you soon.'
        )
        _send_admin_return_notification(return_request_obj, request)

        messages.success(request, 'Return request submitted successfully.')
        return redirect('return_status', return_id=return_request_obj.id)

    return render(request, 'return_request.html', {
        'order': order,
        'eligible_items': eligible_items,
        'is_eligible': is_eligible,
        'ineligible_reason': ineligible_reason,
        'return_deadline': deadline,
        'existing_return': existing_return,
        'reason_choices': ReturnRequest.RETURN_REASON_CHOICES,
        'condition_choices': ReturnItem.CONDITION_CHOICES,
        'refund_options': refund_options,
    })


@login_required(login_url='login')
def return_status(request, return_id):
    """Show return request status for a user"""
    return_request_obj = get_object_or_404(ReturnRequest, id=return_id, user=request.user)
    return render(request, 'return_status.html', {
        'return_request': return_request_obj,
        'history': return_request_obj.history.select_related('changed_by')
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_returns(request):
    """Admin return requests list with operations summary."""
    returns_qs = ReturnRequest.objects.select_related('order', 'user').prefetch_related('items')

    status_filter = request.GET.get('status', '').strip()
    reason_filter = request.GET.get('reason', '').strip()
    search_query = request.GET.get('search', '').strip()

    if status_filter:
        returns_qs = returns_qs.filter(status=status_filter)
    if reason_filter:
        returns_qs = returns_qs.filter(reason=reason_filter)
    if search_query:
        returns_qs = returns_qs.filter(
            Q(order__order_number__icontains=search_query) |
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

    returns = list(returns_qs.order_by('-requested_at'))
    ops_profiles = _build_customer_ops_profiles([ret.user_id for ret in returns])
    for ret in returns:
        ret.customer_ops_profile = ops_profiles.get(ret.user_id, {'label': 'Low Risk', 'color': 'success', 'return_count': 0, 'rto_count': 0, 'issue_rate': 0})
        ret.open_days = max((timezone.now() - ret.requested_at).days, 0)
        ret.ops_recommendation = _return_ops_recommendation(ret.status)

    refund_total = ReturnRequest.objects.filter(status='REFUNDED').aggregate(
        total=Sum(Coalesce('refund_amount_net', 'refund_amount'))
    )['total'] or Decimal('0.00')

    pending_returns = ReturnRequest.objects.exclude(status__in=['REFUNDED', 'REPLACED', 'REJECTED', 'CANCELLED']).count()
    qc_queue_count = ReturnRequest.objects.filter(status__in=['RECEIVED', 'QC_PENDING', 'QC_PASSED', 'QC_FAILED']).count()
    refund_pending_count = ReturnRequest.objects.filter(status='REFUND_PENDING').count()
    active_rto_cases = RTOCase.objects.exclude(status='RTO_CLOSED').count()

    risky_user_ids = set(ReturnRequest.objects.values_list('user_id', flat=True))
    risky_user_ids.update(RTOCase.objects.values_list('order__user_id', flat=True))
    all_profiles = _build_customer_ops_profiles(list(risky_user_ids))
    high_risk_customers = sum(1 for profile in all_profiles.values() if profile['level'] == 'high')

    return render(request, 'admin_panel/returns.html', {
        'returns': returns,
        'status_filter': status_filter,
        'reason_filter': reason_filter,
        'search_query': search_query,
        'refund_total': refund_total,
        'status_choices': ReturnRequest.STATUS_CHOICES,
        'reason_choices': ReturnRequest.RETURN_REASON_CHOICES,
        'pending_returns': pending_returns,
        'qc_queue_count': qc_queue_count,
        'refund_pending_count': refund_pending_count,
        'active_rto_cases': active_rto_cases,
        'high_risk_customers': high_risk_customers,
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_return_detail(request, return_id):
    """Admin return request detail and workflow"""
    return_request_obj = get_object_or_404(ReturnRequest, id=return_id)
    customer_profile = UserProfile.objects.filter(user=return_request_obj.user).first()
    full_refund_amount = _calculate_return_refund_amount(return_request_obj)
    half_refund_amount = (full_refund_amount / Decimal('2')).quantize(Decimal('0.01')) if full_refund_amount else Decimal('0.00')
    allowed_next = RETURN_STATUS_FLOW.get(return_request_obj.status, [])

    if request.method == 'POST':
        action = request.POST.get('action')
        notes = request.POST.get('notes', '').strip()

        if action not in allowed_next:
            messages.error(request, 'Invalid status transition.')
            return redirect('admin_return_detail', return_id=return_request_obj.id)

        old_status = return_request_obj.status
        now = timezone.now()
        return_request_obj.status = action

        refund_notes = ''

        if action == 'APPROVED':
            return_request_obj.approved_at = now
        elif action == 'PICKUP_SCHEDULED':
            pickup_at = request.POST.get('pickup_scheduled_at')
            if pickup_at:
                parsed = parse_datetime(pickup_at)
                if not parsed:
                    date_only = parse_date(pickup_at)
                    if date_only:
                        parsed = timezone.make_aware(datetime.combine(date_only, time(hour=10, minute=0)))
                return_request_obj.pickup_scheduled_at = parsed if parsed else now
            else:
                return_request_obj.pickup_scheduled_at = now
        elif action == 'RECEIVED':
            return_request_obj.received_at = now
        elif action in ['QC_PASSED', 'QC_FAILED']:
            return_request_obj.qc_checked_at = now
            if action == 'QC_PASSED':
                for item in return_request_obj.items.select_related('product'):
                    if item.product:
                        item.product.stock = F('stock') + item.quantity
                        item.product.save(update_fields=['stock'])
        elif action == 'QC_PENDING':
            return_request_obj.qc_checked_at = now
        elif action == 'WRONG_RETURN':
            fee = Decimal('20.00') if half_refund_amount > 0 else Decimal('0.00')
            net_amount = max(half_refund_amount - fee, Decimal('0.00'))
            return_request_obj.refund_amount = half_refund_amount
            return_request_obj.refund_fee = fee
            return_request_obj.refund_amount_net = net_amount
        elif action in ['REFUNDED', 'REPLACED']:
            return_request_obj.resolved_at = now
            if action == 'REFUNDED':
                refund_amount_raw = request.POST.get('refund_amount', '').strip()
                refund_amount = None
                if refund_amount_raw:
                    try:
                        refund_amount = Decimal(refund_amount_raw)
                    except Exception:
                        refund_amount = None
                refund_method = request.POST.get('refund_method', '').strip()
                refund_success, refund_notes = _process_refund(return_request_obj, refund_amount, refund_method)
                if not refund_success:
                    return_request_obj.status = 'REFUND_PENDING'
                    return_request_obj.resolved_at = None

        if notes:
            stamped_notes = f"[{now.strftime('%d %b %Y %H:%M')}] {notes}"
            return_request_obj.admin_notes = f"{stamped_notes}\n{return_request_obj.admin_notes}".strip()

        if refund_notes:
            stamped_refund_note = f"[{now.strftime('%d %b %Y %H:%M')}] {refund_notes}"
            return_request_obj.admin_notes = f"{stamped_refund_note}\n{return_request_obj.admin_notes}".strip()

        return_request_obj.save()
        final_action = return_request_obj.status
        history_notes = notes
        if refund_notes:
            history_notes = f"{notes}\n{refund_notes}".strip()

        _log_return_history(return_request_obj, old_status, final_action, request.user, history_notes)
        _send_return_notification(
            return_request_obj,
            f'Return Update - {return_request_obj.order.order_number}',
            f'Your return request is now {final_action.replace("_", " ").title()}.'
        )

        if refund_notes and final_action == 'REFUND_PENDING':
            messages.warning(request, f'Return updated to {final_action}. Refund could not be completed: {refund_notes}')
        else:
            messages.success(request, f'Return updated to {final_action}.')
        return redirect('admin_return_detail', return_id=return_request_obj.id)

    item_conditions = list(return_request_obj.items.values_list('condition', flat=True))
    if item_conditions and all(condition == 'NEW' for condition in item_conditions):
        disposition_recommendation = 'Restock candidate after QC confirmation.'
    elif 'DAMAGED' in item_conditions:
        disposition_recommendation = 'Move to damage hold or clearance review after QC.'
    else:
        disposition_recommendation = 'Manual QC review before restock or refund closure.'

    customer_ops_profile = _build_customer_ops_profiles([return_request_obj.user_id]).get(
        return_request_obj.user_id,
        {'label': 'Low Risk', 'color': 'success', 'return_count': 0, 'rto_count': 0, 'issue_rate': 0},
    )

    return render(request, 'admin_panel/return_detail.html', {
        'return_request': return_request_obj,
        'history': return_request_obj.history.select_related('changed_by'),
        'allowed_next': allowed_next,
        'customer_profile': customer_profile,
        'half_refund_amount': half_refund_amount,
        'customer_ops_profile': customer_ops_profile,
        'disposition_recommendation': disposition_recommendation,
        'ops_recommendation': _return_ops_recommendation(return_request_obj.status),
        'open_days': max((timezone.now() - return_request_obj.requested_at).days, 0),
        'attachments': return_request_obj.attachments.all(),
        'linked_rto_case': getattr(return_request_obj.order, 'rto_case', None),
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_rto_cases(request):
    """Admin RTO management list."""
    rto_qs = RTOCase.objects.select_related('order', 'order__user')

    status_filter = request.GET.get('status', '').strip()
    reason_filter = request.GET.get('reason', '').strip()
    search_query = request.GET.get('search', '').strip()

    if status_filter:
        rto_qs = rto_qs.filter(status=status_filter)
    if reason_filter:
        rto_qs = rto_qs.filter(reason=reason_filter)
    if search_query:
        rto_qs = rto_qs.filter(
            Q(order__order_number__icontains=search_query) |
            Q(order__user__username__icontains=search_query) |
            Q(order__user__email__icontains=search_query) |
            Q(tracking_number__icontains=search_query)
        )

    rto_cases = list(rto_qs.order_by('-created_at'))
    ops_profiles = _build_customer_ops_profiles([case.order.user_id for case in rto_cases])
    for case in rto_cases:
        case.customer_ops_profile = ops_profiles.get(case.order.user_id, {'label': 'Low Risk', 'color': 'success', 'return_count': 0, 'rto_count': 0, 'issue_rate': 0})
        case.open_days = max((timezone.now() - case.created_at).days, 0)
        case.ops_recommendation = _rto_ops_recommendation(case.status)

    all_profiles = _build_customer_ops_profiles(list(set(RTOCase.objects.values_list('order__user_id', flat=True))))
    high_risk_cases = sum(1 for case in rto_cases if case.customer_ops_profile.get('level') == 'high')

    return render(request, 'admin_panel/rto_cases.html', {
        'rto_cases': rto_cases,
        'status_filter': status_filter,
        'reason_filter': reason_filter,
        'search_query': search_query,
        'status_choices': RTOCase.STATUS_CHOICES,
        'reason_choices': RTOCase.REASON_CHOICES,
        'total_rto_cases': RTOCase.objects.count(),
        'active_rto_cases': RTOCase.objects.exclude(status='RTO_CLOSED').count(),
        'rto_received_cases': RTOCase.objects.filter(status='RTO_RECEIVED').count(),
        'cod_rto_cases': RTOCase.objects.filter(order__payment_method='COD').count(),
        'high_risk_customers': sum(1 for profile in all_profiles.values() if profile['level'] == 'high'),
        'high_risk_cases': high_risk_cases,
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_rto_detail(request, rto_id):
    """Admin RTO case detail and workflow."""
    rto_case = get_object_or_404(RTOCase.objects.select_related('order', 'order__user'), id=rto_id)
    customer_profile = UserProfile.objects.filter(user=rto_case.order.user).first()
    allowed_next = RTO_STATUS_FLOW.get(rto_case.status, [])

    if request.method == 'POST':
        next_status = (request.POST.get('status') or '').strip()
        notes = (request.POST.get('notes') or '').strip()
        now = timezone.now()
        old_status = rto_case.status
        final_status = old_status

        if next_status and next_status != old_status:
            if next_status not in allowed_next:
                messages.error(request, 'Invalid RTO status transition.')
                return redirect('admin_rto_detail', rto_id=rto_case.id)
            rto_case.status = next_status
            final_status = next_status
            if next_status == 'RTO_INITIATED' and not rto_case.initiated_at:
                rto_case.initiated_at = now
            if next_status == 'RTO_RECEIVED':
                rto_case.received_at = now
            if next_status == 'RTO_CLOSED':
                rto_case.closed_at = now

        reason = (request.POST.get('reason') or '').strip()
        if reason in dict(RTOCase.REASON_CHOICES):
            rto_case.reason = reason

        resolution_action = (request.POST.get('resolution_action') or '').strip()
        if resolution_action in dict(RTOCase.RESOLUTION_CHOICES):
            rto_case.resolution_action = resolution_action

        rto_case.reason_notes = (request.POST.get('reason_notes') or '').strip()
        rto_case.courier_name = (request.POST.get('courier_name') or '').strip()
        rto_case.tracking_number = (request.POST.get('tracking_number') or '').strip()

        last_attempted_at = (request.POST.get('last_attempted_at') or '').strip()
        if last_attempted_at:
            parsed = parse_datetime(last_attempted_at)
            if not parsed:
                date_only = parse_date(last_attempted_at)
                if date_only:
                    parsed = timezone.make_aware(datetime.combine(date_only, time(hour=10, minute=0)))
            if parsed:
                rto_case.last_attempted_at = parsed

        if notes:
            stamped_notes = f"[{now.strftime('%d %b %Y %H:%M')}] {notes}"
            rto_case.admin_notes = f"{stamped_notes}\n{rto_case.admin_notes}".strip()

        rto_case.save()

        if final_status != old_status:
            _log_rto_history(rto_case, old_status, final_status, request.user, notes)
            messages.success(request, f'RTO case updated to {rto_case.get_status_display()}.')
        else:
            messages.success(request, 'RTO case details updated.')
        return redirect('admin_rto_detail', rto_id=rto_case.id)

    customer_ops_profile = _build_customer_ops_profiles([rto_case.order.user_id]).get(
        rto_case.order.user_id,
        {'label': 'Low Risk', 'color': 'success', 'return_count': 0, 'rto_count': 0, 'issue_rate': 0},
    )

    return render(request, 'admin_panel/rto_detail.html', {
        'rto_case': rto_case,
        'history': rto_case.history.select_related('changed_by'),
        'allowed_next': allowed_next,
        'customer_profile': customer_profile,
        'customer_ops_profile': customer_ops_profile,
        'ops_recommendation': _rto_ops_recommendation(rto_case.status),
        'open_days': max((timezone.now() - rto_case.created_at).days, 0),
        'reason_choices': RTOCase.REASON_CHOICES,
        'resolution_choices': RTOCase.RESOLUTION_CHOICES,
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_return_analytics(request):
    """Comprehensive return analytics dashboard"""
    from django.db.models.functions import TruncMonth, TruncWeek, TruncDate
    from collections import Counter
    
    # Date filters
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    all_returns = ReturnRequest.objects.select_related('order', 'user').prefetch_related('items', 'items__product')
    
    if date_from:
        all_returns = all_returns.filter(requested_at__date__gte=date_from)
    if date_to:
        all_returns = all_returns.filter(requested_at__date__lte=date_to)
    
    # ===== 1. OVERVIEW METRICS =====
    total_returns_count = all_returns.count()
    total_delivered = Order.objects.filter(order_status='DELIVERED').count()
    return_rate = (total_returns_count / total_delivered * 100) if total_delivered > 0 else 0
    
    refunded_returns = all_returns.filter(status='REFUNDED')
    total_refund_value = refunded_returns.aggregate(
        total=Sum(Coalesce('refund_amount_net', 'refund_amount'))
    )['total'] or Decimal('0.00')
    
    avg_refund = refunded_returns.aggregate(
        avg=Avg(Coalesce('refund_amount_net', 'refund_amount'))
    )['avg'] or Decimal('0.00')
    
    pending_returns = all_returns.exclude(status__in=['REFUNDED', 'REPLACED', 'REJECTED']).count()
    resolved_returns = all_returns.filter(status__in=['REFUNDED', 'REPLACED']).count()
    
    # Average resolution time
    resolved_with_time = all_returns.filter(
        status__in=['REFUNDED', 'REPLACED'],
        resolved_at__isnull=False
    ).annotate(
        resolution_days=ExpressionWrapper(
            F('resolved_at') - F('requested_at'),
            output_field=DecimalField()
        )
    )
    
    avg_resolution_days = 0
    if resolved_with_time.exists():
        total_seconds = sum([
            (r.resolved_at - r.requested_at).total_seconds() 
            for r in resolved_with_time 
            if r.resolved_at and r.requested_at
        ])
        avg_resolution_days = total_seconds / len(resolved_with_time) / 86400 if len(resolved_with_time) > 0 else 0
    
    # ===== 2. RETURN REASON ANALYSIS =====
    reason_breakdown = all_returns.values('reason').annotate(
        count=Count('id')
    ).order_by('-count')
    
    total_for_percentage = total_returns_count if total_returns_count > 0 else 1
    reason_data = []
    for item in reason_breakdown:
        reason_label = dict(ReturnRequest.RETURN_REASON_CHOICES).get(item['reason'], item['reason'])
        reason_data.append({
            'reason': reason_label,
            'count': item['count'],
            'percentage': round(item['count'] / total_for_percentage * 100, 1)
        })
    
    defective_count = all_returns.filter(reason='DEFECTIVE').count()
    changed_mind_count = all_returns.filter(reason='CHANGED_MIND').count()
    
    # ===== 3. PRODUCT ANALYSIS =====
    returned_items = ReturnItem.objects.filter(return_request__in=all_returns)
    
    product_return_counts = returned_items.values(
        'product__name', 'product__id'
    ).annotate(
        return_count=Count('id'),
        total_qty=Sum('quantity')
    ).order_by('-return_count')[:10]
    
    # Category-wise returns
    category_returns = returned_items.filter(product__category__isnull=False).values(
        'product__category'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Brand-wise returns
    brand_returns = returned_items.filter(product__brand__isnull=False).exclude(
        product__brand=''
    ).values(
        'product__brand'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    brand_returns = list(brand_returns)
    brand_total = sum(item['count'] for item in brand_returns) or 1
    for item in brand_returns:
        item['percentage'] = round(item['count'] / brand_total * 100, 1)
    
    # ===== 4. CUSTOMER PATTERNS =====
    user_return_counts = all_returns.values(
        'user__username', 'user__email', 'user__id'
    ).annotate(
        return_count=Count('id')
    ).order_by('-return_count')[:10]
    
    # Repeat returners (>2 returns)
    repeat_returners = all_returns.values('user').annotate(
        count=Count('id')
    ).filter(count__gt=2).count()
    
    # Potential abuse (>5 returns)
    potential_abuse = all_returns.values('user').annotate(
        count=Count('id')
    ).filter(count__gt=5).count()
    
    # ===== 5. TIME-BASED TRENDS =====
    monthly_returns = all_returns.annotate(
        month=TruncMonth('requested_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')[:12]
    
    weekly_returns = all_returns.filter(
        requested_at__gte=timezone.now() - timedelta(days=90)
    ).annotate(
        week=TruncWeek('requested_at')
    ).values('week').annotate(
        count=Count('id')
    ).order_by('week')
    
    daily_returns = all_returns.filter(
        requested_at__gte=timezone.now() - timedelta(days=30)
    ).annotate(
        day=TruncDate('requested_at')
    ).values('day').annotate(
        count=Count('id')
    ).order_by('day')
    
    # ===== 6. QC & RESOLUTION STATS =====
    qc_passed = all_returns.filter(status='QC_PASSED').count()
    qc_failed = all_returns.filter(status='QC_FAILED').count()
    wrong_returns = all_returns.filter(status='WRONG_RETURN').count()
    
    # Status-wise processing time
    status_times = {}
    for status_code, status_label in ReturnRequest.STATUS_CHOICES:
        status_returns = all_returns.filter(status=status_code, resolved_at__isnull=False)
        if status_returns.exists():
            total_secs = sum([
                (r.resolved_at - r.requested_at).total_seconds()
                for r in status_returns
                if r.resolved_at and r.requested_at
            ])
            avg_days = total_secs / status_returns.count() / 86400 if status_returns.count() > 0 else 0
            status_times[status_label] = round(avg_days, 1)
    
    # Refund method breakdown
    refund_methods = refunded_returns.exclude(refund_method='').values('refund_method').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # ===== 7. FINANCIAL IMPACT =====
    total_refund_fees = refunded_returns.aggregate(
        total=Sum('refund_fee')
    )['total'] or Decimal('0.00')
    
    net_refund_impact = total_refund_value - total_refund_fees
    
    # Payment method wise refunds
    payment_refunds = refunded_returns.values('order__payment_method').annotate(
        count=Count('id'),
        total_amount=Sum(Coalesce('refund_amount_net', 'refund_amount'))
    ).order_by('-total_amount')
    
    context = {
        # Overview
        'total_returns_count': total_returns_count,
        'return_rate': round(return_rate, 2),
        'total_refund_value': total_refund_value,
        'avg_refund': avg_refund,
        'pending_returns': pending_returns,
        'resolved_returns': resolved_returns,
        'avg_resolution_days': round(avg_resolution_days, 1),
        
        # Reasons
        'reason_data': reason_data,
        'defective_count': defective_count,
        'changed_mind_count': changed_mind_count,
        
        # Products
        'top_returned_products': product_return_counts,
        'category_returns': category_returns,
        'brand_returns': brand_returns,
        
        # Customers
        'top_returners': user_return_counts,
        'repeat_returners': repeat_returners,
        'potential_abuse': potential_abuse,
        
        # Time trends
        'monthly_returns': list(monthly_returns),
        'weekly_returns': list(weekly_returns),
        'daily_returns': list(daily_returns),
        
        # QC Stats
        'qc_passed': qc_passed,
        'qc_failed': qc_failed,
        'wrong_returns': wrong_returns,
        'status_times': status_times,
        'refund_methods': refund_methods,
        
        # Financial
        'total_refund_fees': total_refund_fees,
        'net_refund_impact': net_refund_impact,
        'payment_refunds': payment_refunds,
        
        # Filters
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'admin_panel/return_analytics.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_marketing_studio(request):
    return render(request, 'admin_panel/marketing_studio.html')


@login_required(login_url='login')
@require_POST
def customer_cancel_order(request, order_id):
    """Allow customer to request cancellation (before shipping)"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    cancel_request = getattr(order, 'cancellation_request', None)

    if cancel_request:
        if cancel_request.status == 'REQUESTED':
            messages.info(request, 'Cancellation request is already pending.')
        elif cancel_request.status == 'APPROVED':
            messages.info(request, 'This order has already been cancelled.')
        else:
            messages.info(request, 'Cancellation request was rejected.')
        return redirect(request.META.get('HTTP_REFERER', 'order_list'))

    is_allowed, reason, _deadline = _cancel_eligibility(order)
    if not is_allowed:
        messages.error(request, reason)
        return redirect(request.META.get('HTTP_REFERER', 'order_list'))

    reason_text = request.POST.get('reason', '').strip()
    notes = request.POST.get('notes', '').strip()
    refund_method = request.POST.get('refund_method', '').strip().upper()
    allowed_refunds = {'WALLET', 'BANK', 'RAZORPAY'}

    bank_account_name = request.POST.get('bank_account_name', '').strip()
    bank_account_number = request.POST.get('bank_account_number', '').strip()
    bank_ifsc = request.POST.get('bank_ifsc', '').strip().upper()
    bank_name = request.POST.get('bank_name', '').strip()
    upi_id = request.POST.get('upi_id', '').strip().lower()
    upi_name = ''

    if order.payment_method == 'COD' or order.payment_status != 'PAID':
        refund_method = ''
    else:
        if refund_method not in allowed_refunds:
            messages.error(request, 'Please select a valid refund method.')
            return redirect(request.META.get('HTTP_REFERER', 'order_list'))
        if refund_method == 'BANK':
            if not bank_account_name or not bank_account_number or not bank_ifsc or not bank_name:
                messages.error(request, 'Please enter all bank transfer details.')
                return redirect(request.META.get('HTTP_REFERER', 'order_list'))
            if not re.fullmatch(r"[0-9]{6,34}", bank_account_number):
                messages.error(request, 'Please enter a valid bank account number.')
                return redirect(request.META.get('HTTP_REFERER', 'order_list'))
            if not re.fullmatch(r"[A-Z]{4}0[A-Z0-9]{6}", bank_ifsc):
                messages.error(request, 'Please enter a valid IFSC code.')
                return redirect(request.META.get('HTTP_REFERER', 'order_list'))
        if refund_method == 'RAZORPAY':
            if not upi_id:
                messages.error(request, 'Please enter a valid UPI ID.')
                return redirect(request.META.get('HTTP_REFERER', 'order_list'))
            if not re.fullmatch(r"[a-z0-9._-]{2,256}@[a-z]{2,64}", upi_id):
                messages.error(request, 'Please enter a valid UPI ID.')
                return redirect(request.META.get('HTTP_REFERER', 'order_list'))
            valid_upi, resolved_name, error_msg = _verify_upi_with_razorpay(upi_id, logger=logger)
            if not valid_upi:
                messages.error(request, error_msg or 'UPI ID not found.')
                return redirect(request.META.get('HTTP_REFERER', 'order_list'))
            upi_name = resolved_name

    OrderCancellationRequest.objects.create(
        order=order,
        user=request.user,
        reason=reason_text,
        notes=notes,
        refund_method=refund_method,
        bank_account_name=bank_account_name if refund_method == 'BANK' else '',
        bank_account_number=bank_account_number if refund_method == 'BANK' else '',
        bank_ifsc=bank_ifsc if refund_method == 'BANK' else '',
        bank_name=bank_name if refund_method == 'BANK' else '',
        upi_id=upi_id if refund_method == 'RAZORPAY' else '',
        upi_name=upi_name if refund_method == 'RAZORPAY' else '',
    )

    OrderStatusHistory.objects.create(
        order=order,
        old_status=order.order_status,
        new_status='CANCEL_REQUESTED',
        changed_by=request.user,
        notes=notes or 'Cancellation requested by customer.'
    )

    messages.success(request, 'Cancellation request submitted. Admin will review it shortly.')
    return redirect(request.META.get('HTTP_REFERER', 'order_list'))


@login_required(login_url='login')
@never_cache
def download_invoice(request, order_number):
    """Generate and download PDF invoice for an order"""
    import traceback
    import os
    from io import BytesIO

    def _generate_reportlab_fallback_pdf(order_obj, invoice_context):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib import colors
        from reportlab.lib.utils import ImageReader
        from reportlab.pdfgen import canvas
        import requests

        def _draw_line(pdf, text, y_pos, x_pos=18 * mm, font_name='Helvetica', font_size=10):
            pdf.setFont(font_name, font_size)
            pdf.drawString(x_pos, y_pos, text)
            return y_pos - 5.5 * mm

        def _draw_wrapped_lines(pdf, text, x_pos, y_pos, max_width, line_height=4.6 * mm, font_name='Helvetica', font_size=9):
            words = str(text or '').split()
            if not words:
                return y_pos

            pdf.setFont(font_name, font_size)
            current_line = ''
            for word in words:
                candidate = f"{current_line} {word}".strip()
                if pdf.stringWidth(candidate, font_name, font_size) <= max_width:
                    current_line = candidate
                else:
                    pdf.drawString(x_pos, y_pos, current_line)
                    y_pos -= line_height
                    current_line = word

            if current_line:
                pdf.drawString(x_pos, y_pos, current_line)
                y_pos -= line_height

            return y_pos

        def _get_image_reader(image_url):
            if not image_url:
                return None

            try:
                if image_url.startswith(('http://', 'https://')):
                    response = requests.get(image_url, timeout=6)
                    response.raise_for_status()
                    return ImageReader(BytesIO(response.content))

                relative_path = ''
                if image_url.startswith('/media/'):
                    relative_path = image_url.replace('/media/', '', 1)
                elif image_url.startswith('media/'):
                    relative_path = image_url.replace('media/', '', 1)

                if relative_path:
                    file_path = os.path.join(str(settings.MEDIA_ROOT), relative_path)
                    if os.path.exists(file_path):
                        return ImageReader(file_path)
            except Exception:
                return None

            return None

        pdf_buffer = BytesIO()
        pdf = canvas.Canvas(pdf_buffer, pagesize=A4)
        page_width, page_height = A4
        cursor_y = page_height - 18 * mm

        company_name = 'VibeMall'
        payment_status = 'Paid (COD)' if order_obj.payment_method == 'COD' and order_obj.order_status == 'DELIVERED' else order_obj.get_payment_status_display()
        shipping_lines = invoice_context.get('shipping_lines') or [line.strip() for line in (order_obj.shipping_address or '').replace('\r', '\n').split('\n') if line.strip()]

        pdf.setFont('Helvetica-Bold', 23)
        pdf.drawCentredString(page_width / 2, cursor_y, company_name.upper())
        cursor_y -= 8 * mm
        pdf.setFont('Helvetica-Bold', 9)
        pdf.setFillColor(colors.HexColor('#666666'))
        pdf.drawCentredString(page_width / 2, cursor_y, 'INVOICE')
        pdf.setFillColor(colors.black)
        cursor_y -= 11 * mm

        pdf.setStrokeColor(colors.HexColor('#D8D4CC'))
        pdf.setLineWidth(0.7)
        pdf.line(18 * mm, cursor_y, page_width - 18 * mm, cursor_y)
        cursor_y -= 9 * mm

        pdf.setFont('Helvetica', 8)
        pdf.setFillColor(colors.HexColor('#777777'))
        pdf.drawString(18 * mm, cursor_y, 'ORDER NUMBER')
        pdf.drawCentredString(page_width / 2, cursor_y, 'DATE')
        pdf.drawRightString(page_width - 18 * mm, cursor_y, 'CUSTOMER ID')
        pdf.setFillColor(colors.black)
        cursor_y -= 5 * mm

        pdf.setFont('Helvetica-Bold', 10)
        pdf.drawString(18 * mm, cursor_y, f"#{order_obj.order_number}")
        pdf.drawCentredString(page_width / 2, cursor_y, order_obj.created_at.strftime('%B %d, %Y'))
        pdf.drawRightString(page_width - 18 * mm, cursor_y, f"CU-{order_obj.user_id:05d}")
        cursor_y -= 9 * mm

        pdf.line(18 * mm, cursor_y, page_width - 18 * mm, cursor_y)
        cursor_y -= 9 * mm

        left_col_x = 18 * mm
        right_col_x = page_width / 2 + 10 * mm
        box_top = cursor_y

        pdf.setFont('Helvetica-Bold', 8)
        pdf.setFillColor(colors.HexColor('#6f5c37'))
        pdf.drawString(left_col_x, cursor_y, 'BILLED TO')
        pdf.drawString(right_col_x, cursor_y, 'FROM')
        pdf.setFillColor(colors.black)
        cursor_y -= 5.5 * mm

        pdf.setFont('Helvetica-Bold', 11)
        pdf.drawString(left_col_x, cursor_y, invoice_context.get('customer_name') or (order_obj.user.get_full_name() or order_obj.user.username))
        pdf.drawString(right_col_x, cursor_y, company_name)
        cursor_y -= 5 * mm

        left_y = cursor_y
        right_y = cursor_y

        pdf.setFont('Helvetica', 9)
        for line in shipping_lines[:5]:
            pdf.drawString(left_col_x, left_y, str(line))
            left_y -= 4.7 * mm

        company_lines = invoice_context.get('company_address_lines') or ['katargam 395004 surat ,Gujarat']
        for line in company_lines[:4]:
            pdf.drawString(right_col_x, right_y, str(line))
            right_y -= 4.7 * mm

        company_email = invoice_context.get('company_email') or 'info.vibemall@gmail.com'
        if company_email:
            pdf.drawString(right_col_x, right_y, company_email)

        cursor_y = min(left_y, right_y) - 5 * mm

        pdf.setStrokeColor(colors.HexColor('#D8D4CC'))
        pdf.line(18 * mm, cursor_y, page_width - 18 * mm, cursor_y)
        cursor_y -= 8 * mm

        pdf.setFont('Helvetica-Bold', 8)
        pdf.setFillColor(colors.HexColor('#6b6861'))
        pdf.drawString(18 * mm, cursor_y, 'Item')
        pdf.drawString(128 * mm, cursor_y, 'Qty')
        pdf.drawString(145 * mm, cursor_y, 'Unit')
        pdf.drawRightString(page_width - 18 * mm, cursor_y, 'Total')
        pdf.setFillColor(colors.black)
        cursor_y -= 5 * mm
        pdf.line(18 * mm, cursor_y, page_width - 18 * mm, cursor_y)
        cursor_y -= 6 * mm

        fallback_items = invoice_context.get('order_items') or []
        for item in fallback_items:
            row_height = 25 * mm
            if cursor_y < 40 * mm:
                pdf.showPage()
                cursor_y = page_height - 20 * mm
                pdf.setFont('Helvetica-Bold', 10)
                pdf.drawString(18 * mm, cursor_y, 'Item')
                pdf.drawString(128 * mm, cursor_y, 'Qty')
                pdf.drawString(145 * mm, cursor_y, 'Unit')
                pdf.drawRightString(page_width - 18 * mm, cursor_y, 'Total')
                cursor_y -= 6 * mm

            image_x = 18 * mm
            image_y = cursor_y - 19 * mm
            image_w = 15 * mm
            image_h = 20 * mm

            pdf.setFillColor(colors.HexColor('#EEEAE2'))
            pdf.rect(image_x, image_y, image_w, image_h, stroke=0, fill=1)
            pdf.setFillColor(colors.black)

            image_reader = _get_image_reader(item.get('image_url'))
            if image_reader:
                try:
                    pdf.drawImage(image_reader, image_x, image_y, width=image_w, height=image_h, preserveAspectRatio=True, anchor='c', mask='auto')
                except Exception:
                    pass

            text_x = image_x + image_w + 4 * mm
            text_max_width = 84 * mm

            item_name = item.get('name') or 'Item'
            pdf.setFont('Helvetica-Bold', 10)
            name_y = cursor_y - 1 * mm
            name_y = _draw_wrapped_lines(pdf, item_name, text_x, name_y, text_max_width, line_height=4.2 * mm, font_name='Helvetica-Bold', font_size=9.8)

            variant_text = item.get('variant_text') or ''
            if variant_text:
                pdf.setFont('Helvetica', 8)
                pdf.setFillColor(colors.HexColor('#7a766f'))
                _draw_wrapped_lines(pdf, variant_text.upper(), text_x, name_y + 0.8 * mm, text_max_width, line_height=3.8 * mm, font_name='Helvetica', font_size=7.2)
                pdf.setFillColor(colors.black)

            pdf.setFont('Helvetica', 9)
            pdf.drawString(130 * mm, cursor_y - 2 * mm, str(item.get('quantity') or 0))
            pdf.drawString(145 * mm, cursor_y - 2 * mm, f"Rs {float(item.get('unit_price') or 0):.2f}")
            pdf.drawRightString(page_width - 18 * mm, cursor_y - 2 * mm, f"Rs {float(item.get('line_total') or 0):.2f}")

            cursor_y -= row_height
            pdf.setStrokeColor(colors.HexColor('#EEEAE2'))
            pdf.line(18 * mm, cursor_y + 4 * mm, page_width - 18 * mm, cursor_y + 4 * mm)

        cursor_y -= 2 * mm
        pdf.line(18 * mm, cursor_y, page_width - 18 * mm, cursor_y)
        cursor_y -= 7 * mm

        summary_lines = [
            ('Subtotal', float(order_obj.subtotal or 0)),
            ('Shipping', float(order_obj.shipping_cost or 0)),
        ]
        if order_obj.tax:
            summary_lines.append(('Tax', float(order_obj.tax or 0)))
        if order_obj.coupon_discount:
            summary_lines.append(('Discount', -float(order_obj.coupon_discount or 0)))
        summary_lines.append(('Grand Total', float(order_obj.total_amount or 0)))

        for label, amount in summary_lines:
            is_total = label == 'Grand Total'
            pdf.setFont('Helvetica-Bold' if is_total else 'Helvetica', 11 if is_total else 10)
            pdf.drawString(128 * mm, cursor_y, label)
            display = f"Rs {abs(amount):.2f}"
            if amount < 0:
                display = f"- Rs {abs(amount):.2f}"
            pdf.drawRightString(page_width - 18 * mm, cursor_y, display)
            cursor_y -= 6 * mm

        cursor_y -= 5 * mm
        pdf.setStrokeColor(colors.HexColor('#D8D4CC'))
        pdf.line(18 * mm, cursor_y, page_width - 18 * mm, cursor_y)
        cursor_y -= 6 * mm

        pdf.setFillColor(colors.HexColor('#6f5c37'))
        pdf.setFont('Helvetica-Oblique', 9)
        pdf.drawString(18 * mm, cursor_y, 'Thank you for choosing VibeMall.')
        pdf.setFillColor(colors.black)

        pdf.showPage()
        pdf.save()

        pdf_buffer.seek(0)
        return pdf_buffer.read()
    
    try:
        # Get order
        order_filter = {'order_number': order_number}
        if not request.user.is_staff:
            order_filter['user'] = request.user

        try:
            order = Order.objects.get(**order_filter)
        except Order.DoesNotExist:
            if request.user.is_staff:
                return redirect('admin_invoice_inventory')
            return HttpResponse('Order not found.', status=404)

        try:
            from weasyprint import HTML

            invoice_context = build_invoice_context(order)
            invoice_html = render_to_string('invoice_pdf.html', invoice_context)

            pdf_file = BytesIO()
            HTML(string=invoice_html, base_url=invoice_context['site_url']).write_pdf(pdf_file)
            pdf_file.seek(0)
            pdf_data = pdf_file.read()
            pdf_file.close()
        except (ImportError, OSError, TypeError) as weasy_error:
            logger.warning("WeasyPrint unavailable for order %s, using ReportLab fallback: %s", order_number, weasy_error)
            try:
                invoice_context = build_invoice_context(order)
                pdf_data = _generate_reportlab_fallback_pdf(order, invoice_context)
            except Exception as fallback_error:
                logger.exception("ReportLab fallback also failed for order %s: %s", order_number, fallback_error)
                return HttpResponse(
                    'Invoice PDF service is temporarily unavailable because PDF rendering dependencies are missing on the server. Please contact support.',
                    status=503,
                )
        
        # Return PDF with proper headers
        response = HttpResponse(pdf_data, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="Invoice_{order.order_number}.pdf"'
        response['Content-Length'] = len(pdf_data)
        response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response['Pragma'] = 'no-cache'
        response['Expires'] = '0'
        return response
        
    except Exception as e:
        logger.exception("Error generating invoice for %s: %s", order_number, str(e))
        if getattr(settings, 'DEBUG', False):
            error_details = traceback.format_exc()
            return HttpResponse(error_details, content_type='text/plain', status=500)
        # Return error response without trying to use messages
        return HttpResponse('Error generating invoice. Please contact support.', status=500)


# Import Reel views
from .views_reel import admin_generate_reel


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reels(request):
    """Admin reel management list"""
    from Hub.models import Reel
    
    reels = Reel.objects.all().order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '').strip()
    if search_query:
        reels = reels.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter == 'published':
        reels = reels.filter(is_published=True)
    elif status_filter == 'draft':
        reels = reels.filter(is_published=False)
    elif status_filter == 'processing':
        reels = reels.filter(is_processing=True)
    
    context = {
        'reels': reels,
        'search_query': search_query,
        'status_filter': status_filter,
    }
    
    return render(request, 'admin_panel/reels.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def reel_studio(request):
    """Professional Reel Studio - Canva-style editor"""
    return render(request, 'admin_panel/reel_studio.html')


@login_required(login_url='login')
@staff_member_required(login_url='login')
def reel_studio_export(request):
    """Export reel from studio"""
    import json
    from django.http import JsonResponse
    from Hub.models import Reel, ReelImage
    from django.core.files.base import ContentFile
    from django.views.decorators.csrf import csrf_exempt
    import base64
    
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=400)
    
    try:
        data = json.loads(request.body)
        
        # Create reel
        reel = Reel.objects.create(
            title=data.get('name', 'Studio Reel'),
            description='Created in Reel Studio',
            duration_per_image=3,
            transition_type='zoom',
            add_end_screen=data.get('settings', {}).get('endScreen', True),
            watermark_opacity=0.7,
            watermark_position='top-right',
            created_by=request.user
        )
        
        # Save layers as images
        image_count = 0
        for idx, layer in enumerate(data.get('layers', [])):
            if layer.get('type') == 'image' and layer.get('url'):
                try:
                    # Check if it's base64
                    if 'base64' in layer['url']:
                        image_data = layer['url'].split(',')[1]
                        image_file = ContentFile(
                            base64.b64decode(image_data), 
                            name=f'studio_layer_{idx}.jpg'
                        )
                        
                        ReelImage.objects.create(
                            reel=reel,
                            image=image_file,
                            order=idx,
                            text_overlay=layer.get('content', ''),
                            text_position='center',
                            text_color='white',
                            text_size=70
                        )
                        image_count += 1
                except Exception:
                    logger.exception("Failed saving reel layer index=%s for reel export", idx)
        
        if image_count == 0:
            reel.delete()
            return JsonResponse({
                'error': 'No images to export. Please add images first.'
            }, status=400)
        
        return JsonResponse({
            'success': True,
            'reel_id': reel.id,
            'message': f'Reel created with {image_count} images! Go to All Reels and click Generate.'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.exception("Reel studio export failed")
        return JsonResponse({'error': str(e)}, status=500)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_reel(request):
    """Add new reel"""
    from Hub.models import Reel, ReelImage
    
    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        duration_per_image = request.POST.get('duration_per_image', 3)
        transition_type = request.POST.get('transition_type', 'zoom')
        background_music = request.FILES.get('background_music')
        
        # Branding fields
        watermark_logo = request.FILES.get('watermark_logo')
        watermark_position = request.POST.get('watermark_position', 'top-right')
        watermark_opacity = request.POST.get('watermark_opacity', 0.7)
        add_end_screen = request.POST.get('add_end_screen') == 'on'
        end_screen_duration = request.POST.get('end_screen_duration', 3)
        product_id = (request.POST.get('product_id') or '').strip()
        order = int(request.POST.get('order', 0))
        linked_product = None
        
        if not title:
            messages.error(request, 'Title is required')
            return redirect('admin_add_reel')
        
        try:
            duration_per_image = int(duration_per_image)
            if duration_per_image < 1 or duration_per_image > 10:
                duration_per_image = 3
        except:
            duration_per_image = 3
        
        try:
            watermark_opacity = float(watermark_opacity)
        except:
            watermark_opacity = 0.7
        
        try:
            end_screen_duration = int(end_screen_duration)
        except:
            end_screen_duration = 3

        if product_id:
            linked_product = Product.objects.filter(id=product_id, is_active=True).first()
            if not linked_product:
                messages.error(request, 'Selected product is invalid or inactive.')
                return redirect('admin_add_reel')
        
        reel = Reel.objects.create(
            title=title,
            description=description,
            product=linked_product,
            duration_per_image=duration_per_image,
            transition_type=transition_type,
            order=order,
            watermark_position=watermark_position,
            watermark_opacity=watermark_opacity,
            add_end_screen=add_end_screen,
            end_screen_duration=end_screen_duration,
            created_by=request.user
        )
        
        if background_music:
            reel.background_music = background_music
            reel.save()
        
        if watermark_logo:
            reel.watermark_logo = watermark_logo
            reel.save()
        
        # Handle multiple image uploads
        images = request.FILES.getlist('images')
        for idx, image in enumerate(images):
            text_overlay = request.POST.get(f'text_overlay_{idx}', '')
            text_position = request.POST.get(f'text_position_{idx}', 'center')
            text_color = request.POST.get(f'text_color_{idx}', 'white')
            text_size = request.POST.get(f'text_size_{idx}', 70)
            
            try:
                text_size = int(text_size)
            except:
                text_size = 70
            
            ReelImage.objects.create(
                reel=reel,
                image=image,
                order=idx,
                text_overlay=text_overlay,
                text_position=text_position,
                text_color=text_color,
                text_size=text_size
            )
        
        messages.success(request, f'✨ Professional reel "{title}" created successfully!')
        return redirect('admin_edit_reel', reel_id=reel.id)
    
    products = Product.objects.filter(is_active=True).only('id', 'name').order_by('name')
    return render(request, 'admin_panel/add_reel.html', {'products': products})


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_reel(request, reel_id):
    """Edit existing reel"""
    from Hub.models import Reel, ReelImage
    
    reel = get_object_or_404(Reel, id=reel_id)
    
    if request.method == 'POST':
        product_id = (request.POST.get('product_id') or '').strip()
        reel.title = request.POST.get('title', '').strip()
        reel.description = request.POST.get('description', '').strip()
        if product_id:
            linked_product = Product.objects.filter(id=product_id, is_active=True).first()
            if not linked_product:
                messages.error(request, 'Selected product is invalid or inactive.')
                return redirect('admin_edit_reel', reel_id=reel.id)
            reel.product = linked_product
        else:
            reel.product = None
        
        try:
            reel.duration_per_image = int(request.POST.get('duration_per_image', 3))
        except:
            reel.duration_per_image = 3
        
        reel.transition_type = request.POST.get('transition_type', 'zoom')
        reel.is_published = request.POST.get('is_published') == 'on'
        
        # Branding fields
        reel.watermark_position = request.POST.get('watermark_position', 'top-right')
        try:
            reel.watermark_opacity = float(request.POST.get('watermark_opacity', 0.7))
        except:
            reel.watermark_opacity = 0.7
        
        reel.add_end_screen = request.POST.get('add_end_screen') == 'on'
        try:
            reel.end_screen_duration = int(request.POST.get('end_screen_duration', 3))
        except:
            reel.end_screen_duration = 3

        try:
            reel.view_count = max(int(request.POST.get('view_count', reel.view_count or 0)), 0)
        except (TypeError, ValueError):
            reel.view_count = reel.view_count or 0

        try:
            reel.like_count = max(int(request.POST.get('like_count', reel.like_count or 0)), 0)
        except (TypeError, ValueError):
            reel.like_count = reel.like_count or 0

        try:
            reel.order = max(int(request.POST.get('order', reel.order or 0)), 0)
        except (TypeError, ValueError):
            reel.order = reel.order or 0
        
        if 'background_music' in request.FILES:
            reel.background_music = request.FILES['background_music']
        
        if 'watermark_logo' in request.FILES:
            reel.watermark_logo = request.FILES['watermark_logo']
        
        reel.save()
        
        # Handle new images
        new_images = request.FILES.getlist('new_images')
        if new_images:
            max_order = reel.images.count()
            for idx, image in enumerate(new_images):
                text_overlay = request.POST.get(f'new_text_overlay_{idx}', '')
                text_position = request.POST.get(f'new_text_position_{idx}', 'center')
                text_color = request.POST.get(f'new_text_color_{idx}', 'white')
                text_size = request.POST.get(f'new_text_size_{idx}', 70)
                
                try:
                    text_size = int(text_size)
                except:
                    text_size = 70
                
                ReelImage.objects.create(
                    reel=reel,
                    image=image,
                    order=max_order + idx,
                    text_overlay=text_overlay,
                    text_position=text_position,
                    text_color=text_color,
                    text_size=text_size
                )
        
        messages.success(request, f'✅ Reel "{reel.title}" updated successfully!')
        return redirect('admin_edit_reel', reel_id=reel.id)
    
    context = {
        'reel': reel,
        'images': reel.images.all().order_by('order'),
        'products': Product.objects.filter(is_active=True).only('id', 'name').order_by('name'),
    }
    
    return render(request, 'admin_panel/edit_reel.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_reel(request, reel_id):
    """Delete reel"""
    from Hub.models import Reel
    
    reel = get_object_or_404(Reel, id=reel_id)
    
    # If published, unpublish first
    if reel.is_published:
        reel.is_published = False
        reel.save(update_fields=['is_published'])
        messages.info(request, f'ℹ️ Reel "{reel.title}" was unpublished before deletion.')
    
    title = reel.title
    reel.delete()
    
    messages.success(request, f'✅ Reel "{title}" deleted successfully!')
    return redirect('admin_reels')


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_generate_reel(request, reel_id):
    """Generate video from reel images"""
    from Hub.models import Reel
    from Hub.reel_generator_v2 import EnhancedReelGenerator
    
    reel = get_object_or_404(Reel, id=reel_id)
    
    # Check if already processing
    if reel.is_processing:
        messages.warning(request, 'Reel is already being processed. Please wait.')
        return redirect('admin_reels')
    
    # Check if images exist
    if not reel.images.exists():
        messages.error(request, 'Please add images to the reel before generating video.')
        return redirect('admin_edit_reel', reel_id=reel.id)
    
    try:
        logger.info(
            "Starting enhanced reel generation reel_id=%s title=%s images=%s animation=%s watermark=%s end_screen=%s",
            reel.id,
            reel.title,
            reel.images.count(),
            reel.transition_type,
            bool(reel.watermark_logo),
            bool(reel.add_end_screen),
        )
        
        # Generate reel with enhanced generator
        generator = EnhancedReelGenerator(reel)
        success = generator.generate_video()
        
        if success:
            logger.info(
                "Enhanced reel generated successfully reel_id=%s video_file=%s duration=%s",
                reel.id,
                reel.video_file.name if reel.video_file else None,
                reel.duration,
            )
            messages.success(request, f'✅ Professional reel "{reel.title}" generated successfully!')
        else:
            logger.warning("Enhanced reel generation returned False reel_id=%s", reel.id)
            messages.error(request, '❌ Failed to generate reel. Check server console for details.')
    
    except Exception as e:
        logger.exception("Exception during enhanced reel generation reel_id=%s", reel.id)
        
        messages.error(request, f'❌ Error: {str(e)}')
        reel.is_processing = False
        reel.save(update_fields=['is_processing'])
    
    return redirect('admin_reels')


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reels(request):
    """Reel list page"""
    from Hub.models import Reel
    
    reels = Reel.objects.all().prefetch_related('images').order_by('-created_at')
    
    # Calculate stats
    total_reels = reels.count()
    published_reels = reels.filter(is_published=True).count()
    draft_reels = reels.filter(is_published=False, is_processing=False).count()
    
    context = {
        'reels': reels,
        'total_reels': total_reels,
        'published_reels': published_reels,
        'draft_reels': draft_reels,
    }
    
    return render(request, 'admin_panel/reels.html', context)



@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_reel_details(request, reel_id):
    """Get reel details (AJAX)"""
    from Hub.models import Reel
    
    reel = get_object_or_404(Reel, id=reel_id)
    
    return JsonResponse({
        'id': reel.id,
        'title': reel.title,
        'description': reel.description,
        'video_url': reel.video_file.url if reel.video_file else '',
        'thumbnail_url': reel.thumbnail.url if reel.thumbnail else '',
        'duration': reel.duration,
        'is_published': reel.is_published,
        'created_at': reel.created_at.strftime('%Y-%m-%d %H:%M:%S')
    })


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_upload_reel_file(request):
    """Upload ready-made reel video file and link it to a product."""
    from Hub.models import Reel

    products = Product.objects.filter(is_active=True).only('id', 'name').order_by('name')

    if request.method == 'POST':
        title = (request.POST.get('title') or '').strip()
        description = (request.POST.get('description') or '').strip()
        product_id = (request.POST.get('product_id') or '').strip()
        duration_raw = (request.POST.get('duration') or '0').strip()
        is_published = request.POST.get('is_published') == 'on'
        order = int(request.POST.get('order', 0))
        video_file = request.FILES.get('video_file')
        thumbnail = request.FILES.get('thumbnail')

        if not title:
            messages.error(request, 'Reel title is required.')
            return redirect('admin_upload_reel_file')

        if not product_id:
            messages.error(request, 'Please select a linked product.')
            return redirect('admin_upload_reel_file')

        linked_product = Product.objects.filter(id=product_id, is_active=True).first()
        if not linked_product:
            messages.error(request, 'Selected product is invalid or inactive.')
            return redirect('admin_upload_reel_file')

        if not video_file:
            messages.error(request, 'Please upload a reel video file.')
            return redirect('admin_upload_reel_file')

        content_type = (getattr(video_file, 'content_type', '') or '').lower()
        if content_type and not content_type.startswith('video/'):
            messages.error(request, 'Invalid file type. Please upload a video file.')
            return redirect('admin_upload_reel_file')

        duration = 0
        try:
            duration = max(int(duration_raw), 0)
        except (TypeError, ValueError):
            duration = 0

        Reel.objects.create(
            title=title,
            description=description,
            product=linked_product,
            video_file=video_file,
            thumbnail=thumbnail,
            duration=duration,
            order=order,
            is_published=is_published,
            is_processing=False,
            created_by=request.user,
        )

        messages.success(request, f'Reel "{title}" uploaded successfully.')
        return redirect('admin_upload_reel_file')

    uploaded_reels = (
        Reel.objects
        .filter(video_file__isnull=False)
        .exclude(video_file='')
        .select_related('product')
        .order_by('-created_at')[:20]
    )

    return render(
        request,
        'admin_panel/upload_reel_file.html',
        {
            'products': products,
            'uploaded_reels': uploaded_reels,
        }
    )



def terms_and_conditions(request):
    """Display Terms and Conditions page"""
    from datetime import datetime
    context = {
        'current_date': datetime.now().strftime('%B %d, %Y'),
        'current_year': datetime.now().year,
    }
    return render(request, 'terms_and_conditions.html', context)


def privacy_policy(request):
    """Display Privacy Policy page"""
    from datetime import datetime
    context = {
        'current_date': datetime.now().strftime('%B %d, %Y'),
        'current_year': datetime.now().year,
    }
    return render(request, 'privacy_policy.html', context)


@login_required(login_url='accounts_login')
def address_book_view(request):
    """Address Book: list and manage saved addresses."""
    if request.method == 'POST':
        action = request.POST.get('action', 'add')

        if action == 'delete':
            address_id = request.POST.get('address_id', '')
            if address_id.isdigit():
                Address.objects.filter(id=int(address_id), user=request.user).delete()
                messages.success(request, 'Address removed.')
            return redirect('address_book')

        if action == 'set_default':
            address_id = request.POST.get('address_id', '')
            if address_id.isdigit():
                Address.objects.filter(user=request.user, is_default=True).update(is_default=False)
                Address.objects.filter(id=int(address_id), user=request.user).update(is_default=True)
                messages.success(request, 'Default address updated.')
            return redirect('address_book')

        full_name = request.POST.get('full_name', '').strip()
        mobile_number = request.POST.get('mobile_number', '').strip()
        address_line1 = request.POST.get('address_line1', '').strip()
        address_line2 = request.POST.get('address_line2', '').strip()
        city = request.POST.get('city', '').strip()
        state = request.POST.get('state', '').strip()
        pincode = request.POST.get('pincode', '').strip()
        country = request.POST.get('country', 'India').strip() or 'India'
        address_type = request.POST.get('address_type', 'HOME')
        is_default = request.POST.get('is_default') == 'on'

        if full_name and address_line1 and city and state and pincode:
            if is_default:
                Address.objects.filter(user=request.user, is_default=True).update(is_default=False)
            Address.objects.create(
                user=request.user,
                full_name=full_name,
                mobile_number=mobile_number,
                address_line1=address_line1,
                address_line2=address_line2,
                city=city,
                state=state,
                pincode=pincode,
                country=country,
                address_type=address_type,
                is_default=is_default,
            )
            messages.success(request, 'Address added successfully.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('address_book')

    addresses = Address.objects.filter(user=request.user)
    return render(request, 'address_book.html', {'addresses': addresses})


@login_required(login_url='accounts_login')
def payment_methods_view(request):
    """Payment Methods: show a summary of past payment methods used."""
    payment_methods_used = list(
        Order.objects.filter(user=request.user)
        .exclude(payment_method__isnull=True)
        .exclude(payment_method__exact='')
        .values_list('payment_method', flat=True)
        .distinct()
    )
    payment_meta = {
        'COD': {'label': 'Cash on Delivery', 'icon': 'fas fa-money-bill-wave', 'color': '#28a745'},
        'RAZORPAY': {'label': 'Razorpay (Cards / UPI / Net Banking)', 'icon': 'fas fa-credit-card', 'color': '#2196f3'},
        'UPI': {'label': 'UPI', 'icon': 'fas fa-qrcode', 'color': '#9c27b0'},
        'ONLINE': {'label': 'Online Payment', 'icon': 'fas fa-globe', 'color': '#ff9800'},
        'CARD': {'label': 'Credit / Debit Card', 'icon': 'fas fa-credit-card', 'color': '#e91e63'},
    }
    methods = [
        {**payment_meta.get(method, {'label': method, 'icon': 'fas fa-circle', 'color': '#6c757d'}), 'key': method}
        for method in payment_methods_used
    ]
    return render(request, 'payment_methods.html', {'methods': methods})


# ============================================
# COUPON SYSTEM VIEWS
# ============================================

@require_POST
def validate_coupon(request):
    """Validate coupon code and return discount details."""
    try:
        data = json.loads(request.body)
        code = data.get('code', '').upper().strip()
        cart_total = float(data.get('cart_total', 0))
        
        if not code:
            return JsonResponse({'valid': False, 'message': 'Please enter a coupon code'})
        
        # Get coupon
        try:
            coupon = Coupon.objects.get(code=code)
        except Coupon.DoesNotExist:
            return JsonResponse({'valid': False, 'message': 'Invalid coupon code'})
        
        # Check if valid
        if not coupon.is_valid():
            return JsonResponse({'valid': False, 'message': 'This coupon has expired'})
        
        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({'valid': False, 'message': 'Please login to use coupons'})
        
        # Check if already used by this user
        already_used = CouponUsage.objects.filter(
            coupon=coupon, 
            user=request.user
        ).exists()
        
        if already_used:
            return JsonResponse({'valid': False, 'message': 'Coupon already used'})
        
        # Check first order coupon
        if coupon.coupon_type == 'FIRST_ORDER':
            has_orders = Order.objects.filter(user=request.user, payment_status='PAID').exists()
            if has_orders:
                return JsonResponse({'valid': False, 'message': 'This coupon is only for first-time orders'})
        
        # Check minimum purchase
        if cart_total < coupon.min_purchase_amount:
            return JsonResponse({
                'valid': False, 
                'message': f'Minimum purchase of ₹{coupon.min_purchase_amount} required'
            })
        
        # Check usage limit
        if coupon.usage_limit:
            total_uses = CouponUsage.objects.filter(coupon=coupon).count()
            if total_uses >= coupon.usage_limit:
                return JsonResponse({'valid': False, 'message': 'Coupon usage limit reached'})
        
        # Calculate discount
        discount_amount = coupon.get_discount_amount(cart_total)
        
        return JsonResponse({
            'valid': True,
            'coupon_id': coupon.id,
            'code': coupon.code,
            'discount_amount': float(discount_amount),
            'discount_type': coupon.discount_type,
            'discount_value': float(coupon.discount_value),
            'message': f'Coupon applied! You saved ₹{discount_amount:.2f}'
        })
        
    except Exception as e:
        return JsonResponse({'valid': False, 'message': f'Error validating coupon: {str(e)}'})


@require_POST
def get_available_coupons(request):
    """Get list of available coupons for the user."""
    try:
        if not request.user.is_authenticated:
            return JsonResponse({'coupons': []})
        
        user = request.user
        data = json.loads(request.body)
        cart_total = float(data.get('cart_total', 0))
        
        available_coupons = []
        
        # 1. First Order Coupon (15%)
        has_paid_orders = Order.objects.filter(user=user, payment_status='PAID').exists()
        if not has_paid_orders:
            first_coupon = Coupon.objects.filter(
                coupon_type='FIRST_ORDER',
                is_active=True
            ).first()
            
            if first_coupon and first_coupon.is_valid():
                used = CouponUsage.objects.filter(coupon=first_coupon, user=user).exists()
                discount_amount = first_coupon.get_discount_amount(cart_total)
                
                available_coupons.append({
                    'code': first_coupon.code,
                    'title': 'First Order Discount',
                    'description': first_coupon.description,
                    'discount': f"{first_coupon.discount_value}% OFF",
                    'discount_amount': float(discount_amount),
                    'min_purchase': float(first_coupon.min_purchase_amount),
                    'used': used,
                    'type': 'FIRST_ORDER'
                })
        
        # 2. Spend 5K Coupon (15%)
        tracker, _ = UserSpendTracker.objects.get_or_create(user=user)
        
        if tracker.can_earn_5k_coupon():
            spend_coupon_code = f"SPEND5K{user.id}"
            
            # Create or get the coupon
            spend_coupon, created = Coupon.objects.get_or_create(
                code=spend_coupon_code,
                defaults={
                    'coupon_type': 'SPEND_5K',
                    'description': f'Congratulations! You spent ₹5000. Enjoy 5% off!',
                    'discount_type': 'PERCENTAGE',
                    'discount_value': 5,
                    'min_purchase_amount': 0,
                    'max_discount_amount': 250,  # Max ₹250 discount
                    'usage_per_user': 1,
                    'valid_from': timezone.now(),
                    'valid_to': timezone.now() + timedelta(days=30),
                    'is_active': True
                }
            )
            
            used = CouponUsage.objects.filter(coupon=spend_coupon, user=user).exists()
            discount_amount = spend_coupon.get_discount_amount(cart_total)
            
            available_coupons.append({
                'code': spend_coupon.code,
                'title': 'Spend ₹5000 Reward',
                'description': spend_coupon.description,
                'discount': '5% OFF',
                'discount_amount': float(discount_amount),
                'min_purchase': 0,
                'used': used,
                'type': 'SPEND_5K',
                'spent_amount': float(tracker.current_cycle_spent)
            })
        
        return JsonResponse({
            'success': True,
            'coupons': available_coupons,
            'total_spent': float(tracker.total_spent),
            'cycle_spent': float(tracker.current_cycle_spent)
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'coupons': [], 'error': str(e)})

# ===== MAIN PAGE BANNER MANAGEMENT VIEWS =====

@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_main_page_banners(request):
    """Admin List Main Page Banners"""
    banners = MainPageBanner.objects.all().order_by('banner_area', 'order', 'id')
    
    # Get preview banners (first 3 active banners for promotion area, 2 for marketing area)
    first_preview = MainPageBanner.objects.filter(banner_area='first', is_active=True).order_by('order', 'id')[:3]
    second_preview = MainPageBanner.objects.filter(banner_area='second', is_active=True).order_by('order', 'id')[:2]
    
    context = {
        'banners': banners,
        'first_preview': first_preview,
        'second_preview': second_preview,
    }
    return render(request, 'admin_panel/main_page_banners.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_add_main_page_banner(request):
    """Admin Add Main Page Banner"""
    if request.method == 'POST':
        try:
            banner_area = request.POST.get('banner_area', 'first')
            link_url = request.POST.get('link_url', '#').strip() or '#'
            order = request.POST.get('order', '0')
            is_active = request.POST.get('is_active') == 'on'
            image = request.FILES.get('image')
            
            if not image:
                messages.error(request, 'Banner image is required.')
                return render(request, 'admin_panel/add_main_page_banner.html')
            
            try:
                order = int(order or 0)
            except (TypeError, ValueError):
                order = 0
            
            # For promotion area (first), only image and link are needed
            if banner_area == 'first':
                badge_text = ''
                title = f'Banner {order}' if order else 'Banner'
                description = ''
            else:
                # For marketing area (second), title is required
                title = request.POST.get('title', '').strip()
                badge_text = request.POST.get('badge_text', '').strip()
                description = request.POST.get('description', '').strip()
                
                if not title:
                    messages.error(request, 'Banner title is required for Marketing area 2 card.')
                    return render(request, 'admin_panel/add_main_page_banner.html')
            
            MainPageBanner.objects.create(
                banner_area=banner_area,
                badge_text=badge_text,
                title=title,
                description=description,
                link_url=link_url,
                order=order,
                is_active=is_active,
                image=image
            )
            
            messages.success(request, f'Banner "{title}" added successfully!')
            return redirect('admin_main_page_banners')
            
        except Exception as e:
            messages.error(request, f'Error adding banner: {str(e)}')
    
    return render(request, 'admin_panel/add_main_page_banner.html')


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_edit_main_page_banner(request, banner_id):
    """Admin Edit Main Page Banner"""
    banner = get_object_or_404(MainPageBanner, id=banner_id)
    
    if request.method == 'POST':
        try:
            banner.banner_area = request.POST.get('banner_area', banner.banner_area)
            banner.link_url = request.POST.get('link_url', '#').strip() or '#'
            banner.order = int(request.POST.get('order', banner.order) or 0)
            banner.is_active = request.POST.get('is_active') == 'on'
            
            # For promotion area (first), only image and link are needed
            if banner.banner_area == 'first':
                banner.badge_text = ''
                banner.title = f'Banner {banner.order}' if banner.order else 'Banner'
                banner.description = ''
            else:
                # For marketing area (second), update all fields
                banner.badge_text = request.POST.get('badge_text', '').strip()
                title = request.POST.get('title', '').strip()
                if not title:
                    messages.error(request, 'Banner title is required for Marketing area 2 card.')
                    return render(request, 'admin_panel/edit_main_page_banner.html', {'banner': banner})
                banner.title = title
                banner.description = request.POST.get('description', '').strip()
            
            if 'image' in request.FILES:
                # Delete old image if exists
                if banner.image:
                    banner.image.delete()
                banner.image = request.FILES['image']
            
            banner.save()
            messages.success(request, f'Banner updated successfully!')
            return redirect('admin_main_page_banners')
            
        except Exception as e:
            messages.error(request, f'Error updating banner: {str(e)}')
    
    context = {
        'banner': banner,
    }
    return render(request, 'admin_panel/edit_main_page_banner.html', context)


@login_required(login_url='login')
@staff_member_required(login_url='login')
def admin_delete_main_page_banner(request, banner_id):
    """Admin Delete Main Page Banner"""
    banner = get_object_or_404(MainPageBanner, id=banner_id)
    
    if request.method == 'POST':
        banner_title = banner.title
        if banner.image:
            banner.image.delete()
        banner.delete()
        messages.success(request, f'Banner "{banner_title}" deleted successfully!')
    
    return redirect('admin_main_page_banners')


# ===== SLIDER MANAGEMENT VIEWS =====
